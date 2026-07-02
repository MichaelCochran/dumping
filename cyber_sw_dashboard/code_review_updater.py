#!/usr/bin/env python3
"""
Code Review Updater
Downloads Excel from source SharePoint, processes data, uploads to destination SharePoint
"""

import sys
import tempfile
import os
import shutil
import time
import re
import webbrowser
import warnings
import json
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, date, timedelta, timezone

# Suppress openpyxl warnings about unsupported Excel features
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# ============================================================================
# CONFIGURATION
# ============================================================================

# Source SharePoint configuration - where we download the original review status file
SOURCE_CONFIG = {
    "site_url": "https://space-us.external.lmco.com/sites/ngi",
    "file_path": "/fa/Software/Software/01-Common-Folder/NGI FSW Formal Review Status.xlsx",
    "output_path": str(Path(__file__).parent / "NGI FSW Formal Review Status.xlsx"),
    "session_file": str(Path(__file__).parent / "sp_session.json"),
}

# Read from production file for dashboard display
# This is the file that the dashboard reads for display day-to-day.
DEST_READ_CONFIG = {
    "site_url": "https://lmco.sharepoint.us/sites/US-NGI-Cyber-Software",
    "file_path": "/Shared Documents/General/Software Support/New Code Review Tracker.xlsx",
    "session_file": str(Path(__file__).parent / "sp_dest_session.json"),
}

# Upload to staging file (staging area before manual update to production)
# Use this "staging" workbook as a sandbox before copying results into production.
DEST_UPLOAD_CONFIG = {
    "site_url": "https://lmco.sharepoint.us/sites/US-NGI-Cyber-Software",
    "file_path": "/Shared Documents/General/Software Support/Staging Tracker.xlsx",
    "session_file": str(Path(__file__).parent / "sp_dest_session.json"),
}

# Target sheets in the destination Excel file for categorizing reviews
# These sheet names must match the tabs in the destination tracker workbook.
TARGET_SHEETS = ["Code"]

# Date cutoff for considering rows new (processed date column is a date-only string)
# Rows with dates BEFORE this cutoff are ignored when generating new entries.
DATE_CUTOFF = "2026-05-01"

# Maximum age for saved SharePoint session files before forcing re-authentication.
SESSION_MAX_AGE = timedelta(hours=5)

# Track whether dashboard-mode login confirmation has already been used in this
# process. We only want the dashboard flag-based flow for the first interactive
# login; subsequent authentications in the same run fall back to the original
# console behavior (which usually won't be needed if the first login succeeds).
DASHBOARD_LOGIN_USED = False

class TermColors:
    INFO = '\033[96m'     # bright cyan
    ERROR = '\033[91m'    # bright red
    WARNING = '\033[33m'  # yellow
    DEBUG = '\033[94m'    # bright blue
    SECTION = '\033[96m'  # bright cyan
    SUCCESS = '\033[92m'  # bright green
    PROMPT = '\033[95m'   # bright magenta
    RESET = '\033[0m'


def cprint(kind: str, message: str) -> None:
    colors = TermColors
    color = colors.INFO  # default

    if kind == "ERROR":
        color = colors.ERROR
    elif kind == "WARNING":
        color = colors.WARNING
    elif kind == "DEBUG":
        color = colors.DEBUG
    elif kind == "SECTION":
        color = colors.SECTION
    elif kind == "SUCCESS":
        color = colors.SUCCESS
    elif kind == "PROMPT":
        color = colors.PROMPT

    text = f"{color}{message}{colors.RESET}"
    try:
        print(text)
    except UnicodeEncodeError:
        # Fallback for terminals that cannot encode certain Unicode characters.
        safe = message.encode("ascii", "replace").decode("ascii")
        try:
            print(safe)
        except Exception:
            # As a last resort, ignore logging rather than crash the script.
            pass

# ============================================================================
# SHAREPOINT AUTHENTICATION & DOWNLOAD
# ============================================================================

def is_login_page(url: str) -> bool:
    """Return True if the URL is a Microsoft login/auth page or corporate SSO page.
    
    Used to detect if authentication has expired and user needs to re-login.
    
    Args:
        url: The current page URL
        
    Returns:
        bool: True if URL is an authentication or identity provider page
    """
    auth_domains = [
        "login.microsoftonline.com",
        "login.live.com",
        "login.microsoft.com",
        "aadcdn.msftauth.net",
        "auth.p.external.lmco.com",  # Lockheed Martin auth
        "idp",  # Generic identity provider in path
        "sso",  # Single sign-on pages
        "saml",  # SAML authentication
    ]
    return any(d in url.lower() for d in auth_domains)

def check_session_valid(playwright, session_path: Path, site_url: str) -> bool:
    """Check if saved session is still valid.
    
    Attempts to navigate to the site using saved session cookies.
    If redirected to login page, session has expired.
    
    Args:
        playwright: Playwright instance
        session_path: Path to saved session file
        site_url: SharePoint site URL to test
        
    Returns:
        bool: True if session is valid, False otherwise
    """
    if not session_path.exists():
        cprint("INFO", f"No saved session found at: {session_path}")
        return False
    
    cprint("INFO", "Checking saved session...")
    browser = playwright.chromium.launch(headless=True)
    try:
        context = browser.new_context(storage_state=str(session_path))
        page = context.new_page()
        # Use a slightly lower timeout to avoid long waits on slow responses
        page.goto(site_url, wait_until="domcontentloaded", timeout=8000)
        valid = not is_login_page(page.url)
        if valid:
            cprint("SUCCESS", "Session is valid.")
        else:
            cprint("WARNING", "Session expired - re-authentication required.")
        return valid
    except Exception as e:
        error_type = type(e).__name__
        if "TimeoutError" in error_type:
            cprint("WARNING", "Session check timed out - network may be slow or unavailable")
        else:
            cprint("WARNING", f"Session check failed ({error_type}): {e}")
        cprint("WARNING", "  Will attempt fresh authentication")
        return False
    finally:
        browser.close()

def authenticate(playwright, session_path: Path, site_url: str) -> bool:
    """Perform interactive login and save session.
    
    Opens a browser window for user to complete Microsoft authentication.
    Session is saved for future use to avoid repeated logins.
    The goal is for the user to only need to re-authenticate when cookies
    expire or when the browser profile changes.
    
    Args:
        playwright: Playwright instance
        session_path: Path where session will be saved
        site_url: SharePoint site URL to authenticate to
        
    Returns:
        bool: True if authentication successful, False otherwise
    """
    colors = TermColors
    sep_line = "=" * 60
    print(f"\n{colors.SECTION}{sep_line}{colors.RESET}")
    print(f"{colors.PROMPT} ACTION REQUIRED: Browser opening for SharePoint login.{colors.RESET}")
    print(f"{colors.PROMPT} Complete the FULL login process:{colors.RESET}")
    print(f"{colors.PROMPT}   1. Choose your identity provider{colors.RESET}")
    print(f"{colors.PROMPT}   2. Enter credentials and complete MFA{colors.RESET}")
    print(f"{colors.PROMPT}   3. WAIT for SharePoint site to load completely{colors.RESET}")
    print(f"{colors.PROMPT}   4. Verify you see SharePoint (NOT a login/auth page){colors.RESET}")
    print(f"{colors.PROMPT}   5. Press ENTER here ONLY after SharePoint loads{colors.RESET}")
    print(f"{colors.SECTION}{sep_line}{colors.RESET}\n")

    # Keep track of the browser so we can close it cleanly on error/interrupt.
    browser = None
    try:
        browser = playwright.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()

        try:
            page.goto(site_url, timeout=15000)
        except Exception as e:
            cprint("WARNING", f"Initial navigation had an issue: {e}")
            cprint("INFO", "Browser is open - please navigate to the SharePoint site manually")

        # When running under the dashboard, we coordinate login using two
        # small files in the script directory:
        # - dashboard_mode_login_needed.flag : created by the updater to
        #   indicate that a login dialog should be shown.
        # - dashboard_mode_continue.flag     : created by the dashboard when
        #   the user clicks "Login Complete".
        #
        # In dashboard mode we always use this flag-based flow so that the
        # user never needs to interact with the terminal.
        base_dir = Path(__file__).parent
        dashboard_login_needed = base_dir / "dashboard_mode_login_needed.flag"
        dashboard_flag = base_dir / "dashboard_mode_continue.flag"
        use_dashboard_mode = os.environ.get("CODE_REVIEW_DASHBOARD_MODE") == "1"

        if use_dashboard_mode:
            cprint("INFO", "Signaling dashboard that login is required...")
            # Ensure any stale flags from a previous run do not immediately
            # trigger continuation.
            try:
                if dashboard_flag.exists():
                    dashboard_flag.unlink()
                if dashboard_login_needed.exists():
                    dashboard_login_needed.unlink()
            except Exception:
                pass

            # Create the "login needed" marker so the dashboard can show or
            # re-use its login dialog for this authentication step.
            try:
                dashboard_login_needed.write_text("login", encoding="utf-8")
            except Exception:
                # If we cannot signal the dashboard, fall back to console
                # behavior to avoid hanging indefinitely.
                use_dashboard_mode = False

        if use_dashboard_mode:
            cprint("INFO", "Waiting for dashboard confirmation of successful login...")
            
            # Poll for the flag file up to a reasonable timeout. The dashboard
            # will create this file when the user clicks a 'Continue' button
            # after finishing login in the browser.
            start = time.time()
            timeout_seconds = 15 * 60  # 15 minutes
            while True:
                if dashboard_flag.exists():
                    break
                if time.time() - start > timeout_seconds:
                    cprint("ERROR", "Timed out waiting for dashboard login confirmation.")
                    if browser:
                        browser.close()
                    return False
                time.sleep(1)

        else:
            try:
                input("Press ENTER after successful login...")
                cprint("INFO", "User confirmed login.")
            except (KeyboardInterrupt, EOFError):
                cprint("ERROR", "Login cancelled by user.")
                if browser:
                    browser.close()
                return False

        try:
            # Verify we're not still on login page
            current_url = page.url
            if is_login_page(current_url):
                cprint("ERROR", "❌ Login incomplete - you're still on an authentication page!")
                cprint("ERROR", f"  Current URL: {current_url}")
                cprint("ERROR", "")
                cprint("ERROR", "  → Please wait for the FULL login process to complete:")
                cprint("ERROR", "     1. Choose your identity provider")
                cprint("ERROR", "     2. Enter credentials and complete MFA")
                cprint("ERROR", "     3. Wait for SharePoint to load completely")
                cprint("ERROR", "     4. Verify you see the SharePoint site (NOT a login/auth page)")
                cprint("ERROR", "")
                cprint("ERROR", "  → Then try running the script again.")
                if browser:
                    browser.close()
                return False
            
            # Verify we actually reached a SharePoint domain
            expected_domains = ["sharepoint", "external.lmco.com"]
            if not any(domain in current_url.lower() for domain in expected_domains):
                cprint("WARNING", f"⚠️  Current page may not be the SharePoint site:")
                cprint("WARNING", f"  URL: {current_url}")
                cprint("WARNING", "  Expected a URL containing 'sharepoint' or 'external.lmco.com'")
                cprint("WARNING", "  Attempting to continue anyway...")
            
            context.storage_state(path=str(session_path))
            cprint("SUCCESS", f"✓ Session saved successfully!")
            cprint("INFO", f"  Session file: {session_path}")
        except Exception as e:
            cprint("ERROR", f"Failed to save session: {e}")
            cprint("ERROR", f"  This may indicate the login was not successful")
            if browser:
                browser.close()
            return False

        browser.close()
        return True
    
    except Exception as e:
        if browser:
            browser.close()
        return False

def _ensure_valid_session(playwright, session_path: Path, site_url: str) -> None:
    """Ensure there is a valid SharePoint session, re-authenticating if needed.

    Uses `check_session_valid` to test an existing session file and falls back to
    interactive `authenticate` if the session is missing or expired. Exits the
    program if authentication ultimately fails.
    """
    session_ok = check_session_valid(playwright, session_path, site_url)
    if not session_ok:
        success = authenticate(playwright, session_path, site_url)
        if not success:
            cprint("ERROR", "Authentication failed - exiting.")
            sys.exit(1)


def _session_is_too_old(path: Path) -> bool:
    """Return True if the session file exists and is older than SESSION_MAX_AGE."""
    if not path.exists():
        return False

    try:
        mtime = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc)
    except Exception:
        return False

    age = datetime.now(timezone.utc) - mtime
    return age > SESSION_MAX_AGE


def _maybe_expire_sessions_by_age() -> None:
    """Automatically delete session files that are older than SESSION_MAX_AGE."""
    # Use dict to track unique session files (DEST_READ and DEST_UPLOAD share same file)
    session_files = {
        "SOURCE (external SharePoint)": Path(SOURCE_CONFIG["session_file"]),
        "DEST (internal SharePoint)": Path(DEST_READ_CONFIG["session_file"]),
    }

    for label, path in session_files.items():
        try:
            if _session_is_too_old(path) and path.exists():
                path.unlink()
                cprint("INFO", f"Session file too old; cleared {label} session: {path}")
        except Exception as e:
            cprint("WARNING", f"Failed to auto-expire {label} session file {path}: {e}")

def get_file_metadata(playwright, session_path: Path, site_url: str, file_path: str) -> dict:
    """Get SharePoint file metadata (ETag, Last-Modified) without downloading.
    
    Args:
        playwright: Playwright instance
        session_path: Path to saved session file
        site_url: SharePoint site URL
        file_path: Relative path to file on SharePoint
        
    Returns:
        dict: Metadata with 'etag', 'last_modified', 'size', or empty dict on error
    """
    metadata_url = f"{site_url.rstrip('/')}{file_path}"
    
    browser = playwright.chromium.launch(headless=True)
    try:
        context = browser.new_context(storage_state=str(session_path))
        page = context.new_page()
        
        metadata = {}
        
        def handle_response(response):
            if response.url == metadata_url or file_path in response.url:
                headers = response.headers
                metadata['etag'] = headers.get('etag', '')
                metadata['last_modified'] = headers.get('last-modified', '')
                metadata['size'] = headers.get('content-length', '0')
        
        page.on("response", handle_response)
        
        try:
            # Use HEAD request would be better but SharePoint doesn't always support it
            # So we'll start a GET and cancel it
            page.goto(metadata_url, wait_until="commit", timeout=10000)
        except:
            pass  # We just need the response headers
        
        return metadata
    except Exception as e:
        cprint("WARNING", f"Could not get file metadata: {e}")
        return {}
    finally:
        browser.close()


def should_download_file(playwright, session_path: Path, site_url: str, 
                        file_path: str, local_path: str) -> bool:
    """Check if file needs to be downloaded by comparing metadata.
    
    Args:
        playwright: Playwright instance
        session_path: Path to saved session file
        site_url: SharePoint site URL
        file_path: Relative path to file on SharePoint
        local_path: Local file path to compare against
        
    Returns:
        bool: True if file should be downloaded, False if local file is current
    """
    local_file = Path(local_path)
    
    # Always download if local file doesn't exist
    if not local_file.exists():
        return True
    
    # Get remote file metadata
    remote_metadata = get_file_metadata(playwright, session_path, site_url, file_path)
    if not remote_metadata:
        # Can't determine, download to be safe
        return True
    
    # Check if we have cached metadata for this file
    cache_path = local_file.with_suffix(local_file.suffix + '.meta')
    if cache_path.exists():
        try:
            with open(cache_path, 'r') as f:
                cached_metadata = json.load(f)
            
            # Compare ETags or Last-Modified
            if remote_metadata.get('etag') and cached_metadata.get('etag'):
                if remote_metadata['etag'] == cached_metadata['etag']:
                    cprint("INFO", f"  File unchanged (ETag match), skipping download")
                    return False
            
            if remote_metadata.get('last_modified') and cached_metadata.get('last_modified'):
                if remote_metadata['last_modified'] == cached_metadata['last_modified']:
                    cprint("INFO", f"  File unchanged (timestamp match), skipping download")
                    return False
        except Exception as e:
            cprint("DEBUG", f"  Could not read cached metadata: {e}")
    
    # File needs to be downloaded
    return True


def save_file_metadata(local_path: str, metadata: dict) -> None:
    """Save file metadata to cache for future comparisons.
    
    Args:
        local_path: Local file path
        metadata: Metadata dict to save
    """
    cache_path = Path(local_path).with_suffix(Path(local_path).suffix + '.meta')
    try:
        with open(cache_path, 'w') as f:
            json.dump(metadata, f, indent=2)
    except Exception as e:
        cprint("DEBUG", f"Could not save metadata cache: {e}")


def download_file_from_sharepoint(playwright, session_path: Path, site_url: str, 
                                  file_path: str, output_path: str) -> bool:
    """Download file from SharePoint using Playwright.
    
    Uses the ?download=1 parameter to trigger direct file download.
    
    Args:
        playwright: Playwright instance
        session_path: Path to saved session file
        site_url: SharePoint site URL
        file_path: Relative path to file on SharePoint
        output_path: Local path where file will be saved
        
    Returns:
        bool: True if download successful, False otherwise
    """
    download_url = f"{site_url.rstrip('/')}{file_path}?download=1"
    
    browser = playwright.chromium.launch(headless=True)
    page = None
    try:
        context = browser.new_context(storage_state=str(session_path), accept_downloads=True)
        page = context.new_page()
        
        download_info = {"download": None, "error": None}
        response_info = {"status": None, "url": None}
        
        def handle_download(download):
            try:
                download_info["download"] = download
            except Exception as e:
                download_info["error"] = str(e)
        
        def handle_response(response):
            response_info["status"] = response.status
            response_info["url"] = response.url
        
        page.on("download", handle_download)
        page.on("response", handle_response)
        
        navigation_error = None
        try:
            # Reduced timeout to avoid long stalls when SharePoint is slow/unresponsive
            page.goto(download_url, wait_until="commit", timeout=15000)
        except Exception as e:
            navigation_error = e
            error_str = str(e)
            if "Download is starting" in error_str or "download" in error_str.lower():
                cprint("INFO", "Navigation interrupted by download (expected)")
            elif "TimeoutError" in type(e).__name__ or "timeout" in error_str.lower():
                cprint("WARNING", "Navigation timed out - will check if download started")
            else:
                cprint("WARNING", f"Navigation error: {error_str}")
        
        # Shorter fixed wait; download events should already be triggered by now
        page.wait_for_timeout(1000)
        
        if download_info["error"]:
            return False
        
        if not download_info["download"]:
            cprint("ERROR", "Download did not start")
            cprint("ERROR", f"  URL attempted: {download_url}")
            
            if navigation_error:
                cprint("ERROR", f"  Navigation error occurred: {str(navigation_error)}")
            
            if response_info["status"]:
                if response_info["status"] == 404:
                    cprint("ERROR", "  File not found - check if the path is correct")
                elif response_info["status"] == 403:
                    cprint("ERROR", "  Access denied - check permissions")
                elif response_info["status"] >= 500:
                    cprint("ERROR", "  Server error - SharePoint may be experiencing issues")
            
            try:
                page_title = page.title()
                page_url = page.url
                cprint("ERROR", f"  Current page: {page_url}")
                cprint("ERROR", f"  Page title: {page_title}")
                
                if "sign" in page_title.lower() or "login" in page_title.lower():
                    cprint("ERROR", "  Redirected to login page - session may have expired")
                elif "error" in page_title.lower():
                    cprint("ERROR", "  Page shows an error - file may not exist or access denied")
            except:
                pass
            
            return False
        
        download = download_info["download"]
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            temp_path = tmp_file.name
        
        try:
            download.save_as(temp_path)
        except Exception as e:
            cprint("ERROR", f"Failed to save download: {str(e)}")
            cprint("ERROR", "  This may indicate disk space issues or permission problems")
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return False
        
        if not os.path.exists(temp_path):
            cprint("ERROR", f"Download file was not created at: {temp_path}")
            return False
        
        temp_size = os.path.getsize(temp_path)
        if temp_size == 0:
            cprint("ERROR", "Downloaded file is empty (0 bytes)")
            cprint("ERROR", "  This usually means the download failed or returned an error page")
            os.remove(temp_path)
            return False
        elif temp_size < 1024:
            cprint("WARNING", f"Downloaded file is very small ({temp_size} bytes)")
            cprint("WARNING", "  This may indicate an error page instead of the actual file")
        
        normalized_path = os.path.normpath(output_path)
        
        if os.path.exists(normalized_path):
            os.remove(normalized_path)
        
        try:
            with open(temp_path, 'rb') as src:
                file_data = src.read()
                with open(normalized_path, 'wb') as dst:
                    dst.write(file_data)
        except Exception as e:
            cprint("ERROR", f"Failed to copy file to destination: {str(e)}")
            cprint("ERROR", f"  Source: {temp_path}")
            cprint("ERROR", f"  Destination: {normalized_path}")
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return False
        
        os.remove(temp_path)
        
        size_kb = os.path.getsize(normalized_path) / 1024
        if size_kb < 10:
            cprint("WARNING", "File size is unusually small for an Excel file")
            cprint("WARNING", "  Verify the file content is correct")
        
        # Try to save metadata for future change detection
        try:
            metadata = get_file_metadata(playwright, session_path, site_url, file_path)
            if metadata:
                save_file_metadata(output_path, metadata)
        except:
            pass  # Not critical if metadata save fails
        
        return True
        
    except Exception as e:
        error_type = type(e).__name__
        cprint("ERROR", f"Unexpected error during download ({error_type}): {str(e)}")
        cprint("ERROR", f"  URL: {download_url}")
        return False
    finally:
        if browser:
            browser.close()

def upload_file_to_sharepoint(playwright, session_path: Path, site_url: str, 
                               file_path: str, local_file: str) -> bool:
    """Upload file to SharePoint using REST API.
    
    Uses SharePoint REST API with form digest for authentication.
    Overwrites existing file at the destination path.
    
    Args:
        playwright: Playwright instance
        session_path: Path to saved session file
        site_url: SharePoint site URL
        file_path: Relative path where file will be uploaded on SharePoint
        local_file: Local path of file to upload
        
    Returns:
        bool: True if upload successful, False otherwise
    """
    if not os.path.exists(local_file):
        cprint("ERROR", f"Local file not found: {local_file}")
        return False
    
    try:
        with open(local_file, 'rb') as f:
            file_content = f.read()
    except Exception as e:
        cprint("ERROR", f"Failed to read local file: {e}")
        cprint("ERROR", f"  File: {local_file}")
        return False
    
    file_size = len(file_content)
    if file_size == 0:
        cprint("ERROR", "Cannot upload empty file")
        return False
    
    file_name = os.path.basename(file_path)
    folder_path = os.path.dirname(file_path)
    
    from urllib.parse import urlparse
    parsed_url = urlparse(site_url)
    site_relative_path = parsed_url.path.rstrip('/')
    
    full_folder_path = f"{site_relative_path}{folder_path}" if site_relative_path else folder_path
    
    upload_endpoint = f"{site_url}/_api/web/GetFolderByServerRelativeUrl('{full_folder_path}')/Files/add(url='{file_name}',overwrite=true)"
    
    cprint("INFO", f"Upload endpoint: {upload_endpoint}")
    
    browser = None
    try:
        browser = playwright.chromium.launch(headless=True)
        context = browser.new_context(storage_state=str(session_path))
        page = context.new_page()
        
        digest_url = f"{site_url}/_api/contextinfo"
        
        try:
            response = page.request.post(
                digest_url,
                headers={
                    "Accept": "application/json;odata=verbose",
                    "Content-Type": "application/json;odata=verbose",
                }
            )
        except Exception as e:
            cprint("ERROR", f"Failed to request form digest: {e}")
            cprint("ERROR", f"  Digest URL: {digest_url}")
            cprint("ERROR", "  This usually indicates session expiration or network issues")
            return False
        
        if response.status != 200:
            cprint("ERROR", "Failed to get request digest")
            cprint("ERROR", f"  HTTP Status: {response.status}")
            cprint("ERROR", f"  URL: {digest_url}")
            
            if response.status == 401:
                cprint("ERROR", "  Unauthorized - session may have expired")
            elif response.status == 403:
                cprint("ERROR", "  Forbidden - insufficient permissions")
            elif response.status == 404:
                cprint("ERROR", "  Not found - check site URL")
            
            return False
        
        try:
            digest_data = response.json()
            form_digest = digest_data["d"]["GetContextWebInformation"]["FormDigestValue"]
        except Exception as e:
            cprint("ERROR", f"Failed to parse digest response: {e}")
            cprint("ERROR", "  This indicates an unexpected response format")
            return False
        
        try:
            response = page.request.post(
                upload_endpoint,
                headers={
                    "Accept": "application/json;odata=verbose",
                    "X-RequestDigest": form_digest,
                    "Content-Type": "application/octet-stream",
                },
                data=file_content
            )
        except Exception as e:
            cprint("ERROR", f"Failed to upload file: {e}")
            cprint("ERROR", f"  Upload endpoint: {upload_endpoint}")
            cprint("ERROR", "  This may indicate network timeout or connection issues")
            return False
        
        if response.status in (200, 201):
            cprint("SUCCESS", "File uploaded successfully!")
            return True
        else:
            cprint("ERROR", "Upload failed")
            cprint("ERROR", f"  HTTP Status: {response.status}")
            cprint("ERROR", f"  Upload endpoint: {upload_endpoint}")
            
            if response.status == 401:
                cprint("ERROR", "  Unauthorized - form digest may have expired")
            elif response.status == 403:
                cprint("ERROR", "  Forbidden - check write permissions to the folder")
            elif response.status == 404:
                cprint("ERROR", "  Folder not found - check the folder path")
            elif response.status == 409:
                cprint("ERROR", "  Conflict - file may be locked or checked out by another user")
            elif response.status == 423:
                cprint("ERROR", "\n  File is LOCKED - someone has it open in Excel or SharePoint")
                cprint("ERROR", "  → ACTION REQUIRED:")
                cprint("ERROR", "     1. Close the file if you have it open")
                cprint("ERROR", "     2. Ask others to close the file")
                cprint("ERROR", "     3. Check if the file is checked out in SharePoint")
                cprint("ERROR", "     4. Wait a moment and try running the script again")
            elif response.status >= 500:
                cprint("ERROR", "  Server error - SharePoint may be experiencing issues")
            return False
            
    except Exception as e:
        error_type = type(e).__name__
        cprint("ERROR", f"Unexpected upload error ({error_type}): {str(e)}")
        cprint("ERROR", f"  Target: {upload_endpoint}")
        return False
    finally:
        if browser:
            browser.close()

# ============================================================================
# DATA PROCESSING & TRANSFORMATION
# ============================================================================

def map_column_d(col_c_value):
    """Map column C value to repository name.
    
    Extracts repository identifier from description field based on keywords.
    Supports multiple variants: MSV, KV, COMMON, NGICON, HSP, HS.
    
    Args:
        col_c_value: Value from column C (description field)
        
    Returns:
        str: Standardized repository name or 'ENTER REPO NAME' if no match
    """
    val = str(col_c_value or "").strip()
    val_upper = val.upper()
    
    if "MSV" in val_upper and "EM" in val_upper:
        return "MSV-EM"
    if "MSV" in val_upper and "FSW" in val_upper:
        return "MSV-FSW"
    if "MSV" in val_upper and "NGC" in val_upper:
        return "MSV-NGC"
    if "MSV" in val_upper and "PTS" in val_upper:
        return "MSV-PTS"
    
    if "KV" in val_upper and "EM" in val_upper:
        return "KV-EM"
    if "KV" in val_upper and "FSW" in val_upper:
        return "KV-FSW"
    if "KV" in val_upper and "NGC" in val_upper:
        return "KV-NGC"
    if "KV" in val_upper and "PTS" in val_upper:
        return "KV-PTS"
    
    if "COMMON" in val_upper and "ALGO" in val_upper:
        return "COMMON-ALGO"
    if "COMMON" in val_upper and "FSW" in val_upper:
        return "COMMON-FSW"
    
    if "NGICON" in val_upper:
        return "NGICON"
    if "HSP" in val_upper:
        return "HSP"
    if "HAWKEYE" in val_upper or "HS" in val_upper:
        return "HS"
    
    if val_upper == "FSW":
        return "COMMON-FSW"
    
    return "ENTER REPO NAME"

def extract_date_only(value):
    """Extract just the date from a value that may contain date and time.
    
    Handles multiple date formats and converts to mm/dd/yyyy format.
    Removes time portion if present.
    
    Args:
        value: Date value (datetime object, date object, or string)
        
    Returns:
        str: Date string in mm/dd/yyyy format, or None if no date found
    """
    if value is None:
        return None
    
    if isinstance(value, (datetime, date)):
        return value.strftime('%m/%d/%Y')
    
    value_str = str(value).strip()
    if not value_str:
        return None
    
    # Try to parse YYYY-MM-DD format and convert to mm/dd/yyyy
    match = re.match(r'(\d{4})-(\d{2})-(\d{2})', value_str)
    if match:
        year, month, day = match.groups()
        return f"{int(month)}/{int(day)}/{year}"
    
    # Already in M/D/YYYY format - return as is
    match = re.match(r'(\d{1,2}/\d{1,2}/\d{4})', value_str)
    if match:
        return match.group(1)
    
    # Convert M-D-YYYY format to mm/dd/yyyy
    match = re.match(r'(\d{1,2})-(\d{1,2})-(\d{4})', value_str)
    if match:
        month, day, year = match.groups()
        return f"{int(month)}/{int(day)}/{year}"
    
    # Convert D-Mon-YYYY format to mm/dd/yyyy
    match = re.match(r'(\d{1,2})-([A-Za-z]{3})-(\d{4})', value_str)
    if match:
        day, month_str, year = match.groups()
        try:
            month_num = datetime.strptime(month_str, '%b').month
            return f"{month_num}/{int(day)}/{year}"
        except:
            pass
    
    return value_str


def _extract_review_number(raw_value):
    """Best-effort conversion of a cell value to an int review number.

    Returns the integer review number, or None if the value cannot be parsed.
    Handles whitespace, extracts digits from mixed strings.
    """
    if raw_value is None:
        return None
    
    # If already an int or float, convert directly
    if isinstance(raw_value, (int, float)):
        try:
            return int(raw_value)
        except (ValueError, TypeError):
            return None
    
    # Convert to string and strip whitespace
    str_value = str(raw_value).strip()
    if not str_value:
        return None
    
    # Try direct conversion first
    try:
        return int(str_value)
    except (ValueError, TypeError):
        pass
    
    # Try to extract first sequence of 6+ digits
    import re
    match = re.search(r'\d{6,}', str_value)
    if match:
        try:
            return int(match.group())
        except (ValueError, TypeError):
            pass
    
    return None


def _passes_date_cutoff(date_str, log_failures=False, context=""):
    """Return True if the given date string passes the global DATE_CUTOFF.

    Accepts any value; non-string or unparsable values are treated as passing
    the cutoff (the caller may choose to handle warnings separately).
    
    Also filters out dates more than 3 months in the future (likely typos).
    
    Args:
        date_str: Date string to validate
        log_failures: If True, log detailed failure reasons
        context: Additional context for logging (e.g., review number)
    """
    if date_str is None:
        # No date: treat as passing the cutoff (same as previous behavior).
        return True
    
    # Allow "NO DATE" marker to pass through
    if str(date_str).strip().upper() == "NO DATE":
        return True

    # Try to normalize to a date object using the same formats supported by
    # extract_date_only. If parsing fails, we treat the value as NOT passing
    # the cutoff (Option 2: unknown dates are considered old and excluded).
    cutoff_dt = None
    try:
        cutoff_dt = datetime.strptime(DATE_CUTOFF, "%Y-%m-%d").date()
    except Exception:
        # If the cutoff itself is somehow invalid, do not block any rows.
        return True

    # Calculate max future date (3 months from today)
    today = datetime.now().date()
    max_future_dt = today + timedelta(days=90)  # ~3 months

    try:
        normalized = extract_date_only(date_str)
        if not normalized:
            if log_failures:
                cprint("WARNING", f"Date parsing failed for '{date_str}'{context} - date could not be normalized")
                cprint("WARNING", f"  → Suggested fix: Ensure date is in format MM/DD/YYYY, YYYY-MM-DD, or DD-Mon-YYYY")
            return False

        # If the normalized string has no plausible 4-digit year (19xx or 20xx),
        # treat it as an ambiguous/invalid date (e.g., "28 June 0900 - 1000")
        # and do not consider it as passing the cutoff.
        # This prevents time values like "1200" or "1315" from being mistaken for years.
        if not re.search(r"\b(19|20)\d{2}\b", str(normalized)):
            if log_failures:
                cprint("WARNING", f"Date parsing failed for '{date_str}'{context} - no valid year (19xx or 20xx) found")
                cprint("WARNING", f"  → Suggested fix: Ensure year is 4 digits (e.g., 2026, not 26)")
            return False

        # Try multiple patterns, mirroring extract_date_only semantics.
        # Now that extract_date_only returns mm/dd/yyyy format, prioritize that.
        parsed_date = None
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%d-%b-%Y"):
            try:
                candidate = datetime.strptime(normalized, fmt).date()
                parsed_date = candidate
                # Block dates strictly before the cutoff.
                if candidate < cutoff_dt:
                    if log_failures:
                        cprint("WARNING", f"Date '{date_str}'{context} is before cutoff date {DATE_CUTOFF}")
                        cprint("WARNING", f"  → Row will be ignored (only processing dates on or after {DATE_CUTOFF})")
                    return False
                # Block dates more than 3 months in the future (likely typos).
                if candidate > max_future_dt:
                    if log_failures:
                        cprint("WARNING", f"Date '{date_str}'{context} is more than 3 months in the future (parsed as {candidate})")
                        cprint("WARNING", f"  → Suggested fix: Check for year typo (e.g., 2027 should be 2026)")
                    return False
                return True
            except Exception:
                continue

        # Unknown/unparseable date formats are treated as not passing.
        if log_failures:
            cprint("WARNING", f"Date parsing failed for '{date_str}'{context} - format not recognized")
            cprint("WARNING", f"  → Suggested fix: Use format MM/DD/YYYY (e.g., 06/16/2026)")
        return False
    except Exception as e:
        # Any unexpected parsing error: treat as not passing the cutoff.
        if log_failures:
            cprint("WARNING", f"Date parsing error for '{date_str}'{context}: {e}")
        return False


def _row_is_empty(row) -> bool:
    """Return True if a row is None/empty or all cells are None."""
    if not row:
        return True
    return all(cell is None for cell in row)


def _collect_review_numbers(ws, start_row: int = 1) -> set:
    """Collect integer review numbers from a worksheet's second column.

    Args:
        ws: openpyxl worksheet.
        start_row: First row index to consider.

    Returns:
        set of parsed integer review numbers.
    """
    review_nums = set()
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if not row or len(row) <= 1:
            continue
        parsed = _extract_review_number(row[1])
        if parsed is not None:
            review_nums.add(parsed)
    return review_nums

def classify(row):
    """Determine which target sheet this row belongs to.
    
    Classification rules:
    - 'test' in column C -> ignore (None)
    - 'code' in column D -> Code sheet
    - Otherwise -> ignore (None)
    
    Args:
        row: Excel row with cell values
        
    Returns:
        str: Target sheet name ('Code') or None to ignore
    """
    col_c = str(row[2].value or "").lower()
    col_d = str(row[3].value or "").lower()

    if "test" in col_c:
        return None
    if "code" in col_d:
        return "Code"
    return None

def create_filtered_copy(source_path: str, output_path: str) -> None:
    """Create a filtered copy of the source Excel file.
    
    Creates a new xlsx file containing all rows from the source document that:
    - Do NOT contain the words 'design', 'document', or 'test' (case-insensitive)
    - Have a review number >= 105000
    
    Args:
        source_path: Path to source Excel file
        output_path: Path to output filtered Excel file
    """
    import openpyxl
    
    cprint("INFO", "Creating filtered copy of source document...")
    cprint("INFO", f"  Excluding rows containing: 'design', 'document', or 'test'")
    cprint("INFO", f"  Only including review numbers >= 105000")
    
    src_wb = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    src_ws = src_wb.active
    
    # Create new workbook for filtered data
    dst_wb = openpyxl.Workbook()
    dst_ws = dst_wb.active
    
    # Keywords to filter out (case-insensitive)
    filter_keywords = ['design', 'document', 'test']
    
    rows_included = 0
    rows_excluded_keywords = 0
    rows_excluded_review_num = 0
    
    # Process each row
    for row in src_ws.iter_rows():
        # Convert row to values for checking
        row_values = [str(cell.value or "").lower() for cell in row]
        
        # Check if any filter keyword appears in any cell
        contains_keyword = any(
            keyword in cell_value 
            for keyword in filter_keywords 
            for cell_value in row_values
        )
        
        if contains_keyword:
            rows_excluded_keywords += 1
            continue
        
        # Check review number (column A, index 0)
        review_num = _extract_review_number(row[0].value)
        if review_num is None or review_num < 105000:
            rows_excluded_review_num += 1
            continue
        
        # Include this row - copy cell values
        dst_ws.append([cell.value for cell in row])
        rows_included += 1
    
    # Save the filtered workbook
    dst_wb.save(output_path)
    src_wb.close()
    
    cprint("SUCCESS", f"✓ Filtered copy created: {output_path}")
    cprint("INFO", f"  Rows included: {rows_included}")
    cprint("INFO", f"  Rows excluded (keywords): {rows_excluded_keywords}")
    cprint("INFO", f"  Rows excluded (review # < 105000): {rows_excluded_review_num}")
    cprint("INFO", f"  Total rows processed: {rows_included + rows_excluded_keywords + rows_excluded_review_num}")


def process_and_categorize(source_path: str, dest_path: str) -> None:
    """Process Excel data and categorize into different sheets.
    
    Reads source Excel file and:
    1. Classifies each row into Code sheet (ignores non-code rows)
    2. Maps repository names from descriptions
    3. Extracts date-only values
    4. Writes categorized data to destination file
    
    Args:
        source_path: Path to source Excel file
        dest_path: Path to output categorized Excel file
    """
    import openpyxl
    
    src_wb = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    src_ws = src_wb.active

    dst_wb = openpyxl.Workbook()
    dst_wb.active.title = "Other"
    
    for name in TARGET_SHEETS:
        if name not in dst_wb.sheetnames:
            dst_wb.create_sheet(name)

    counts = {sheet: 0 for sheet in TARGET_SHEETS}
    ignored = 0
    row_number = 0  # Track row number for debugging
    blank_date_reviews = []  # Track reviews filtered due to blank dates

    # Peek at the first row to decide whether to skip header without
    # materializing the entire sheet into memory
    rows_iter = src_ws.iter_rows()
    first_row = next(rows_iter, None)

    def _row_iter():
        if first_row is None:
            return
        # Determine if the first row is header based on first cell type
        if isinstance(first_row[0].value, (int, float)):
            yield first_row
            for r in rows_iter:
                yield r
        else:
            for r in rows_iter:
                yield r

    # Process each row and categorize
    for row in _row_iter():
        row_number += 1
        
        # Skip empty rows
        if all(cell.value is None for cell in row):
            continue

        # Log every row we're examining
        raw_review = row[0].value
        cprint("DEBUG", f"[ROW {row_number}] Processing | Raw Review: {raw_review} | Col C: {str(row[2].value)[:30] if row[2].value else 'None'} | Col D: {str(row[3].value)[:30] if row[3].value else 'None'}")
        
        # Additional ignore rule: if original source column E contains 'test', treat as test row
        # Source columns are 0-based in 'row': A=0, B=1, C=2, D=3, E=4
        col_e = str(row[4].value or "").lower()
        if "test" in col_e:
            ignored += 1
            cprint("DEBUG", f"[FILTERED] Row {row_number} ignored: Column E contains 'test' | Review: {raw_review}")
            continue

        # Determine target sheet for this row
        target = classify(row)
        if target is None:
            ignored += 1
            cprint("DEBUG", f"[FILTERED] Row {row_number} ignored: Classified as None (test in column C) | Review: {raw_review}")
            continue
        
        cprint("DEBUG", f"[ROW {row_number}] Classified as: {target}")

        col_a = row[0].value
        if col_a is not None:
            try:
                col_a = int(col_a)
            except (ValueError, TypeError):
                pass

        dst_ws = dst_wb[target]

        # Extract source data (columns A-K, need index 10 for column K)
        src = [col_a] + [cell.value for cell in row[1:11]]
        
        # Build destination row with remapped columns
        new_row = [None] * 10
        new_row[1] = src[0]  # Review number: A → B
        new_row[3] = map_column_d(src[2])  # Repository: C → D (mapped)
        new_row[4] = src[10]  # Author: K → E
        new_row[2] = extract_date_only(src[6])  # Date: G → C (date only)
        
        # Handle blank/None dates - include them with "NO DATE" marker
        if not new_row[2]:
            new_row[2] = ""
            blank_date_reviews.append({
                'row': row_number,
                'review': new_row[1],
                'raw_date': src[6],
                'repo': new_row[3],
                'target': target
            })
            cprint("WARNING", f"[ROW {row_number}] Blank date detected - using 'NO DATE' | Review: {new_row[1]} | Raw date from col G: {src[6]}")
        else:
            cprint("DEBUG", f"[ROW {row_number}] Date extracted: {new_row[2]} (from raw: {src[6]})")
        
        # Skip rows with review numbers that have fewer than 6 digits
        review_num = _extract_review_number(new_row[1])
        if review_num is None:
            ignored += 1
            cprint("DEBUG", f"[FILTERED] Row {row_number} ignored: Could not parse review number '{new_row[1]}'")
            continue
        if review_num < 105000:
            ignored += 1
            cprint("DEBUG", f"[FILTERED] Row {row_number} ignored: Review number {review_num} < 100000")
            continue
        
        cprint("DEBUG", f"[ROW {row_number}] Review number parsed: {review_num}")
        
        # Apply date cutoff: skip rows that don't pass the cutoff
        context = f" (Review: {review_num}, Row: {row_number})"
        if not _passes_date_cutoff(new_row[2], log_failures=True, context=context):
            ignored += 1
            cprint("DEBUG", f"[FILTERED] Row {row_number} ignored: Date cutoff check failed | Review: {review_num} | Date: {new_row[2]}")
            continue
        
        cprint("SUCCESS", f"[ROW {row_number}] ✓ ACCEPTED | Review: {review_num} | Date: {new_row[2]} | Repo: {new_row[3]} | Target: {target}")
        dst_ws.append(new_row)
        counts[target] += 1

    dst_wb.save(dest_path)
    src_wb.close()
    
    # Log summary
    total_processed = sum(counts.values())
    cprint("INFO", f"\n{'='*60}")
    cprint("INFO", f"Processing Summary:")
    cprint("INFO", f"  Total rows processed: {total_processed}")
    cprint("INFO", f"  Total rows filtered/ignored: {ignored}")
    for sheet, count in counts.items():
        if count > 0:
            cprint("INFO", f"  {sheet}: {count} rows")
    
    # Report blank date reviews
    if blank_date_reviews:
        cprint("WARNING", f"\n{'='*60}")
        cprint("WARNING", f"BLANK DATE REVIEWS ({len(blank_date_reviews)} rows included with 'NO DATE'):")
        cprint("WARNING", f"These reviews were added to the tracker with date='NO DATE':")
        for item in blank_date_reviews:
            cprint("WARNING", f"  Row {item['row']}: Review {item['review']} | Repo: {item['repo']} | Target: {item['target']}")
        cprint("WARNING", f"\n  → Consider filling in the date field in the source file for accurate tracking")
        cprint("WARNING", f"{'='*60}")
    
    cprint("INFO", f"{'='*60}\n")


# ============================================================================
# EXCEL TABLE OPERATIONS
# ============================================================================

def append_to_excel_tables(file_path: str, sheet_data: dict) -> None:
    """Append data to existing Excel tables in the file.
    
    For each sheet with data:
    1. Finds existing table and expands its range
    2. Appends new rows after existing data
    3. Preserves existing data and formatting
    
    Args:
        file_path: Path to Excel file with tables
        sheet_data: Dictionary mapping sheet names to lists of row data
    """
    import openpyxl
    from openpyxl.utils import range_boundaries, get_column_letter
    
    wb = openpyxl.load_workbook(file_path)
    
    for sheet_name, data in sheet_data.items():
        if not data:
            continue
            
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Find and expand existing table if present
            if ws.tables:
                table_name = list(ws.tables.keys())[0]
                table = ws.tables[table_name]
                table_ref = table.ref
                
                # Calculate new table range to include appended rows
                min_col, min_row, max_col, max_row = range_boundaries(table_ref)
                start_row = max_row + 1
                
                new_max_row = max_row + len(data)
                new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
                table.ref = new_ref
            else:
                last_row = ws.max_row
                start_row = last_row + 1 if last_row > 0 else 1
        else:
            ws = wb.create_sheet(sheet_name)
            start_row = 1
        
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
    wb.save(file_path)

def open_sharepoint_urls():
    """Open SharePoint Excel files in browser after new data is added.
    
    Opens both the Code Review Tracker and Staging Tracker
    in the default web browser.
    """
    urls = [
        "https://lmco.sharepoint.us/:x:/r/sites/US-NGI-Cyber-Software/Shared%20Documents/General/Software%20Support/New%20Code%20Review%20Tracker.xlsx?d=w9c09dec9da2f44e09ef9bf9f0179d308&csf=1&web=1&e=isrm1W",
        "https://lmco.sharepoint.us/:x:/r/sites/US-NGI-Cyber-Software/Shared%20Documents/General/Software%20Support/Staging%20Tracker.xlsx?d=wfad7298975de4045a95a03e74cd9b673&csf=1&web=1&e=Ui8ySs"
    ]
    
    for url in urls:
        try:
            webbrowser.open_new_tab(url)
            time.sleep(1)  # Delay to ensure tabs open separately
        except Exception as e:
            cprint("WARNING", f"Failed to open URL: {e}")

# ============================================================================
# MAIN WORKFLOW
# ============================================================================

def run_step_1_download(playwright, args, source_file: str, dest_file: Path) -> None:
    """STEP 1: Download files from SharePoint.
    
    Downloads both source (original data) and destination (current state).
    Respects the --skip-download flag to reuse an existing source file.
    """
    # Auto-expire stale session files before validating sessions.
    _maybe_expire_sessions_by_age()

    if args.skip_download:
        cprint("SECTION", "=" * 60)
        cprint("INFO", "STEP 1: SKIPPED - Using existing file")
        cprint("SECTION", "=" * 60)
        
        if not Path(source_file).exists():
            cprint("ERROR", f"File not found: {source_file}")
            sys.exit(1)
        
        size_kb = Path(source_file).stat().st_size / 1024
        cprint("INFO", f"Using existing file: {source_file} ({size_kb:.1f} KB)")
        
        # Download destination for comparison
        dest_session_path = Path(DEST_READ_CONFIG["session_file"])
        _ensure_valid_session(playwright, dest_session_path, DEST_READ_CONFIG["site_url"])
        
        success = download_file_from_sharepoint(
            playwright, dest_session_path, 
            DEST_READ_CONFIG["site_url"], DEST_READ_CONFIG["file_path"], 
            str(dest_file)
        )
        if not success:
            sys.exit(1)
        
        # Save a copy for the dashboard to read
        cprint("INFO", f"Saving copy for dashboard: {dest_file}")
        return

    cprint("SECTION", "=" * 60)
    cprint("INFO", "STEP 1: Downloading files from SharePoint (PARALLEL)")
    cprint("SECTION", "=" * 60)
    
    source_session_path = Path(SOURCE_CONFIG["session_file"])
    dest_session_path = Path(DEST_READ_CONFIG["session_file"])
    
    # Validate sessions
    _ensure_valid_session(playwright, source_session_path, SOURCE_CONFIG["site_url"])
    _ensure_valid_session(playwright, dest_session_path, DEST_READ_CONFIG["site_url"])

    # Check which files need downloading
    downloads_needed = []
    
    cprint("INFO", "Checking for file changes...")
    if should_download_file(playwright, source_session_path, SOURCE_CONFIG["site_url"], 
                           SOURCE_CONFIG["file_path"], source_file):
        downloads_needed.append(("source", source_session_path, SOURCE_CONFIG["site_url"], 
                                SOURCE_CONFIG["file_path"], source_file))
    else:
        cprint("SUCCESS", f"✓ Source file is up to date: {source_file}")
    
    if should_download_file(playwright, dest_session_path, DEST_READ_CONFIG["site_url"],
                           DEST_READ_CONFIG["file_path"], str(dest_file)):
        downloads_needed.append(("destination", dest_session_path, DEST_READ_CONFIG["site_url"],
                                DEST_READ_CONFIG["file_path"], str(dest_file)))
    else:
        cprint("SUCCESS", f"✓ Destination file is up to date: {dest_file}")
    
    if not downloads_needed:
        cprint("SUCCESS", "\n✓ All files are up to date - no downloads needed!")
        cprint("SUCCESS", f"✓ Dashboard file ready: {dest_file}")
        return
    
    # Download files in parallel using ThreadPoolExecutor
    cprint("INFO", f"\nDownloading {len(downloads_needed)} file(s) in parallel...")
    
    from concurrent.futures import ThreadPoolExecutor, as_completed
    
    def download_task(file_type, session_path, site_url, file_path, output_path):
        """Task for parallel download execution."""
        try:
            from playwright.sync_api import sync_playwright
            with sync_playwright() as pw:
                success = download_file_from_sharepoint(
                    pw, session_path, site_url, file_path, output_path
                )
                return (file_type, success, output_path)
        except Exception as e:
            cprint("ERROR", f"Download error for {file_type}: {e}")
            return (file_type, False, output_path)
    
    results = []
    with ThreadPoolExecutor(max_workers=2) as executor:
        futures = [
            executor.submit(download_task, *task_args)
            for task_args in downloads_needed
        ]
        
        for future in as_completed(futures):
            try:
                file_type, success, output_path = future.result()
                results.append((file_type, success))
                
                if success:
                    cprint("SUCCESS", f"✓ Downloaded {file_type} to: {output_path}")
                else:
                    cprint("ERROR", f"✗ Failed to download {file_type}")
            except Exception as e:
                cprint("ERROR", f"Download exception: {e}")
                results.append((None, False))
    
    # Check if all downloads succeeded
    if not all(success for _, success in results):
        cprint("ERROR", "✗ One or more downloads failed")
        sys.exit(1)
    
    # Confirm destination file is ready for dashboard
    cprint("SUCCESS", f"\n✓ Dashboard file ready: {dest_file}")


def run_step_2_process(source_file: str) -> str:
    """STEP 2: Process and categorize data.
    
    Reads source file, classifies reviews by type, maps repositories,
    and writes the processed data to a temporary Excel file.
    Returns the path to the processed file.
    """
    cprint("SECTION", "=" * 60)
    cprint("INFO", "STEP 2: Processing and categorizing data")
    cprint("SECTION", "=" * 60)
    
    processed_file = tempfile.mktemp(suffix='.xlsx')
    
    try:
        process_and_categorize(source_file, processed_file)
        cprint("SUCCESS", f"✓ Processed data saved to: {processed_file}")
    except Exception as e:
        cprint("ERROR", f"Failed to process data: {e}")
        sys.exit(1)
    
    return processed_file


def _filter_new_rows_against_production(dest_file: Path, sheet_data: dict) -> dict:
    """Filter processed data to rows missing from the production destination file."""
    import openpyxl
    dest_wb = openpyxl.load_workbook(dest_file, read_only=True, data_only=True)
    filtered_data = {}
    
    for sheet_name, source_data in sheet_data.items():
        if sheet_name in dest_wb.sheetnames:
            dest_ws = dest_wb[sheet_name]
            existing_review_nums = _collect_review_numbers(dest_ws, start_row=1)

            cprint("INFO", f"  Sheet '{sheet_name}': {len(existing_review_nums)} unique review #s in production")
            
            new_rows = []
            for source_row in source_data:
                review_num = source_row[1] if len(source_row) > 1 else None

                source_review_num = _extract_review_number(review_num)
                if source_review_num is None:
                    cprint("DEBUG", f"[FILTERED] Stage 2: Could not parse review number '{review_num}' | sheet={sheet_name}")
                    continue

                # Date cutoff already validated in Stage 1 (process_and_categorize)
                # No need to re-check here

                if source_review_num not in existing_review_nums:
                    new_rows.append(source_row)
                    cprint("DEBUG", f"[NEW] vs production | sheet={sheet_name} review={source_review_num}")
                else:
                    cprint("DEBUG", f"[FILTERED] Stage 2: Review {source_review_num} already exists in production | sheet={sheet_name}")
            
            filtered_data[sheet_name] = new_rows
            cprint("INFO", f"  Sheet '{sheet_name}': {len(new_rows)} new rows to upload (missing from production)")
        else:
            filtered_data[sheet_name] = source_data
    
    dest_wb.close()
    return filtered_data


def _load_sheet_data(wb, target_sheets):
    """Load data from the processed workbook for the given target sheets.

    Args:
        wb: An openpyxl Workbook instance for the processed file.
        target_sheets: Iterable of sheet names to load.

    Returns:
        dict: Mapping sheet name -> list of row lists (including empty columns),
        with completely empty rows skipped.
    """
    sheet_data = {}

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if _row_is_empty(row):
                continue
            rows.append(list(row))

        if rows:
            sheet_data[sheet_name] = rows

    return sheet_data


def _count_rows(sheet_data: dict) -> int:
    """Return total number of rows across all sheets in the given dict.

    Args:
        sheet_data: Mapping of sheet name -> list of row lists.

    Returns:
        int: Sum of lengths of all row lists.
    """
    if not sheet_data:
        return 0
    return sum(len(rows) for rows in sheet_data.values())


def _log_no_data_and_cleanup(message: str, processed_file: str) -> None:
    """Log a no-new-data message and remove the temporary processed file.

    Args:
        message: Human-readable message to display to the user.
        processed_file: Path to the temporary processed Excel file to delete.
    """
    cprint("SUCCESS", "=" * 60)
    cprint("SUCCESS", message)
    cprint("SUCCESS", "=" * 60)

    try:
        if processed_file and os.path.exists(processed_file):
            os.remove(processed_file)
    except Exception as e:
        cprint("WARNING", f"Failed to delete processed file '{processed_file}': {e}")


def _filter_against_staging(playwright, filtered_data: dict) -> dict:
    """Filter rows that already exist in the staging SharePoint file."""
    import openpyxl
    staging_file = Path(__file__).parent / "staging_comparison.xlsx"
    dest_session_path = Path(DEST_UPLOAD_CONFIG["session_file"])
    success = download_file_from_sharepoint(
        playwright, dest_session_path,
        DEST_UPLOAD_CONFIG["site_url"], DEST_UPLOAD_CONFIG["file_path"],
        str(staging_file)
    )
    
    if not success:
        cprint("WARNING", "Could not download staging file - will proceed with all rows")
        return filtered_data
    
    staging_wb = openpyxl.load_workbook(staging_file, data_only=True, read_only=True)
    final_filtered_data = {}
    
    for sheet_name in TARGET_SHEETS:
        if sheet_name not in filtered_data or not filtered_data[sheet_name]:
            continue
        
        if sheet_name not in staging_wb.sheetnames:
            final_filtered_data[sheet_name] = filtered_data[sheet_name]
            continue
        
        staging_sheet = staging_wb[sheet_name]
        staging_review_nums = _collect_review_numbers(staging_sheet, start_row=4)
        
        new_to_staging = []
        for row in filtered_data[sheet_name]:
            review_num = row[1] if len(row) > 1 else None
            r = _extract_review_number(review_num)
            if r is None:
                cprint("DEBUG", f"[FILTERED] Stage 3: Could not parse review number '{review_num}' | sheet={sheet_name}")
                continue

            if r in staging_review_nums:
                cprint("DEBUG", f"[FILTERED] Stage 3: Review {r} already exists in staging | sheet={sheet_name}")
            else:
                new_to_staging.append(row)
                cprint("INFO", f"[NEW] Review {r} will be added (not in staging) | sheet={sheet_name}")

        if new_to_staging:
            final_filtered_data[sheet_name] = new_to_staging
            cprint("INFO", f"  Sheet '{sheet_name}': {len(new_to_staging)} rows new to staging (from {len(filtered_data[sheet_name])} new vs production)")
        else:
            cprint("INFO", f"  Sheet '{sheet_name}': 0 rows new to staging (all {len(filtered_data[sheet_name])} new vs production already exist in staging)")
    
    staging_wb.close()
    os.remove(staging_file)
    return final_filtered_data


def run_step_3_diff_and_upload(playwright, processed_file: str, dest_file: Path) -> None:
    """STEP 3: Check for duplicates and upload new rows to staging SharePoint file."""
    import openpyxl
    cprint("SECTION", "=" * 60)
    cprint("INFO", "STEP 3: Checking for duplicates and uploading")
    cprint("SECTION", "=" * 60)
    
    wb = openpyxl.load_workbook(processed_file)
    sheet_data = _load_sheet_data(wb, TARGET_SHEETS)
    
    wb.close()
    
    filtered_data = _filter_new_rows_against_production(dest_file, sheet_data)
    
    # Check if there's any new data to upload vs production
    total_rows = _count_rows(filtered_data)
    if total_rows == 0:
        _log_no_data_and_cleanup("✓ NO NEW DATA - All rows already exist!", processed_file)
        return
    
    cprint("INFO", f"\nFound {total_rows} rows new to production file...")
    cprint("INFO", "Checking staging file for duplicates...")
    
    filtered_data = _filter_against_staging(playwright, filtered_data)
    
    # Final check after filtering against staging file
    total_rows = _count_rows(filtered_data)
    if total_rows == 0:
        _log_no_data_and_cleanup("✓ NO NEW DATA - All rows already exist in staging file!", processed_file)
        return
    
    cprint("INFO", f"\nUploading {total_rows} new rows to staging file...")
    cprint("INFO", "  Browser tabs will be opened after successful upload")
    
    # Append to destination file in a temporary copy
    temp_upload = tempfile.mktemp(suffix='.xlsx')
    shutil.copy(dest_file, temp_upload)
    
    append_to_excel_tables(temp_upload, filtered_data)
    
    # Upload modified file back to SharePoint (to staging file)
    dest_session_path = Path(DEST_UPLOAD_CONFIG["session_file"])
    success = upload_file_to_sharepoint(
        playwright, dest_session_path,
        DEST_UPLOAD_CONFIG["site_url"], DEST_UPLOAD_CONFIG["file_path"],
        temp_upload
    )
    
    if success:
        cprint("SUCCESS", "✓ Successfully uploaded all data to SharePoint!")
        
        # Open SharePoint files in browser since new data was added
        open_sharepoint_urls()
    else:
        cprint("ERROR", "✗ Upload failed.")
        os.remove(processed_file)
        os.remove(temp_upload)
        sys.exit(1)
    
    # Cleanup temporary files
    os.remove(processed_file)
    os.remove(temp_upload)


def main():
    """Main workflow function.
    
    Workflow steps:
    1. Download files from SharePoint (source and destination)
    2. Create filtered copy of source (excluding design/document/test rows)
    3. Process and categorize source data
    4. Check for duplicates against destination and staging
    5. Append only new rows to destination staging file
    6. Upload updated staging file to SharePoint
    """
    cprint("SECTION", "=" * 60)
    print("Code Review Updater - Starting Workflow")
    cprint("SECTION", "=" * 60)

    import argparse
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        cprint("ERROR", "Playwright is not installed.")
        print("Fix: pip install playwright openpyxl")
        print("     playwright install chromium")
        sys.exit(1)

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        cprint("ERROR", "openpyxl is not installed.")
        print("Fix: pip install openpyxl")
        sys.exit(1)

    parser = argparse.ArgumentParser(description="Code Review Updater Workflow")
    parser.add_argument('--skip-download', action='store_true', 
                       help='Skip download step and use existing file')
    args = parser.parse_args()

    source_file = SOURCE_CONFIG["output_path"]
    dest_file = Path(__file__).parent / "destination_current.xlsx"
    
    # Define output path for filtered copy (same directory as source)
    filtered_copy_path = str(Path(source_file).parent / "NGI FSW Formal Review Status - Filtered.xlsx")

    with sync_playwright() as playwright:
        run_step_1_download(playwright, args, source_file, dest_file)
        
        # Create filtered copy after downloading source
        cprint("SECTION", "=" * 60)
        cprint("INFO", "STEP 1.5: Creating filtered copy of source document")
        cprint("SECTION", "=" * 60)
        try:
            create_filtered_copy(source_file, filtered_copy_path)
        except Exception as e:
            cprint("ERROR", f"Failed to create filtered copy: {e}")
            cprint("WARNING", "Continuing with main workflow...")
        
        processed_file = run_step_2_process(source_file)
        run_step_3_diff_and_upload(playwright, processed_file, dest_file)

    cprint("SUCCESS", "=" * 60)
    # Use ASCII-only text for compatibility with Windows consoles
    print("WORKFLOW COMPLETED SUCCESSFULLY")
    cprint("SUCCESS", "=" * 60)

if __name__ == "__main__":
    main()
