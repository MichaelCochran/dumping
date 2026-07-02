#!/usr/bin/env python3
"""
Code Review Dashboard - Tkinter App
Shows outstanding reports from destination_current.xlsx
"""

import sys
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
from pathlib import Path
import openpyxl
from datetime import datetime, timedelta
import json
import threading
import subprocess
import os
import time

# ============================================================================
# CONFIGURATION
# ============================================================================
# CUSTOMIZATION: Update file paths if your data files are in different locations.
# The dashboard is intentionally file-based so it can run without direct
# SharePoint access – it simply reads whatever the updater last wrote.

SCRIPT_DIR = Path(__file__).parent
# CUSTOMIZATION: Path to the Excel file containing review data
# This should match the file written by the "super" updater script.
DEST_FILE = SCRIPT_DIR / "destination_current.xlsx"  # Production data source file
STAGING_FILE = SCRIPT_DIR / "staging_current.xlsx"  # Staging data source file
# CUSTOMIZATION: Path to the JSON file containing repository filter settings
# This file allows each user to limit the dashboard to specific repos.
FILTER_CONFIG = SCRIPT_DIR / "user_filters.json"  # User filter configuration

# SharePoint configuration for downloading production file after refresh
DEST_READ_CONFIG = {
    "site_url": "https://lmco.sharepoint.us/sites/US-NGI-Cyber-Software",
    "file_path": "/Shared Documents/General/Software Support/New Code Review Tracker.xlsx",
    "session_file": str(SCRIPT_DIR / "sp_dest_session.json"),
}

# SharePoint configuration for downloading staging file after refresh
# Uses same session as production since it's the same SharePoint site
STAGING_READ_CONFIG = {
    "site_url": "https://lmco.sharepoint.us/sites/US-NGI-Cyber-Software",
    "file_path": "/Shared Documents/General/Software Support/Staging Tracker.xlsx",
    "session_file": str(SCRIPT_DIR / "sp_dest_session.json"),
}

# CUSTOMIZATION: Excel sheet and column configuration
# Change these if your Excel file has a different structure
SHEET_NAME = "Code Reviews"  # Name of the sheet containing review data
HEADER_ROW_COUNT = 3  # Number of header rows to skip (data starts after this)

# Column indices (0-based) - CUSTOMIZATION: Adjust if your columns are different
COL_REVIEW_NUM = 1  # Column B - Review Number
COL_DATE = 2  # Column C - Date
COL_REPO = 3  # Column D - Repository
COL_DESCRIPTION = 4  # Column E - Description
COL_REPORT_UPLOADED = 5  # Column F - Report Uploaded status
COL_NOTES = 9  # Column J - Notes

# CUSTOMIZATION: Keywords to identify cancelled/broken reviews in notes
CANCELLED_KEYWORDS = ["cancelled", "pipeline"]

# CUSTOMIZATION: UI Color Scheme
# Background colors
COLOR_BG_MAIN = "#f0f0f0"  # Main window background (light grey)
COLOR_BG_CARD_BLUE = "#dbeafe"  # Default summary card (blue)
COLOR_BG_CARD_RED = "#fef2f2"  # High count alert (red)
COLOR_BG_CARD_YELLOW = "#fefce8"  # Medium count warning (yellow)
COLOR_BG_CARD_GREEN = "#f0fdf4"  # Low count ok (green)
COLOR_BG_DIALOG = "#f9fafb"  # Dialog/modal background
COLOR_BG_TERMINAL = "#1e1e1e"  # Terminal/console background

# Text colors
COLOR_TEXT_BLUE = "#2563eb"  # Primary blue text
COLOR_TEXT_BLUE_DARK = "#1e40af"  # Darker blue text
COLOR_TEXT_BLUE_LIGHT = "#60a5fa"  # Lighter blue text
COLOR_TEXT_RED = "#dc2626"  # Red text for high counts
COLOR_TEXT_RED_DARK = "#991b1b"  # Darker red text
COLOR_TEXT_YELLOW = "#eab308"  # Yellow text for medium counts
COLOR_TEXT_YELLOW_DARK = "#854d0e"  # Darker yellow text
COLOR_TEXT_GREEN = "#16a34a"  # Green text for low counts
COLOR_TEXT_GREEN_DARK = "#166534"  # Darker green text
COLOR_TEXT_GREY = "#6b7280"  # Grey text
COLOR_TEXT_GREY_LIGHT = "#9ca3af"  # Light grey text
COLOR_TEXT_GREY_DARK = "#1f2937"  # Dark grey text
COLOR_TEXT_GREY_DARKER = "#374151"  # Even darker grey text
COLOR_TEXT_ORANGE = "#ff8800"  # Bright orange for tomorrow dates
COLOR_TEXT_TERMINAL = "#d4d4d4"  # Terminal text
COLOR_TEXT_TERMINAL_HEADER = "#569cd6"  # Terminal header (VS Code blue)
COLOR_TEXT_WHITE = "white"  # White text

# Background colors - Additional
COLOR_BG_BLUE_LIGHT = "#eff6ff"  # Very light blue background
COLOR_BG_GREY_LIGHT = "#e5e7eb"  # Light grey background  
COLOR_BG_GREY_LIGHTER = "#f3f4f6"  # Even lighter grey background

# Button colors
COLOR_BTN_PRIMARY = "#3b82f6"  # Primary blue button
COLOR_BTN_SUCCESS = "#10b981"  # Success green button
COLOR_BTN_DANGER = "#ef4444"  # Danger red button
COLOR_BTN_GREY = "#6b7280"  # Inactive grey button

# CUSTOMIZATION: UI Fonts
FONT_WINDOW_TITLE = ("Arial", 16, "bold")  # Large window headers
FONT_TITLE = ("Arial", 14, "bold")
FONT_REPO_NAME = ("Arial", 13, "bold")  # Repository names in cards
FONT_SUBTITLE = ("Arial", 12)
FONT_BADGE = ("Arial", 12, "bold")  # Count badges
FONT_BUTTON = ("Arial", 11)
FONT_BODY = ("Arial", 10)
FONT_SMALL = ("Arial", 9)
FONT_SMALL_ITALIC = ("Arial", 9, "italic")  # Notes text
FONT_TINY = ("Arial", 8)
FONT_ICON = ("Arial", 20)  # Emoji icons in cards
FONT_MODAL_COUNT = ("Arial", 28, "bold")  # Large count in modal windows
FONT_LARGE_COUNT = ("Arial", 36, "bold")  # Main dashboard count
FONT_TERMINAL = ("Consolas", 9)
FONT_TERMINAL_HEADER = ("Consolas", 10, "bold")

# CUSTOMIZATION: Window sizes and padding
WINDOW_SIZE_MAIN = "400x600"
WINDOW_SIZE_PROGRESS = "450x300"
WINDOW_SIZE_LOGIN = "420x260"
WINDOW_SIZE_ERROR = "600x400"
PADDING_STANDARD = 20
PADDING_SMALL = 10
PADDING_LARGE = 50

# CUSTOMIZATION: Count thresholds for color coding
THRESHOLD_HIGH = 10  # Red alert - many outstanding reviews
THRESHOLD_MEDIUM = 6  # Yellow warning - moderate outstanding reviews
THRESHOLD_LOW = 1    # Green ok - few outstanding reviews

# Repo card thresholds (used when all reviews in single repo)
THRESHOLD_REPO_HIGH = 5  # Red alert for single repo
THRESHOLD_REPO_MEDIUM = 3  # Yellow warning for single repo

def create_button(parent, text, command, bg=None, fg=COLOR_TEXT_WHITE, 
                 font=FONT_BUTTON, **kwargs):
    """Create a standardized button with consistent styling.
    
    Args:
        parent: Parent widget
        text: Button text
        command: Command to execute on click
        bg: Background color (defaults to primary blue)
        fg: Foreground color (defaults to white)
        font: Font tuple (defaults to FONT_BUTTON)
        **kwargs: Additional button parameters
        
    Returns:
        tk.Button: Configured button widget
    """
    if bg is None:
        bg = COLOR_BTN_PRIMARY
    
    return tk.Button(
        parent,
        text=text,
        command=command,
        bg=bg,
        fg=fg,
        font=font,
        relief="flat",
        cursor="hand2",
        **kwargs
    )

def load_excel_sheet(file_path):
    """Load Excel sheet and return rows, or None if error.
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        list: List of row tuples, or None if sheet not found or error
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        if SHEET_NAME not in wb.sheetnames:
            wb.close()
            return None
        
        ws = wb[SHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        
        return rows
    except Exception as e:
        print(f"Error loading Excel file {file_path}: {e}")
        return None

def has_notes_content(notes_value):
    """Check if notes field has meaningful content.
    
    Args:
        notes_value: Value from notes column
        
    Returns:
        tuple: (has_content: bool, notes_str: str)
    """
    notes_str = str(notes_value).strip() if notes_value else ""
    has_content = bool(notes_str) and notes_str.lower() != 'none'
    return has_content, notes_str

def is_cancelled_or_broken(notes_str):
    """Check if notes indicate a cancelled or broken review.
    
    Args:
        notes_str: String content of notes
        
    Returns:
        bool: True if notes contain cancelled/broken keywords
    """
    notes_lower = notes_str.lower()
    return any(keyword in notes_lower for keyword in CANCELLED_KEYWORDS)

def format_date_only(value):
    """Extract date without time from various formats and return in YYYY-MM-DD format"""
    if value is None:
        return 'N/A'
    
    # If it's already a datetime object, format it
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    
    # Convert to string and remove time portion
    value_str = str(value)
    
    # Remove time if present (e.g., "2026-05-06 00:00:00" -> "2026-05-06")
    if ' ' in value_str:
        value_str = value_str.split(' ')[0]
    
    # Try to parse and normalize the date format to YYYY-MM-DD
    try:
        # Try parsing M/D/YYYY format
        if '/' in value_str:
            parsed_date = datetime.strptime(value_str, '%m/%d/%Y')
            return parsed_date.strftime('%Y-%m-%d')
        # Try parsing YYYY-MM-DD format (already correct, but validate)
        elif '-' in value_str:
            parsed_date = datetime.strptime(value_str, '%Y-%m-%d')
            return parsed_date.strftime('%Y-%m-%d')
    except:
        pass
    
    return value_str if value_str else 'N/A'

def normalize_date_for_comparison(date_str):
    """Normalize date string to YYYY-MM-DD format for comparison.
    
    Handles multiple formats:
    - M/D/YYYY (e.g., 6/11/2026)
    - YYYY-MM-DD (e.g., 2026-06-13)
    - N/A or invalid dates return None
    """
    if not date_str or date_str == 'N/A':
        return None
    
    try:
        # Try parsing M/D/YYYY format
        if '/' in date_str:
            parsed_date = datetime.strptime(date_str, '%m/%d/%Y')
            return parsed_date.strftime('%Y-%m-%d')
        # Try parsing YYYY-MM-DD format
        elif '-' in date_str:
            parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
            return parsed_date.strftime('%Y-%m-%d')
    except:
        pass
    
    return None

def is_report_uploaded(report_uploaded_value):
    """Check if report uploaded field indicates the report has been uploaded.
    
    Args:
        report_uploaded_value: Value from the Report Uploaded column
        
    Returns:
        bool: True if report has been uploaded, False otherwise
    """
    # Empty, None, or whitespace means not uploaded
    if not report_uploaded_value:
        return False
    
    value_str = str(report_uploaded_value).strip().lower()
    
    # 'none' or empty string means not uploaded
    if not value_str or value_str == 'none':
        return False
    
    # 'no' anywhere in the value means not uploaded
    if 'no' in value_str:
        return False
    
    # Any other value means uploaded (typically a date)
    return True

class DashboardApp:
    """Main dashboard application for displaying code review status.
    
    Displays outstanding reviews by repository and provides access to cancelled/broken reviews.
    Reviews are loaded from destination_current.xlsx and can be filtered by repository.
    """
    
    def __init__(self, root):
        """Initialize the dashboard application.
        
        Args:
            root: The tkinter root window
        """
        self.root = root
        # CUSTOMIZATION: Change window title as desired
        self.root.title("Code Review Dashboard - Super")
        # CUSTOMIZATION: Adjust window size (width x height)
        self.root.geometry(WINDOW_SIZE_MAIN)  # Main window dimensions
        # CUSTOMIZATION: Change background color (hex format)
        self.root.configure(bg=COLOR_BG_MAIN)  # Light grey background
        
        # Data structures for storing review information
        # These dictionaries are rebuilt from scratch each time data is loaded.
        self.repo_data = {}  # Outstanding reviews grouped by repository
        self.cancelled_data = {}  # Cancelled/broken reviews grouped by repository
        self.other_data = {}  # Other reviews (notes present but not cancelled/broken)
        self.total_outstanding = 0  # Count of actionable outstanding reviews
        self.total_cancelled = 0  # Count of cancelled/broken reviews
        self.total_other = 0  # Count of other reviews
        self.current_file_source = "production"  # Track which file is being displayed
        
        # Track which repo cards are expanded in the UI
        self.expanded_repos = set()  # Expanded outstanding repo cards
        self.expanded_cancelled = set()  # Expanded cancelled repo cards in popup
        self.expanded_other = set()  # Expanded other repo cards in popup
        
        # Load repository filter from config file
        # This lets individual users focus the dashboard on a subset of repos.
        self.user_filter = self.load_user_filter()
        
        # Build UI and load data
        self.create_widgets()
        
        # Download latest production file before loading data
        # This ensures dashboard always shows current SharePoint state
        self._download_production_file_on_startup()
        
        # Schedule hourly auto-refresh (3600000 ms = 1 hour)
        self._schedule_auto_refresh()
    
    def create_widgets(self):
        """Create all UI widgets for the main dashboard window.
        
        Layout includes:
        - Summary card showing total outstanding count with color coding
        - Toolbar with Refresh and Cancelled/Broken buttons
        - Scrollable list of repository cards
        """
        # Summary Frame - top section with total count
        summary_frame = tk.Frame(self.root, bg=COLOR_BG_MAIN)
        summary_frame.pack(fill="x", padx=PADDING_STANDARD, pady=(PADDING_STANDARD, PADDING_SMALL))
        
        # Summary Card
        self.summary_card = tk.Frame(summary_frame, bg=COLOR_BG_CARD_BLUE, relief="solid", bd=1)
        self.summary_card.pack(fill="x")
        
        summary_inner = tk.Frame(self.summary_card, bg=COLOR_BG_CARD_BLUE)
        summary_inner.pack(fill="x", padx=PADDING_STANDARD, pady=15)
        
        summary_label = tk.Label(
            summary_inner,
            text="Total Outstanding Reports",
            font=FONT_TITLE,
            bg=COLOR_BG_CARD_BLUE,
            fg=COLOR_TEXT_BLUE_DARK
        )
        summary_label.pack(anchor="center")
        
        self.total_label = tk.Label(
            summary_inner,
            text="0",
            font=FONT_LARGE_COUNT,
            bg=COLOR_BG_CARD_BLUE,
            fg=COLOR_TEXT_BLUE
        )
        self.total_label.pack(anchor="center")
        
        self.file_info_label = tk.Label(
            summary_inner,
            text="",
            font=FONT_SMALL,
            bg=COLOR_BG_CARD_BLUE,
            fg=COLOR_TEXT_BLUE_LIGHT
        )
        self.file_info_label.pack(anchor="center", pady=(5, 0))
        
        self.filter_info_label = tk.Label(
            summary_inner,
            text="",
            font=FONT_TINY,
            bg=COLOR_BG_CARD_BLUE,
            fg=COLOR_TEXT_BLUE_LIGHT
        )
        self.filter_info_label.pack(anchor="center", pady=(2, 0))
        
        # Toolbar Frame - contains action buttons
        toolbar_frame = tk.Frame(self.root, bg=COLOR_BG_MAIN)
        toolbar_frame.pack(fill="x", padx=PADDING_STANDARD, pady=(0, PADDING_SMALL))
        
        # First row - Refresh button (full width)
        refresh_container = tk.Frame(toolbar_frame, bg=COLOR_BG_MAIN)
        refresh_container.pack(fill="x", pady=(0, 5))
        
        self.refresh_btn = create_button(
            refresh_container,
            text="⭮ Refresh Data",
            command=self.refresh_data,
            pady=8
        )
        self.refresh_btn.pack(fill="x")
        
        # Second row - Cancelled/Broken and Other buttons (split width)
        button_row = tk.Frame(toolbar_frame, bg=COLOR_BG_MAIN)
        button_row.pack(fill="x")
        
        self.cancelled_btn = create_button(
            button_row,
            text="🚫 Cancelled/Broken",
            command=self.show_cancelled_window,
            bg=COLOR_BTN_GREY,
            disabledforeground=COLOR_TEXT_GREY_LIGHT,
            pady=8,
            state="disabled"
        )
        self.cancelled_btn.pack(side="left", fill="x", expand=True, padx=(0, 2.5))
        
        self.other_btn = create_button(
            button_row,
            text="📋 Other",
            command=self.show_other_window,
            bg=COLOR_BTN_GREY,
            disabledforeground=COLOR_TEXT_GREY_LIGHT,
            pady=8,
            state="disabled"
        )
        self.other_btn.pack(side="left", fill="x", expand=True, padx=(2.5, 0))
        
        # Canvas with Scrollbar for Repository List - main scrollable content area
        # All repository cards are children of this scrollable frame.
        canvas_frame = tk.Frame(self.root, bg=COLOR_BG_MAIN)
        canvas_frame.pack(fill="both", expand=True, padx=PADDING_STANDARD, pady=(0, PADDING_STANDARD))
        
        self.canvas = tk.Canvas(canvas_frame, bg=COLOR_BG_MAIN, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.canvas.yview)
        
        self.scrollable_frame = tk.Frame(self.canvas, bg=COLOR_BG_MAIN)
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self._on_frame_configure()
        )
        
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(self.canvas_window, width=e.width))
        
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        # Bind mousewheel
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        
        # Empty state label
        self.empty_label = tk.Label(
            self.scrollable_frame,
            text="",
            font=FONT_TITLE,
            bg=COLOR_BG_MAIN,
            fg=COLOR_TEXT_GREY
        )
        self.empty_label.pack(pady=PADDING_LARGE)
    
    def _on_frame_configure(self):
        """Update scroll region and show/hide scrollbar as needed.
        
        Dynamically shows scrollbar only when content exceeds visible height.
        """
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        
        # Show scrollbar only if content exceeds canvas height
        canvas_height = self.canvas.winfo_height()
        content_height = self.scrollable_frame.winfo_reqheight()
        
        if content_height > canvas_height:
            self.scrollbar.pack(side="right", fill="y")
        else:
            self.scrollbar.pack_forget()
    
    def _on_mousewheel(self, event):
        """Handle mousewheel scrolling for the main canvas."""
        self.canvas.yview_scroll(int(-1 * (event.delta / 100)), "units")  # Scroll speed
    
    def load_user_filter(self):
        """Load repository filter from config file.
        
        Returns:
            list: Repository names/patterns to filter, or empty list for all repos
        """
        if not FILTER_CONFIG.exists():
            # No filter file means "show everything".
            return []
        
        try:
            with open(FILTER_CONFIG, 'r') as f:
                config = json.load(f)
            
            repos = config.get('repos', [])
            return repos if repos else []
        
        except Exception as e:
            print(f"Error loading filter config: {e}")
            return []
    
    def _schedule_auto_refresh(self):
        """Schedule automatic refresh every hour to keep dashboard data current.
        
        This silently checks for file changes and reloads if needed without
        disrupting the user. Only downloads files if they've changed on SharePoint.
        """
        def auto_refresh_task():
            print("\n[Auto-refresh] Checking for updates...")
            
            try:
                import sys
                sys.path.insert(0, str(SCRIPT_DIR))
                from code_review_updater import (download_file_from_sharepoint, _ensure_valid_session,
                                                should_download_file)
                from playwright.sync_api import sync_playwright
                from concurrent.futures import ThreadPoolExecutor, as_completed
                
                with sync_playwright() as playwright:
                    # Validate sessions
                    prod_session_path = Path(DEST_READ_CONFIG["session_file"])
                    staging_session_path = Path(STAGING_READ_CONFIG["session_file"])
                    
                    try:
                        _ensure_valid_session(playwright, prod_session_path, DEST_READ_CONFIG["site_url"])
                        _ensure_valid_session(playwright, staging_session_path, STAGING_READ_CONFIG["site_url"])
                    except:
                        # If authentication fails during auto-refresh, skip this cycle
                        print("[Auto-refresh] Skipping - authentication needed")
                        return
                    
                    # Check which files need downloading
                    downloads_needed = []
                    
                    if should_download_file(playwright, prod_session_path, DEST_READ_CONFIG["site_url"],
                                          DEST_READ_CONFIG["file_path"], str(DEST_FILE)):
                        downloads_needed.append(("production", prod_session_path, DEST_READ_CONFIG["site_url"],
                                               DEST_READ_CONFIG["file_path"], str(DEST_FILE)))
                    
                    if should_download_file(playwright, staging_session_path, STAGING_READ_CONFIG["site_url"],
                                          STAGING_READ_CONFIG["file_path"], str(STAGING_FILE)):
                        downloads_needed.append(("staging", staging_session_path, STAGING_READ_CONFIG["site_url"],
                                               STAGING_READ_CONFIG["file_path"], str(STAGING_FILE)))
                    
                    if not downloads_needed:
                        print("[Auto-refresh] No updates needed")
                        return
                    
                    print(f"[Auto-refresh] Downloading {len(downloads_needed)} updated file(s)...")
                    
                    # Download in parallel
                    def download_one(file_type, session_path, site_url, file_path, output_path):
                        with sync_playwright() as pw:
                            return download_file_from_sharepoint(pw, session_path, site_url, file_path, output_path)
                    
                    with ThreadPoolExecutor(max_workers=2) as executor:
                        futures = [executor.submit(download_one, *args) for args in downloads_needed]
                        results = [f.result() for f in as_completed(futures)]
                    
                    if all(results):
                        print("[Auto-refresh] Files updated, reloading dashboard...")
                        # Reload data on main thread
                        self.root.after(0, self.load_data_from_best_source)
                    
            except Exception as e:
                print(f"[Auto-refresh] Error: {e}")
        
        def schedule_next():
            # Run auto-refresh in background
            threading.Thread(target=auto_refresh_task, daemon=True).start()
            # Schedule next refresh in 1 hour (3600000 ms)
            self.root.after(3600000, schedule_next)
        
        # Start the refresh cycle
        self.root.after(3600000, schedule_next)
    
    def _download_production_file_on_startup(self):
        """Download both production and staging files from SharePoint on startup.
        
        This runs in a background thread to avoid blocking the UI during startup.
        Shows a loading message while downloading. After downloading, loads from
        whichever file has more outstanding reviews.
        """
        # Show loading message
        self.empty_label.config(text="Loading latest data from SharePoint...")
        self.empty_label.pack(pady=50)
        
        # Setup login dialog support (same as refresh_data)
        self._login_window = None
        
        def ensure_login_window():
            if self._login_window is not None and tk.Toplevel.winfo_exists(self._login_window):
                return self._login_window

            login_window = tk.Toplevel(self.root)
            self._login_window = login_window
            login_window.title("Complete SharePoint Login")
            login_window.geometry(WINDOW_SIZE_LOGIN)
            login_window.configure(bg=COLOR_BG_DIALOG)
            login_window.grab_set()

            msg = (
                "If a browser window opened for SharePoint,\n"
                "please complete the FULL login process:\n\n"
                "1. Choose your identity provider\n"
                "2. Enter credentials and complete MFA\n"
                "3. WAIT for SharePoint site to load completely\n"
                "4. Verify you see the SharePoint site (NOT a login page)\n\n"
                "Only click 'Login Complete' after SharePoint loads!"
            )

            label = tk.Label(
                login_window,
                text=msg,
                bg=COLOR_BG_DIALOG,
                fg=COLOR_TEXT_GREY_DARK,
                justify="left",
                wraplength=340,
            )
            label.pack(padx=PADDING_STANDARD, pady=(PADDING_STANDARD, PADDING_SMALL))

            button_frame = tk.Frame(login_window, bg=COLOR_BG_DIALOG)
            button_frame.pack(pady=(0, 15))

            def mark_login_complete():
                flag_path = SCRIPT_DIR / "dashboard_mode_continue.flag"
                login_needed_flag = SCRIPT_DIR / "dashboard_mode_login_needed.flag"
                try:
                    if login_needed_flag.exists():
                        login_needed_flag.unlink()
                except Exception:
                    pass

                try:
                    with open(flag_path, "w", encoding="utf-8") as f:
                        f.write("ok")
                except Exception as e:
                    messagebox.showwarning("Warning", f"Failed to signal login completion: {e}")

                if self._login_window is not None and tk.Toplevel.winfo_exists(self._login_window):
                    self._login_window.destroy()
                self._login_window = None

            continue_btn = create_button(
                button_frame,
                text="Login Complete",
                command=mark_login_complete,
                bg=COLOR_BTN_SUCCESS,
                padx=PADDING_STANDARD,
                pady=6
            )
            continue_btn.pack(side="left", padx=PADDING_SMALL)

            def cancel_login():
                login_window.destroy()
                self._login_window = None
                self.load_data()  # Load from local files

            cancel_btn = create_button(
                button_frame,
                text="Cancel",
                command=cancel_login,
                bg=COLOR_BTN_DANGER,
                padx=PADDING_STANDARD,
                pady=6
            )
            cancel_btn.pack(side="left", padx=PADDING_SMALL)

            return login_window

        login_needed_path = SCRIPT_DIR / "dashboard_mode_login_needed.flag"

        def poll_login_needed():
            if login_needed_path.exists():
                ensure_login_window()
            # Continue polling while download is in progress
            self.root.after(500, poll_login_needed)

        poll_login_needed()
        
        def download_task():
            try:
                import sys
                import os
                sys.path.insert(0, str(SCRIPT_DIR))
                
                from code_review_updater import (download_file_from_sharepoint, _ensure_valid_session,
                                                should_download_file)
                from playwright.sync_api import sync_playwright
                from concurrent.futures import ThreadPoolExecutor, as_completed
                
                # Enable dashboard mode for authentication
                env = os.environ.copy()
                env["CODE_REVIEW_DASHBOARD_MODE"] = "1"
                os.environ["CODE_REVIEW_DASHBOARD_MODE"] = "1"
                
                with sync_playwright() as playwright:
                    # Validate sessions first
                    prod_session_path = Path(DEST_READ_CONFIG["session_file"])
                    _ensure_valid_session(playwright, prod_session_path, DEST_READ_CONFIG["site_url"])
                    
                    staging_session_path = Path(STAGING_READ_CONFIG["session_file"])
                    _ensure_valid_session(playwright, staging_session_path, STAGING_READ_CONFIG["site_url"])
                    
                    # Check which files need downloading
                    downloads_needed = []
                    
                    if should_download_file(playwright, prod_session_path, DEST_READ_CONFIG["site_url"],
                                          DEST_READ_CONFIG["file_path"], str(DEST_FILE)):
                        downloads_needed.append(("production", prod_session_path, DEST_READ_CONFIG["site_url"],
                                               DEST_READ_CONFIG["file_path"], str(DEST_FILE)))
                    
                    if should_download_file(playwright, staging_session_path, STAGING_READ_CONFIG["site_url"],
                                          STAGING_READ_CONFIG["file_path"], str(STAGING_FILE)):
                        downloads_needed.append(("staging", staging_session_path, STAGING_READ_CONFIG["site_url"],
                                               STAGING_READ_CONFIG["file_path"], str(STAGING_FILE)))
                    
                    if not downloads_needed:
                        print("All files are up to date - no downloads needed")
                        return True
                    
                    # Download in parallel
                    def download_one(file_type, session_path, site_url, file_path, output_path):
                        with sync_playwright() as pw:
                            return download_file_from_sharepoint(pw, session_path, site_url, file_path, output_path)
                    
                    results = []
                    with ThreadPoolExecutor(max_workers=2) as executor:
                        futures = [executor.submit(download_one, *args) for args in downloads_needed]
                        results = [f.result() for f in as_completed(futures)]
                    
                    return all(results)
                    
            except Exception as e:
                print(f"Error downloading files on startup: {e}")
                return False
            finally:
                # Clean up dashboard mode
                if "CODE_REVIEW_DASHBOARD_MODE" in os.environ:
                    del os.environ["CODE_REVIEW_DASHBOARD_MODE"]
        
        def on_complete():
            # Close any login dialog and clean up flags
            login_needed_flag = SCRIPT_DIR / "dashboard_mode_login_needed.flag"
            continue_flag = SCRIPT_DIR / "dashboard_mode_continue.flag"
            try:
                if hasattr(self, "_login_window") and self._login_window is not None:
                    if tk.Toplevel.winfo_exists(self._login_window):
                        self._login_window.destroy()
                    self._login_window = None
                if login_needed_flag.exists():
                    login_needed_flag.unlink()
                if continue_flag.exists():
                    continue_flag.unlink()
            except Exception:
                pass
            
            # Load from whichever file has more outstanding reviews
            self.load_data_from_best_source()
        
        def run_task():
            download_task()
            self.root.after(0, on_complete)
        
        threading.Thread(target=run_task, daemon=True).start()
    
    def refresh_data(self):
        """Run the full updater script, then reload data into the dashboard.

        This spawns code_review_updater.py as a background process so the UI
        remains responsive. After the script finishes, the Excel file is
        reloaded and the display is updated.
        """
        # Disable buttons during refresh
        self.refresh_btn.config(state="disabled", text="⭮ Updating from SharePoint...")
        self.cancelled_btn.config(state="disabled")
        self.other_btn.config(state="disabled")

        # Create progress window to show live output
        self.progress_window = tk.Toplevel(self.root)
        self.progress_window.title("Update Progress")
        self.progress_window.geometry(WINDOW_SIZE_PROGRESS)
        self.progress_window.configure(bg=COLOR_BG_DIALOG)
        
        # Header
        header_frame = tk.Frame(self.progress_window, bg=COLOR_BG_DIALOG)
        header_frame.pack(fill="x", padx=PADDING_STANDARD, pady=(PADDING_STANDARD, PADDING_SMALL))
        
        header_label = tk.Label(
            header_frame,
            text="🔄 Update in Progress",
            font=FONT_TITLE,
            bg=COLOR_BG_DIALOG,
            fg=COLOR_TEXT_BLUE
        )
        header_label.pack(anchor="w")
        
        self.progress_status_label = tk.Label(
            header_frame,
            text="Initializing...",
            font=FONT_BODY,
            bg=COLOR_BG_DIALOG,
            fg=COLOR_TEXT_GREY
        )
        self.progress_status_label.pack(anchor="w", pady=(5, 0))
        
        # Output text area
        text_frame = tk.Frame(self.progress_window, bg=COLOR_BG_DIALOG)
        text_frame.pack(fill="both", expand=True, padx=PADDING_STANDARD, pady=(PADDING_SMALL, PADDING_STANDARD))
        
        self.progress_text = scrolledtext.ScrolledText(
            text_frame,
            wrap=tk.WORD,
            font=FONT_TERMINAL,
            bg=COLOR_BG_TERMINAL,
            fg=COLOR_TEXT_TERMINAL,
            state=tk.DISABLED
        )
        self.progress_text.pack(fill="both", expand=True)
        
        # Close button (initially hidden, shown after completion)
        self.progress_close_btn = create_button(
            self.progress_window,
            text="Close",
            command=lambda: self.progress_window.destroy() if hasattr(self, 'progress_window') else None,
            padx=30,
            pady=8,
            state="disabled"
        )
        self.progress_close_btn.pack(pady=(0, PADDING_STANDARD))
        
        # Handle manual close during update
        def on_progress_window_close():
            if self._updater_process is not None:
                result = messagebox.askyesno(
                    "Cancel Update?",
                    "The update is still in progress. Do you want to cancel it?",
                    parent=self.progress_window
                )
                if result:
                    try:
                        self._updater_process.terminate()
                        self._updater_process.wait(timeout=2)
                    except:
                        try:
                            self._updater_process.kill()
                        except:
                            pass
                    self._updater_process = None
                    self.progress_window.destroy()
                    # Re-enable buttons
                    self.refresh_btn.config(state="normal", text="⭮ Refresh Data")
                    self.load_data()
                else:
                    return
            else:
                self.progress_window.destroy()
        
        self.progress_window.protocol("WM_DELETE_WINDOW", on_progress_window_close)

        # Show a small dialog that lets the user confirm login completion.
        # We keep a single dialog instance and re-use it for multiple login
        # steps within the same refresh by responding to small flag files
        # written by the updater.
        self._login_window = None
        self._updater_process = None

        def ensure_login_window():
            if self._login_window is not None and tk.Toplevel.winfo_exists(self._login_window):
                return self._login_window

            login_window = tk.Toplevel(self.root)
            self._login_window = login_window
            login_window.title("Complete SharePoint Login")
            login_window.geometry(WINDOW_SIZE_LOGIN)
            login_window.configure(bg=COLOR_BG_DIALOG)
            login_window.grab_set()

            msg = (
                "If a browser window opened for SharePoint,\n"
                "please complete the FULL login process:\n\n"
                "1. Choose your identity provider\n"
                "2. Enter credentials and complete MFA\n"
                "3. WAIT for SharePoint site to load completely\n"
                "4. Verify you see the SharePoint site (NOT a login page)\n\n"
                "Only click 'Login Complete' after SharePoint loads!"
            )

            label = tk.Label(
                login_window,
                text=msg,
                bg=COLOR_BG_DIALOG,
                fg=COLOR_TEXT_GREY_DARK,
                justify="left",
                wraplength=340,
            )
            label.pack(padx=PADDING_STANDARD, pady=(PADDING_STANDARD, PADDING_SMALL))

            button_frame = tk.Frame(login_window, bg=COLOR_BG_DIALOG)
            button_frame.pack(pady=(0, 15))

            def mark_login_complete():
                """Signal to the updater that login is complete by creating the flag file."""
                flag_path = SCRIPT_DIR / "dashboard_mode_continue.flag"
                login_needed_flag = SCRIPT_DIR / "dashboard_mode_login_needed.flag"
                try:
                    # Clear the "login needed" marker for this step
                    if login_needed_flag.exists():
                        login_needed_flag.unlink()
                except Exception:
                    pass

                try:
                    with open(flag_path, "w", encoding="utf-8") as f:
                        f.write("ok")
                except Exception as e:
                    messagebox.showwarning("Warning", f"Failed to signal login completion: {e}")

                # Close this dialog after signaling completion; if another
                # login is needed later in the workflow, the polling logic
                # will create a fresh dialog when it sees a new
                # dashboard_mode_login_needed.flag.
                if self._login_window is not None and tk.Toplevel.winfo_exists(self._login_window):
                    self._login_window.destroy()
                self._login_window = None

            continue_btn = create_button(
                button_frame,
                text="Login Complete",
                command=mark_login_complete,
                bg=COLOR_BTN_SUCCESS,
                padx=PADDING_STANDARD,
                pady=6
            )
            continue_btn.pack(side="left", padx=PADDING_SMALL)

            def cancel_login():
                """Cancel the login process and terminate the updater script."""
                # Terminate the updater subprocess if it's running
                if self._updater_process is not None:
                    try:
                        self._updater_process.terminate()
                        # Give it a moment to terminate gracefully
                        try:
                            self._updater_process.wait(timeout=2)
                        except subprocess.TimeoutExpired:
                            # Force kill if it doesn't terminate
                            self._updater_process.kill()
                            self._updater_process.wait()
                    except Exception as e:
                        print(f"Warning: Failed to terminate updater process: {e}")
                    finally:
                        self._updater_process = None
                
                # Clean up flag files
                login_needed_flag = SCRIPT_DIR / "dashboard_mode_login_needed.flag"
                continue_flag = SCRIPT_DIR / "dashboard_mode_continue.flag"
                try:
                    if login_needed_flag.exists():
                        login_needed_flag.unlink()
                    if continue_flag.exists():
                        continue_flag.unlink()
                except Exception as e:
                    print(f"Warning: Failed to clean up flag files: {e}")
                
                # Close the login window
                login_window.destroy()
                self._login_window = None
                
                # Reload data to restore UI to previous state
                self.load_data()
                
                # Reset refresh button
                self.refresh_btn.config(state="normal", text="⭮ Refresh Data")
                
                print("\n" + "="*60)
                print("✗ Update cancelled by user.")
                print("="*60 + "\n")

            cancel_btn = create_button(
                button_frame,
                text="Cancel",
                command=cancel_login,
                bg=COLOR_BTN_DANGER,
                padx=PADDING_STANDARD,
                pady=6
            )
            cancel_btn.pack(side="left", padx=PADDING_SMALL)

            return login_window

        # Poll for login requests signaled by the updater and ensure the
        # login window is visible whenever a request is active. We do not
        # show the dialog immediately; it only appears if the updater writes
        # the dashboard_mode_login_needed.flag (i.e., when a session is
        # missing or expired).
        login_needed_path = SCRIPT_DIR / "dashboard_mode_login_needed.flag"

        def poll_login_needed():
            if login_needed_path.exists():
                ensure_login_window()
            # Continue polling while the refresh is in progress
            self.root.after(500, poll_login_needed)

        poll_login_needed()

        def append_to_progress(text, autoscroll=True):
            """Append text to progress window and optionally auto-scroll."""
            if hasattr(self, 'progress_text') and self.progress_text.winfo_exists():
                self.progress_text.config(state=tk.NORMAL)
                self.progress_text.insert(tk.END, text)
                if autoscroll:
                    self.progress_text.see(tk.END)
                self.progress_text.config(state=tk.DISABLED)
        
        def update_progress_status(status):
            """Update the status label in the progress window."""
            if hasattr(self, 'progress_status_label') and self.progress_status_label.winfo_exists():
                self.progress_status_label.config(text=status)
        
        def run_updater():
            script_path = SCRIPT_DIR / "code_review_updater.py"
            env = os.environ.copy()
            env["CODE_REVIEW_DASHBOARD_MODE"] = "1"
            
            print("\n" + "="*60)
            print("Starting Code Review Updater...")
            print("="*60 + "\n")
            
            stdout_lines = []
            stderr_lines = []
            
            try:
                # Run subprocess with streaming output
                process = subprocess.Popen(
                    ["python", str(script_path)],
                    cwd=str(SCRIPT_DIR),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    env=env,
                    text=True,
                    bufsize=1,
                    universal_newlines=True
                )
                
                # Store process so cancel_login can terminate it
                self._updater_process = process
                
                # Stream output in real-time using thread-safe queue
                from queue import Queue, Empty
                output_queue = Queue()
                
                def read_stdout():
                    try:
                        for line in iter(process.stdout.readline, ''):
                            if line:
                                output_queue.put(('stdout', line))
                        process.stdout.close()
                    except Exception as e:
                        output_queue.put(('error', str(e)))
                
                def read_stderr():
                    try:
                        for line in iter(process.stderr.readline, ''):
                            if line:
                                output_queue.put(('stderr', line))
                        process.stderr.close()
                    except Exception as e:
                        output_queue.put(('error', str(e)))
                
                # Start reader threads
                stdout_thread = threading.Thread(target=read_stdout, daemon=True)
                stderr_thread = threading.Thread(target=read_stderr, daemon=True)
                stdout_thread.start()
                stderr_thread.start()
                
                # Process output as it arrives
                while process.poll() is None or not output_queue.empty():
                    try:
                        stream_type, line = output_queue.get(timeout=0.1)
                        
                        if stream_type == 'stdout':
                            stdout_lines.append(line)
                            print(line, end='')
                            self.root.after(0, lambda l=line: append_to_progress(l))
                            
                            # Update status based on output
                            if "STEP 1:" in line:
                                self.root.after(0, lambda: update_progress_status("Downloading files from SharePoint..."))
                            elif "STEP 2:" in line:
                                self.root.after(0, lambda: update_progress_status("Processing and transforming data..."))
                            elif "STEP 3:" in line:
                                self.root.after(0, lambda: update_progress_status("Checking for duplicates and uploading..."))
                            elif "WORKFLOW COMPLETED" in line:
                                self.root.after(0, lambda: update_progress_status("✓ Update completed successfully!"))
                        elif stream_type == 'stderr':
                            stderr_lines.append(line)
                            print(line, end='', file=sys.stderr)
                            self.root.after(0, lambda l=line: append_to_progress(f"[STDERR] {l}"))
                        
                    except Empty:
                        continue
                
                # Wait for threads to finish
                stdout_thread.join(timeout=1)
                stderr_thread.join(timeout=1)
                
                returncode = process.returncode
                success = returncode == 0
                stdout = ''.join(stdout_lines)
                stderr = ''.join(stderr_lines)
                
                # Clear the process reference after completion
                self._updater_process = None
                
            except Exception as e:
                success = False
                stdout = ''.join(stdout_lines)
                stderr = str(e)
                print(f"\nERROR: {stderr}\n")
                self.root.after(0, lambda: append_to_progress(f"\n[ERROR] {stderr}\n"))
                self.root.after(0, lambda: update_progress_status("✗ Update failed"))
                # Clear the process reference on error
                self._updater_process = None

            def on_complete():
                # Print completion message to terminal
                print("\n" + "="*60)
                if success:
                    print("✓ Update process completed successfully!")
                else:
                    print("✗ Update process failed.")
                print("="*60 + "\n")
                
                # Enable close button on progress window
                if hasattr(self, 'progress_close_btn') and self.progress_close_btn.winfo_exists():
                    self.progress_close_btn.config(state="normal", bg=COLOR_BTN_SUCCESS)
                
                # Re-download production file and reload dashboard
                if success:
                    self.refresh_btn.config(text="⭮ Downloading latest data...")
                    if hasattr(self, 'progress_status_label') and self.progress_status_label.winfo_exists():
                        self.progress_status_label.config(text="Downloading latest data from SharePoint...")
                    self._download_and_reload()
                else:
                    if hasattr(self, 'progress_status_label') and self.progress_status_label.winfo_exists():
                        self.progress_status_label.config(text="✗ Update failed - see details above")
                    # Parse and display errors
                    error_messages = self._parse_errors(stdout, stderr)
                    if error_messages:
                        self._show_error_dialog("Refresh Failed", error_messages)
                    else:
                        # Generic error if no specific errors found
                        messagebox.showerror(
                            "Refresh Failed",
                            "The refresh process failed. Check the terminal for details."
                        )
                    # Just reload local data if update failed
                    self.load_data()
                
                self.refresh_btn.config(state="normal", text="⭮ Refresh Data")

                # Close any remaining login dialog and clean up dashboard
                # flag files so that the next refresh starts cleanly.
                login_needed_flag = SCRIPT_DIR / "dashboard_mode_login_needed.flag"
                continue_flag = SCRIPT_DIR / "dashboard_mode_continue.flag"
                try:
                    if hasattr(self, "_login_window") and self._login_window is not None:
                        if tk.Toplevel.winfo_exists(self._login_window):
                            self._login_window.destroy()
                        self._login_window = None
                    if login_needed_flag.exists():
                        login_needed_flag.unlink()
                    if continue_flag.exists():
                        continue_flag.unlink()
                except Exception:
                    pass

            # Schedule UI update back on the main thread
            self.root.after(0, on_complete)

        threading.Thread(target=run_updater, daemon=True).start()
    
    def _download_and_reload(self):
        """Download both production and staging files from SharePoint and reload dashboard.
        
        This is called after the updater completes to ensure the dashboard shows
        whichever file has more outstanding reviews. This is important because:
        1. The user may have manually updated the production file after the updater ran
        2. The updater only uploads to staging, not production
        3. We want to show whichever file has the highest count
        """
        def download_and_load():
            try:
                import sys
                sys.path.insert(0, str(SCRIPT_DIR))
                
                try:
                    from code_review_updater import (download_file_from_sharepoint, _ensure_valid_session,
                                                    should_download_file)
                    from playwright.sync_api import sync_playwright
                    from concurrent.futures import ThreadPoolExecutor, as_completed
                    
                    with sync_playwright() as playwright:
                        # Validate sessions first
                        prod_session_path = Path(DEST_READ_CONFIG["session_file"])
                        _ensure_valid_session(playwright, prod_session_path, DEST_READ_CONFIG["site_url"])
                        
                        staging_session_path = Path(STAGING_READ_CONFIG["session_file"])
                        _ensure_valid_session(playwright, staging_session_path, STAGING_READ_CONFIG["site_url"])
                        
                        # Check which files need downloading
                        downloads_needed = []
                        
                        if should_download_file(playwright, prod_session_path, DEST_READ_CONFIG["site_url"],
                                              DEST_READ_CONFIG["file_path"], str(DEST_FILE)):
                            downloads_needed.append(("production", prod_session_path, DEST_READ_CONFIG["site_url"],
                                                   DEST_READ_CONFIG["file_path"], str(DEST_FILE)))
                        
                        if should_download_file(playwright, staging_session_path, STAGING_READ_CONFIG["site_url"],
                                              STAGING_READ_CONFIG["file_path"], str(STAGING_FILE)):
                            downloads_needed.append(("staging", staging_session_path, STAGING_READ_CONFIG["site_url"],
                                                   STAGING_READ_CONFIG["file_path"], str(STAGING_FILE)))
                        
                        if not downloads_needed:
                            print("All files are up to date - no downloads needed")
                            return True
                        
                        # Download in parallel
                        def download_one(file_type, session_path, site_url, file_path, output_path):
                            with sync_playwright() as pw:
                                return download_file_from_sharepoint(pw, session_path, site_url, file_path, output_path)
                        
                        results = []
                        with ThreadPoolExecutor(max_workers=2) as executor:
                            futures = [executor.submit(download_one, *args) for args in downloads_needed]
                            results = [f.result() for f in as_completed(futures)]
                        
                        return all(results)
                        
                except Exception as e:
                    print(f"Error downloading files: {e}")
                    return False
                    
            except Exception as e:
                print(f"Error in download_and_load: {e}")
                return False
        
        def on_complete():
            # Load from whichever file has more outstanding reviews
            self.load_data_from_best_source()
            
            # Update progress window status
            if hasattr(self, 'progress_status_label') and self.progress_status_label.winfo_exists():
                self.progress_status_label.config(text="✓ Refresh complete! Dashboard updated.")
            
            # Auto-close progress window after a short delay
            if hasattr(self, 'progress_window') and self.progress_window.winfo_exists():
                self.root.after(2000, lambda: self.progress_window.destroy() if hasattr(self, 'progress_window') and self.progress_window.winfo_exists() else None)
        
        # Run download in background and reload when done
        def run_task():
            download_and_load()
            self.root.after(0, on_complete)
        
        threading.Thread(target=run_task, daemon=True).start()
    
    def _parse_errors(self, stdout, stderr):
        """Parse error messages from subprocess output.
        
        Looks for common error patterns and extracts relevant error information.
        
        Args:
            stdout: Standard output from the updater
            stderr: Standard error from the updater
            
        Returns:
            list: List of error message strings, or empty list if no errors found
        """
        errors = []
        combined_output = (stdout or "") + "\n" + (stderr or "")
        
        # Look for common error patterns
        lines = combined_output.split('\n')
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            # Check for error indicators
            if any(indicator in line for indicator in ["Failed to", "ERROR", "✗", "error:", "Error:"]):
                # Collect this line and following context lines
                error_block = [line]
                
                # Look ahead for context (HTTP status, URL, additional details)
                j = i + 1
                while j < len(lines) and j < i + 10:  # Max 10 lines of context
                    next_line = lines[j].strip()
                    if not next_line:
                        break
                    # Include lines that look like error details
                    if any(keyword in next_line for keyword in ["HTTP Status:", "URL:", "Unauthorized", "Forbidden", "session may have expired", "Upload failed", "Download", "File", "may indicate", "This usually"]):
                        error_block.append(next_line)
                        j += 1
                    else:
                        break
                
                if error_block:
                    errors.append('\n'.join(error_block))
                
                i = j
            else:
                i += 1
        
        # Remove duplicates while preserving order
        seen = set()
        unique_errors = []
        for error in errors:
            if error not in seen:
                seen.add(error)
                unique_errors.append(error)
        
        return unique_errors
    
    def _show_error_dialog(self, title, error_messages):
        """Display error messages in a user-friendly dialog.
        
        Args:
            title: Dialog window title
            error_messages: List of error message strings to display
        """
        error_window = tk.Toplevel(self.root)
        error_window.title(title)
        error_window.geometry(WINDOW_SIZE_ERROR)
        error_window.configure(bg=COLOR_BG_CARD_RED)
        error_window.grab_set()  # Make dialog modal
        
        # Header
        header_frame = tk.Frame(error_window, bg=COLOR_BG_CARD_RED)
        header_frame.pack(fill="x", padx=PADDING_STANDARD, pady=(PADDING_STANDARD, PADDING_SMALL))
        
        header_label = tk.Label(
            header_frame,
            text="⚠ Errors occurred during refresh",
            font=FONT_TITLE,
            bg=COLOR_BG_CARD_RED,
            fg=COLOR_TEXT_RED
        )
        header_label.pack(anchor="w")
        
        subtitle_label = tk.Label(
            header_frame,
            text="The following errors were encountered:",
            font=FONT_BODY,
            bg=COLOR_BG_CARD_RED,
            fg=COLOR_TEXT_RED_DARK
        )
        subtitle_label.pack(anchor="w", pady=(5, 0))
        
        # Scrollable error text
        text_frame = tk.Frame(error_window, bg=COLOR_BG_CARD_RED)
        text_frame.pack(fill="both", expand=True, padx=PADDING_STANDARD, pady=(PADDING_SMALL, PADDING_STANDARD))
        
        text_widget = scrolledtext.ScrolledText(
            text_frame,
            wrap=tk.WORD,
            font=FONT_TERMINAL,
            bg="white",
            fg=COLOR_TEXT_GREY_DARK,
            relief="solid",
            borderwidth=1,
            padx=PADDING_SMALL,
            pady=10
        )
        text_widget.pack(fill="both", expand=True)
        
        # Add each error message
        for i, error_msg in enumerate(error_messages):
            if i > 0:
                text_widget.insert(tk.END, "\n" + "-"*70 + "\n\n")
            text_widget.insert(tk.END, error_msg)
        
        # Make read-only
        text_widget.config(state=tk.DISABLED)
        
        # Close button
        button_frame = tk.Frame(error_window, bg=COLOR_BG_CARD_RED)
        button_frame.pack(pady=(0, PADDING_STANDARD))
        
        close_btn = create_button(
            button_frame,
            text="Close",
            command=error_window.destroy,
            bg=COLOR_BTN_DANGER,
            padx=30,
            pady=8
        )
        close_btn.pack()
    
    def _show_output_window(self, title, stdout, stderr):
        """Display updater output in a scrollable window.
        
        Args:
            title: Window title
            stdout: Standard output from the updater
            stderr: Standard error from the updater
        """
        output_window = tk.Toplevel(self.root)
        output_window.title(title)
        output_window.geometry("800x600")
        output_window.configure(bg=COLOR_BG_DIALOG)
        
        # Create scrolled text widget
        text_frame = tk.Frame(output_window, bg=COLOR_BG_DIALOG)
        text_frame.pack(fill="both", expand=True, padx=PADDING_STANDARD, pady=PADDING_STANDARD)
        
        text_widget = scrolledtext.ScrolledText(
            text_frame,
            wrap=tk.WORD,
            font=FONT_TERMINAL,
            bg=COLOR_BG_TERMINAL,
            fg=COLOR_TEXT_TERMINAL,
            insertbackground="white"
        )
        text_widget.pack(fill="both", expand=True)
        
        # Add stdout
        if stdout:
            text_widget.insert(tk.END, "=== STANDARD OUTPUT ===\n\n", "header")
            text_widget.insert(tk.END, stdout)
            text_widget.insert(tk.END, "\n\n")
        
        # Add stderr if present
        if stderr:
            text_widget.insert(tk.END, "=== ERRORS/WARNINGS ===\n\n", "header")
            text_widget.insert(tk.END, stderr)
        
        # Configure tags for styling
        text_widget.tag_config("header", foreground=COLOR_TEXT_TERMINAL_HEADER, font=FONT_TERMINAL_HEADER)
        
        # Make read-only
        text_widget.config(state=tk.DISABLED)
        
        # Add close button
        close_btn = create_button(
            output_window,
            text="Close",
            command=output_window.destroy,
            font=FONT_BODY,
            padx=PADDING_STANDARD,
            pady=8
        )
        close_btn.pack(pady=(0, PADDING_STANDARD))
    
    def _count_total_rows_in_file(self, file_path):
        """Count total data rows in a file (excluding empty rows).
        
        Args:
            file_path: Path to the Excel file to check
            
        Returns:
            int: Total count of non-empty rows (or 0 if file doesn't exist/error)
        """
        if not file_path.exists():
            return 0
        
        rows = load_excel_sheet(file_path)
        if rows is None:
            return 0
        
        try:
            count = 0
            
            # Skip header rows and count all non-empty rows
            for row in rows[HEADER_ROW_COUNT:]:
                if not row or all(cell is None for cell in row):
                    continue
                
                repo = row[COL_REPO] if len(row) > COL_REPO else None
                if repo:  # Count any row with a repository
                    count += 1
            
            return count
            
        except Exception as e:
            print(f"Error counting total rows in {file_path}: {e}")
            return 0
    
    def _get_uploaded_review_numbers_from_production(self):
        """Get set of review numbers from production that have report uploaded filled.
        
        Returns:
            set: Set of review numbers (as strings) that have been marked as uploaded in production
        """
        if not DEST_FILE.exists():
            return set()
        
        rows = load_excel_sheet(DEST_FILE)
        if rows is None:
            return set()
        
        try:
            uploaded_review_nums = set()
            
            # Skip header rows
            for row in rows[HEADER_ROW_COUNT:]:
                if not row or all(cell is None for cell in row):
                    continue
                
                review_num = row[COL_REVIEW_NUM] if len(row) > COL_REVIEW_NUM else None
                report_uploaded = row[COL_REPORT_UPLOADED] if len(row) > COL_REPORT_UPLOADED else None
                
                # Add review number to set if it has been uploaded
                if review_num and is_report_uploaded(report_uploaded):
                    uploaded_review_nums.add(str(review_num).strip())
            
            return uploaded_review_nums
            
        except Exception as e:
            print(f"Error loading uploaded review numbers from production: {e}")
            return set()
    
    def _count_outstanding_in_file(self, file_path):
        """Count outstanding reviews in a file without loading it into the UI.
        
        Args:
            file_path: Path to the Excel file to check
            
        Returns:
            int: Count of outstanding reviews (or 0 if file doesn't exist/error)
        """
        if not file_path.exists():
            return 0
        
        rows = load_excel_sheet(file_path)
        if rows is None:
            return 0
        
        try:
            count = 0
            
            # Skip header rows
            for row in rows[HEADER_ROW_COUNT:]:
                if not row or all(cell is None for cell in row):
                    continue
                
                repo = row[COL_REPO] if len(row) > COL_REPO else None
                report_uploaded = row[COL_REPORT_UPLOADED] if len(row) > COL_REPORT_UPLOADED else None
                notes = row[COL_NOTES] if len(row) > COL_NOTES else None
                
                has_notes, _ = has_notes_content(notes)
                
                # Count only outstanding reviews (no upload, no notes)
                if repo and not is_report_uploaded(report_uploaded) and not has_notes:
                    if self.user_filter and not any(filter_term.lower() in repo.lower() for filter_term in self.user_filter):
                        continue
                    count += 1
            
            return count
            
        except Exception as e:
            print(f"Error counting reviews in {file_path}: {e}")
            return 0
    
    def load_data_from_best_source(self):
        """Load data from production file unless it has fewer total rows than staging.
        
        Workflow:
        1. Production is the source of truth by default
        2. Only use staging if production has FEWER total rows (missing data)
        3. If row counts are equal, production takes precedence
        
        Logic (compares TOTAL rows, not just outstanding):
        - Production >= Staging: Use production (authoritative source)
        - Production < Staging: Use staging (production missing data)
        """
        prod_total = self._count_total_rows_in_file(DEST_FILE)
        staging_total = self._count_total_rows_in_file(STAGING_FILE)
        
        print(f"Production file: {prod_total} total rows")
        print(f"Staging file: {staging_total} total rows")
        
        # Use production unless it has fewer TOTAL rows than staging
        if prod_total < staging_total:
            print("→ Loading from staging file (production has fewer total rows - missing data)")
            self.current_file_source = "staging"
            self._load_data_from_file(STAGING_FILE)
        else:
            # Use production if it has same or more total rows (source of truth)
            print("→ Loading from production file (source of truth)")
            self.current_file_source = "production"
            self._load_data_from_file(DEST_FILE)
    
    def load_data(self):
        """Load code review data from the production Excel file.
        
        This is a simple wrapper that loads from the production file.
        """
        self.current_file_source = "production"
        self._load_data_from_file(DEST_FILE)
    
    def _load_data_from_file(self, file_path):
        """Load code review data from the specified Excel file.
        
        Reads the Excel file and categorizes reviews as:
        - Outstanding: Report Uploaded is blank, Notes is empty
        - Cancelled/Broken: Report Uploaded is blank, Notes has content
        
        Applies user filters if configured.
        
        When loading from staging, cross-references against production to exclude
        reviews that have already been uploaded in production.
        
        Args:
            file_path: Path to the Excel file to load
        """
        if not file_path.exists():
            messagebox.showerror(
                "File Not Found",
                f"Could not find: {file_path}\n\n"
                "Please run first:\n"
                "python code_review_updater.py"
            )
            self.refresh_btn.config(state="normal", text="⭮ Reload Data")
            return
        
        rows = load_excel_sheet(file_path)
        if rows is None:
            messagebox.showerror("Error", f"{SHEET_NAME} sheet not found in workbook")
            return
        
        try:
            self.repo_data = {}
            self.cancelled_data = {}
            self.other_data = {}
            self.total_outstanding = 0
            self.total_cancelled = 0
            self.total_other = 0
            
            # If loading from staging, get uploaded review numbers from production
            # to avoid showing reviews that have already been synced
            uploaded_in_production = set()
            is_loading_staging = (file_path == STAGING_FILE)
            if is_loading_staging:
                uploaded_in_production = self._get_uploaded_review_numbers_from_production()
            
            # Skip header rows
            for row in rows[HEADER_ROW_COUNT:]:
                if not row or all(cell is None for cell in row):
                    continue
                
                # Extract cell values using column constants
                review_num = row[COL_REVIEW_NUM] if len(row) > COL_REVIEW_NUM else None
                date = row[COL_DATE] if len(row) > COL_DATE else None
                repo = row[COL_REPO] if len(row) > COL_REPO else None
                description = row[COL_DESCRIPTION] if len(row) > COL_DESCRIPTION else None
                report_uploaded = row[COL_REPORT_UPLOADED] if len(row) > COL_REPORT_UPLOADED else None
                notes = row[COL_NOTES] if len(row) > COL_NOTES else None
                
                # Cross-reference: If loading staging and this review is already uploaded in production, skip it
                if is_loading_staging and review_num:
                    review_num_str = str(review_num).strip()
                    if review_num_str in uploaded_in_production:
                        continue
                
                # Check if notes has content
                has_notes, notes_str = has_notes_content(notes)
                
                if repo and not is_report_uploaded(report_uploaded):
                    if self.user_filter and not any(filter_term.lower() in repo.lower() for filter_term in self.user_filter):
                        continue
                    
                    review_info = {
                        'reviewNum': str(review_num) if review_num is not None else 'N/A',
                        'date': format_date_only(date),
                        'description': str(description) if description is not None else 'N/A',
                        'notes': notes_str if has_notes else None
                    }
                    
                    # CUSTOMIZATION: Modify categorization logic for your workflow
                    # Categorize based on notes content
                    if has_notes:
                        # CUSTOMIZATION: Edit CANCELLED_KEYWORDS constant to change keywords
                        if is_cancelled_or_broken(notes_str):
                            if repo not in self.cancelled_data:
                                self.cancelled_data[repo] = []
                            self.cancelled_data[repo].append(review_info)
                            self.total_cancelled += 1
                        else:
                            # Other: Notes has content but not cancelled/pipeline
                            if repo not in self.other_data:
                                self.other_data[repo] = []
                            self.other_data[repo].append(review_info)
                            self.total_other += 1
                    else:
                        # Outstanding: No notes
                        if repo not in self.repo_data:
                            self.repo_data[repo] = []
                        self.repo_data[repo].append(review_info)
                        self.total_outstanding += 1
            
            self.update_display()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data:\n{str(e)}")
    
    def update_display(self):
        """Update the UI display with current data.
        
        Updates:
        - Summary card with color-coded count
        - Repository cards for outstanding reviews
        - Cancelled/Broken button state and text
        """
        # CUSTOMIZATION: Modify color thresholds and colors to match your preferences
        # Update summary with color-coding based on count thresholds
        count = self.total_outstanding
        
        # Check if all outstanding reviews are in a single repo
        repos_with_reviews = [repo for repo, reviews in self.repo_data.items() if len(reviews) > 0]
        single_repo_mode = len(repos_with_reviews) == 1
        
        # If all reviews are in one repo, use repo card color logic
        # Otherwise, use the default summary color logic
        if single_repo_mode:
            # Use repo card color thresholds (5+, 3+, 1+)
            badge_text_color, badge_bg_color = self._get_badge_colors(count)
            text_color = badge_text_color
            bg_color = badge_bg_color
            # Derive label color from text color (darker variant)
            if count >= THRESHOLD_REPO_HIGH:
                label_color = COLOR_TEXT_RED_DARK
            elif count >= THRESHOLD_REPO_MEDIUM:
                label_color = COLOR_TEXT_YELLOW_DARK
            elif count >= THRESHOLD_LOW:
                label_color = COLOR_TEXT_GREEN_DARK
            else:
                label_color = COLOR_TEXT_BLUE_DARK
        else:
            # Multiple repos: use default summary color logic
            # Color coding thresholds: red 10+, yellow 6-9, green 1-5, blue 0
            # CUSTOMIZATION: Adjust threshold constants at top of file to change when colors change
            if count >= THRESHOLD_HIGH:
                text_color = COLOR_TEXT_RED
                bg_color = COLOR_BG_CARD_RED
                label_color = COLOR_TEXT_RED_DARK
            elif count >= THRESHOLD_MEDIUM:
                text_color = COLOR_TEXT_YELLOW
                bg_color = COLOR_BG_CARD_YELLOW
                label_color = COLOR_TEXT_YELLOW_DARK
            elif count >= THRESHOLD_LOW:
                text_color = COLOR_TEXT_GREEN
                bg_color = COLOR_BG_CARD_GREEN
                label_color = COLOR_TEXT_GREEN_DARK
            else:
                text_color = COLOR_TEXT_BLUE
                bg_color = COLOR_BG_CARD_BLUE
                label_color = COLOR_TEXT_BLUE_DARK
        
        # Update summary card colors
        self.summary_card.config(bg=bg_color)
        self.total_label.config(text=str(count), fg=text_color, bg=bg_color)
        
        # Show which file source is being displayed
        source_label = "Production" if self.current_file_source == "production" else "Staging"
        self.file_info_label.config(
            text=f"Source: {source_label} | Updated: {datetime.now().strftime('%m-%d %H:%M')}",
            bg=bg_color,
            fg=text_color
        )
        
        filter_text = f"Filter: {', '.join(self.user_filter)}" if self.user_filter else "Filter: All repos"
        self.filter_info_label.config(
            text=filter_text,
            bg=bg_color,
            fg=text_color
        )
        
        # Update summary label and inner frame
        for widget in self.summary_card.winfo_children():
            widget.config(bg=bg_color)
            for child in widget.winfo_children():
                if isinstance(child, tk.Label) and child not in [self.total_label, self.file_info_label, self.filter_info_label]:
                    child.config(bg=bg_color, fg=label_color)
        
        # Clear existing repo cards
        for widget in self.scrollable_frame.winfo_children():
            if widget != self.empty_label:
                widget.destroy()
        
        # Update cancelled button state and display count in label (BEFORE early return!)
        if self.total_cancelled > 0:
            self.cancelled_btn.config(state="normal")
            self.cancelled_btn.config(text=f"🚫 Cancelled/Broken ({self.total_cancelled})")
        else:
            self.cancelled_btn.config(state="disabled")
            self.cancelled_btn.config(text="🚫 Cancelled/Broken")
        
        # Update other button state and display count in label (BEFORE early return!)
        if self.total_other > 0:
            self.other_btn.config(state="normal")
            self.other_btn.config(text=f"📋 Other ({self.total_other})")
        else:
            self.other_btn.config(state="disabled")
            self.other_btn.config(text="📋 Other")
        
        if self.total_outstanding == 0:
            self.empty_label.config(
                text="✓ No Outstanding Reports!\nAll reports have been uploaded."
            )
            self.empty_label.pack(pady=50)  # Vertical centering
            return
        
        self.empty_label.pack_forget()
        
        # Sort repos by count (descending)
        sorted_repos = sorted(self.repo_data.items(), key=lambda x: len(x[1]), reverse=True)  # Most reviews first
        
        # Create repo cards
        for repo, reviews in sorted_repos:
            self.create_repo_card(repo, reviews)
    
    def _create_card(self, parent_frame, repo, reviews, card_config):
        """Create an expandable card for repository reviews (shared helper).
        
        Args:
            parent_frame: Parent frame to attach card to
            repo: Repository name
            reviews: List of review dictionaries
            card_config: Dictionary with styling configuration:
                - bg_color: Background color
                - fg_color: Foreground color for text
                - icon: Icon emoji
                - count_text_template: Template for count text (with {count} placeholder)
                - badge_color_fn: Function to determine badge colors, or None for grey
                - expanded_set: Set tracking expanded state
                - expansion_key: Key for expanded state (repo or cancelled_repo)
                - build_details_fn: Function to build details
                - details_bg: Background color for details frame
        """
        count = len(reviews)
        bg_color = card_config['bg_color']
        fg_color = card_config['fg_color']
        icon = card_config['icon']
        expanded_set = card_config['expanded_set']
        expansion_key = card_config['expansion_key']
        
        # Main card frame
        card_frame = tk.Frame(parent_frame, bg=bg_color, relief="solid", bd=1)  # 1px border
        card_frame.pack(fill="x", pady=5)  # Spacing between cards
        
        # Header (clickable)
        header_frame = tk.Frame(card_frame, bg=bg_color, cursor="hand2")
        header_frame.pack(fill="x", padx=15, pady=12)  # Header padding
        
        # Icon
        icon_label = tk.Label(
            header_frame,
            text=icon,
            font=FONT_ICON,
            bg=bg_color
        )
        icon_label.pack(side="left", padx=(0, 10))  # Icon spacing
        
        # Repo name and count
        text_frame = tk.Frame(header_frame, bg=bg_color)
        text_frame.pack(side="left", fill="x", expand=True)
        
        repo_label = tk.Label(
            text_frame,
            text=repo,
            font=FONT_REPO_NAME,
            bg=bg_color,
            fg=fg_color
        )
        repo_label.pack(anchor="w")
        
        count_text = card_config['count_text_template'].format(count=count, s='s' if count != 1 else '')
        count_label = tk.Label(
            text_frame,
            text=count_text,
            font=FONT_BODY,
            bg=bg_color,
            fg=COLOR_TEXT_GREY_LIGHT if bg_color == COLOR_BG_DIALOG else COLOR_TEXT_GREY
        )
        count_label.pack(anchor="w")
        
        # Badge with color-coding or grey
        if card_config['badge_color_fn']:
            badge_color, badge_bg = card_config['badge_color_fn'](count)
        else:
            badge_color = COLOR_BTN_GREY
            badge_bg = COLOR_BG_GREY_LIGHT
        
        badge = tk.Label(
            header_frame,
            text=str(count),
            font=FONT_BADGE,
            bg=badge_bg,
            fg=badge_color,
            padx=12,
            pady=4
        )
        badge.pack(side="right", padx=(10, 5))  # Badge spacing
        
        # Arrow
        arrow_label = tk.Label(
            header_frame,
            text="▼" if expansion_key in expanded_set else "▶",
            font=FONT_BODY,
            bg=bg_color,
            fg=COLOR_TEXT_GREY_LIGHT
        )
        arrow_label.pack(side="right")
        
        # Details frame (initially hidden, lazy loaded)
        details_frame = tk.Frame(card_frame, bg=card_config['details_bg'])
        
        # Bind click event to toggle expansion with lazy loading of details
        def toggle_details(event=None):
            """Toggle the expansion of this card."""
            if expansion_key in expanded_set:
                expanded_set.remove(expansion_key)
                details_frame.pack_forget()
                arrow_label.config(text="▶")
            else:
                expanded_set.add(expansion_key)
                details_frame.pack(fill="both", padx=15, pady=(0, 12))  # Details padding
                arrow_label.config(text="▼")
                # Lazy load details only when expanded
                card_config['build_details_fn'](details_frame, reviews)
        
        # If already expanded, show and populate details
        if expansion_key in expanded_set:
            details_frame.pack(fill="both", padx=15, pady=(0, 12))
            card_config['build_details_fn'](details_frame, reviews)
        
        # Bind click events to all header elements
        for widget in [header_frame, icon_label, repo_label, count_label, badge, arrow_label]:
            widget.bind("<Button-1>", toggle_details)
    
    def _get_badge_colors(self, count):
        """Get badge colors based on count (for outstanding reviews).
        CUSTOMIZATION: Adjust threshold constants at top of file for repo card badges."""
        # CUSTOMIZATION: Edit THRESHOLD_REPO_* constants at top of file
        if count >= THRESHOLD_REPO_HIGH:
            return COLOR_TEXT_RED, COLOR_BG_CARD_RED
        elif count >= THRESHOLD_REPO_MEDIUM:
            return COLOR_TEXT_YELLOW, COLOR_BG_CARD_YELLOW
        elif count >= THRESHOLD_LOW:
            return COLOR_TEXT_GREEN, COLOR_BG_CARD_GREEN
        else:
            return COLOR_TEXT_BLUE, COLOR_BG_BLUE_LIGHT
    
    def create_repo_card(self, repo, reviews):
        """Create an expandable card for a repository's outstanding reviews.
        
        Args:
            repo: Repository name
            reviews: List of review dictionaries with reviewNum, date, description
        """
        card_config = {
            'bg_color': 'white',
            'fg_color': COLOR_TEXT_GREY_DARK,
            'icon': '📁',
            'count_text_template': '{count} outstanding report{s}',
            'badge_color_fn': self._get_badge_colors,
            'expanded_set': self.expanded_repos,
            'expansion_key': repo,
            'build_details_fn': self._build_details_for_repo,
            'details_bg': COLOR_BG_DIALOG
        }
        self._create_card(self.scrollable_frame, repo, reviews, card_config)
    
    def _build_review_details(self, details_frame, reviews, bg_color, fg_color_header, fg_color_desc, show_notes=False):
        """Build the details view for reviews (shared helper).
        
        Args:
            details_frame: The frame to populate with review details
            reviews: List of review dictionaries to display
            bg_color: Background color for review frames
            fg_color_header: Foreground color for header text
            fg_color_desc: Foreground color for description text
            show_notes: Whether to display notes field
        """
        # Clear any existing details
        for widget in details_frame.winfo_children():
            widget.destroy()
        
        # Get today's date and tomorrow's date for comparison
        today = datetime.now().strftime('%Y-%m-%d')
        tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
        
        # Add reviews to details
        for review in reviews:
            review_frame = tk.Frame(details_frame, bg=bg_color, relief="solid", bd=1)  # 1px border
            review_frame.pack(fill="x", pady=3)  # Spacing between reviews
            
            review_inner = tk.Frame(review_frame, bg=bg_color)
            review_inner.pack(fill="x", padx=10, pady=8)  # Review inner padding
            
            # Determine date color based on review date
            review_date = review['date']
            normalized_date = normalize_date_for_comparison(review_date)
            
            if normalized_date == today:
                date_color = COLOR_TEXT_RED
            elif normalized_date == tomorrow:
                date_color = COLOR_TEXT_ORANGE
            else:
                date_color = fg_color_header
            
            # Create header frame to hold review number and date with different colors
            header_frame = tk.Frame(review_inner, bg=bg_color)
            header_frame.pack(anchor="w")
            
            review_num_label = tk.Label(
                header_frame,
                text=f"Review #{review['reviewNum']} • ",
                font=FONT_BODY + ("bold",),
                bg=bg_color,
                fg=fg_color_header
            )
            review_num_label.pack(side="left")
            
            date_label = tk.Label(
                header_frame,
                text=review['date'],
                font=FONT_BODY + ("bold",),
                bg=bg_color,
                fg=date_color
            )
            date_label.pack(side="left")
            
            desc_label = tk.Label(
                review_inner,
                text=review['description'],
                font=FONT_SMALL,
                bg=bg_color,
                fg=fg_color_desc,
                wraplength=500,
                justify="left"
            )
            desc_label.pack(anchor="w", pady=(3, 0))  # Top padding only
            
            # Show notes if requested and available
            if show_notes and review.get('notes'):
                notes_label = tk.Label(
                    review_inner,
                    text=f"Notes: {review['notes']}",
                    font=FONT_SMALL_ITALIC,
                    bg=bg_color,
                    fg=fg_color_desc,
                    wraplength=500,
                    justify="left"
                )
                notes_label.pack(anchor="w", pady=(3, 0))  # Top padding only
    
    def _build_details_for_repo(self, details_frame, reviews):
        """Build the details view for a repo (lazy loaded).
        
        Args:
            details_frame: The frame to populate with review details
            reviews: List of review dictionaries to display
        """
        self._build_review_details(details_frame, reviews, "white", COLOR_TEXT_GREY_DARKER, COLOR_TEXT_GREY, show_notes=False)
    
    def show_cancelled_window(self):
        """Open a separate window to show cancelled/broken reviews.
        
        Creates a modal-style window with:
        - Header showing total cancelled count
        - Scrollable list of cancelled review cards
        - Close button
        
        These reviews have Report Uploaded blank AND Notes populated.
        """
        if self.total_cancelled == 0:
            return
        
        # Create new window
        cancelled_window = tk.Toplevel(self.root)
        cancelled_window.title("Cancelled/Broken Reviews")
        cancelled_window.geometry(WINDOW_SIZE_MAIN)
        cancelled_window.configure(bg=COLOR_BG_MAIN)
        
        # Header section
        header_frame = tk.Frame(cancelled_window, bg=COLOR_BG_DIALOG, relief="solid", bd=1)
        header_frame.pack(fill="x", padx=PADDING_STANDARD, pady=PADDING_STANDARD)
        
        header_inner = tk.Frame(header_frame, bg=COLOR_BG_DIALOG)
        header_inner.pack(fill="x", padx=PADDING_STANDARD, pady=15)
        
        title_label = tk.Label(
            header_inner,
            text="🚫 Cancelled/Broken Reviews",
            font=FONT_WINDOW_TITLE,
            bg=COLOR_BG_DIALOG,
            fg=COLOR_TEXT_GREY_DARK
        )
        title_label.pack(anchor="center")
        
        count_label = tk.Label(
            header_inner,
            text=f"{self.total_cancelled} {'review' if self.total_cancelled == 1 else 'reviews'}",
            font=FONT_MODAL_COUNT,
            bg=COLOR_BG_DIALOG,
            fg=COLOR_TEXT_GREY_DARK
        )
        count_label.pack(anchor="center", pady=(5, 0))
        
        note_label = tk.Label(
            header_inner,
            text="Not counted in outstanding total",
            font=FONT_BODY,
            bg=COLOR_BG_DIALOG,
            fg=COLOR_TEXT_GREY
        )
        note_label.pack(anchor="center", pady=(5, 0))  # Top padding only
        
        # Scrollable content area
        canvas_frame = tk.Frame(cancelled_window, bg=COLOR_BG_MAIN)
        canvas_frame.pack(fill="both", expand=True, padx=PADDING_STANDARD, pady=(0, PADDING_STANDARD))
        
        canvas = tk.Canvas(canvas_frame, bg=COLOR_BG_MAIN, highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        
        scrollable_frame = tk.Frame(canvas, bg=COLOR_BG_MAIN)
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")  # Always show scrollbar in popup
        
        # Bind mousewheel
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 100)), "units")  # Scroll speed
        
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Sort cancelled repos by count
        sorted_cancelled = sorted(self.cancelled_data.items(), key=lambda x: len(x[1]), reverse=True)  # Most reviews first
        
        # Create cards for cancelled repos
        for repo, reviews in sorted_cancelled:
            self.create_cancelled_card_in_window(scrollable_frame, repo, reviews)
        
        # Close button at bottom
        close_btn = create_button(
            cancelled_window,
            text="Close",
            command=cancelled_window.destroy,
            bg=COLOR_BTN_GREY,
            padx=30,
            pady=8
        )
        close_btn.pack(pady=(0, PADDING_STANDARD))
        
        # Cleanup mousewheel binding when window closes to avoid conflicts
        def on_close():
            canvas.unbind_all("<MouseWheel>")
            cancelled_window.destroy()
        
        cancelled_window.protocol("WM_DELETE_WINDOW", on_close)
    
    def create_cancelled_card_in_window(self, parent_frame, repo, reviews):
        """Create a card for cancelled/broken reviews in the popup window.
        
        Args:
            parent_frame: The parent frame to add the card to
            repo: Repository name
            reviews: List of cancelled review dictionaries
        """
        card_config = {
            'bg_color': COLOR_BG_DIALOG,
            'fg_color': COLOR_TEXT_GREY_DARK,
            'icon': '🚫',
            'count_text_template': '{count} cancelled/broken review{s}',
            'badge_color_fn': None,  # Use grey badge
            'expanded_set': self.expanded_cancelled,
            'expansion_key': f"cancelled_{repo}",
            'build_details_fn': self._build_cancelled_details,
            'details_bg': COLOR_BG_GREY_LIGHTER
        }
        self._create_card(parent_frame, repo, reviews, card_config)
    
    def _build_cancelled_details(self, details_frame, reviews):
        """Build the details view for cancelled reviews (lazy loaded).
        
        Args:
            details_frame: The frame to populate with review details
            reviews: List of cancelled review dictionaries to display
        """
        self._build_review_details(details_frame, reviews, COLOR_BG_DIALOG, COLOR_TEXT_GREY_DARK, COLOR_TEXT_GREY, show_notes=True)
    
    def show_other_window(self):
        """Open a separate window to show other reviews.
        
        Creates a modal-style window with:
        - Header showing total other count
        - Scrollable list of other review cards
        - Close button
        
        These reviews have Report Uploaded blank AND Notes with content (but not cancelled/broken).
        """
        if self.total_other == 0:
            return
        
        # Create new window
        other_window = tk.Toplevel(self.root)
        other_window.title("Other Reviews")
        other_window.geometry(WINDOW_SIZE_MAIN)
        other_window.configure(bg=COLOR_BG_MAIN)
        
        # Header section
        header_frame = tk.Frame(other_window, bg=COLOR_BG_DIALOG, relief="solid", bd=1)
        header_frame.pack(fill="x", padx=PADDING_STANDARD, pady=PADDING_STANDARD)
        
        header_inner = tk.Frame(header_frame, bg=COLOR_BG_DIALOG)
        header_inner.pack(fill="x", padx=PADDING_STANDARD, pady=15)
        
        title_label = tk.Label(
            header_inner,
            text="📋 Other Reviews",
            font=FONT_WINDOW_TITLE,
            bg=COLOR_BG_DIALOG,
            fg=COLOR_TEXT_GREY_DARK
        )
        title_label.pack(anchor="center")
        
        count_label = tk.Label(
            header_inner,
            text=f"{self.total_other} {'review' if self.total_other == 1 else 'reviews'}",
            font=FONT_MODAL_COUNT,
            bg=COLOR_BG_DIALOG,
            fg=COLOR_TEXT_GREY_DARK
        )
        count_label.pack(anchor="center", pady=(5, 0))
        
        note_label = tk.Label(
            header_inner,
            text="Not counted in outstanding total",
            font=FONT_BODY,
            bg=COLOR_BG_DIALOG,
            fg=COLOR_TEXT_GREY
        )
        note_label.pack(anchor="center", pady=(5, 0))  # Top padding only
        
        # Scrollable content area
        canvas_frame = tk.Frame(other_window, bg=COLOR_BG_MAIN)
        canvas_frame.pack(fill="both", expand=True, padx=PADDING_STANDARD, pady=(0, PADDING_STANDARD))
        
        canvas = tk.Canvas(canvas_frame, bg=COLOR_BG_MAIN, highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        
        scrollable_frame = tk.Frame(canvas, bg=COLOR_BG_MAIN)
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")  # Always show scrollbar in popup
        
        # Bind mousewheel
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 100)), "units")
        
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Sort other repos by count
        sorted_other = sorted(self.other_data.items(), key=lambda x: len(x[1]), reverse=True)
        
        # Create cards for other repos
        for repo, reviews in sorted_other:
            self.create_other_card_in_window(scrollable_frame, repo, reviews)
        
        # Close button at bottom
        close_btn = create_button(
            other_window,
            text="Close",
            command=other_window.destroy,
            bg=COLOR_BTN_GREY,
            padx=30,
            pady=8
        )
        close_btn.pack(pady=(0, PADDING_STANDARD))
        
        # Cleanup mousewheel binding when window closes to avoid conflicts
        def on_close():
            canvas.unbind_all("<MouseWheel>")
            other_window.destroy()
        
        other_window.protocol("WM_DELETE_WINDOW", on_close)
    
    def create_other_card_in_window(self, parent_frame, repo, reviews):
        """Create a card for other reviews in the popup window.
        
        Args:
            parent_frame: The parent frame to add the card to
            repo: Repository name
            reviews: List of other review dictionaries
        """
        card_config = {
            'bg_color': COLOR_BG_DIALOG,
            'fg_color': COLOR_TEXT_GREY_DARK,
            'icon': '📋',
            'count_text_template': '{count} other review{s}',
            'badge_color_fn': None,  # Use grey badge
            'expanded_set': self.expanded_other,
            'expansion_key': f"other_{repo}",
            'build_details_fn': self._build_other_details,
            'details_bg': COLOR_BG_GREY_LIGHTER
        }
        self._create_card(parent_frame, repo, reviews, card_config)
    
    def _build_other_details(self, details_frame, reviews):
        """Build the details view for other reviews (lazy loaded).
        
        Args:
            details_frame: The frame to populate with review details
            reviews: List of other review dictionaries to display
        """
        self._build_review_details(details_frame, reviews, COLOR_BG_DIALOG, COLOR_TEXT_GREY_DARK, COLOR_TEXT_GREY, show_notes=True)

def main():
    root = tk.Tk()
    app = DashboardApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
