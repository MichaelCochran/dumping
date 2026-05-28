#!/usr/bin/env python3
"""
Code Review Dashboard - Tkinter App
Shows outstanding reports from destination_current.xlsx
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
from pathlib import Path
import openpyxl
from datetime import datetime
import json
import subprocess
import threading
import sys

# ============================================================================
# CONFIGURATION
# ============================================================================
# CUSTOMIZATION: Update file paths if your data files are in different locations

SCRIPT_DIR = Path(__file__).parent
# CUSTOMIZATION: Path to the Excel file containing review data
DEST_FILE = SCRIPT_DIR / "destination_current.xlsx"  # Main data source file
# CUSTOMIZATION: Path to the JSON file containing repository filter settings
FILTER_CONFIG = SCRIPT_DIR / "user_filters.json"  # User filter configuration
# CUSTOMIZATION: Local updater script to generate the Excel file
UPDATER_SCRIPT = SCRIPT_DIR / "code_review_local.py"

def format_date_only(value):
    """Extract date without time from various formats"""
    if value is None:
        return 'N/A'
    
    # If it's already a datetime object, format it
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    
    # Convert to string and remove time portion
    value_str = str(value)
    
    # Remove time if present (e.g., "2026-05-06 00:00:00" -> "2026-05-06")
    if ' ' in value_str:
        value_str = value_str.split(' ')[0]
    
    return value_str if value_str else 'N/A'

class DashboardApp:
    """Main dashboard application for displaying code review status.
    
    Displays outstanding reviews by repository and provides access to cancelled/broken reviews.
    Reviews are loaded from destination_current.xlsx and can be filtered by repository.
    """
    
    def __init__(self, root):
        """Initialize the dashboard application.
        
        Args:
            root: The tkinter root window
        """
        self.root = root
        # CUSTOMIZATION: Change window title as desired
        self.root.title("Code Review Dashboard - Local")
        # CUSTOMIZATION: Adjust window size (width x height)
        self.root.geometry("400x600")  # Main window dimensions
        # CUSTOMIZATION: Change background color (hex format)
        self.root.configure(bg="#f0f0f0")  # Light grey background
        
        # Data structures for storing review information
        self.repo_data = {}  # Outstanding reviews grouped by repository
        self.cancelled_data = {}  # Cancelled/broken reviews grouped by repository
        self.other_data = {}  # Other reviews (notes present but not cancelled/broken)
        self.total_outstanding = 0  # Count of actionable outstanding reviews
        self.total_cancelled = 0  # Count of cancelled/broken reviews
        self.total_other = 0  # Count of other reviews
        
        # Track which repo cards are expanded in the UI
        self.expanded_repos = set()  # Expanded outstanding repo cards
        self.expanded_cancelled = set()  # Expanded cancelled repo cards in popup
        self.expanded_other = set()  # Expanded other repo cards in popup
        
        # Load repository filter from config file
        self.user_filter = self.load_user_filter()
        
        # Build UI and load data
        self.create_widgets()
        self.load_data()
    
    def create_widgets(self):
        """Create all UI widgets for the main dashboard window.
        
        Layout includes:
        - Summary card showing total outstanding count with color coding
        - Toolbar with Refresh and Cancelled/Broken buttons
        - Scrollable list of repository cards
        """
        # Summary Frame - top section with total count
        summary_frame = tk.Frame(self.root, bg="#f0f0f0")
        summary_frame.pack(fill="x", padx=20, pady=(20, 10))
        
        # Summary Card
        self.summary_card = tk.Frame(summary_frame, bg="#dbeafe", relief="solid", bd=1)  # Blue background, solid border
        self.summary_card.pack(fill="x")
        
        summary_inner = tk.Frame(self.summary_card, bg="#dbeafe")
        summary_inner.pack(fill="x", padx=20, pady=15)  # Inner padding
        
        summary_label = tk.Label(
            summary_inner,
            text="Total Outstanding Reports",
            font=("Arial", 12), # Summary label font
            bg="#dbeafe",
            fg="#1e40af"
        )
        summary_label.pack(anchor="center")
        
        self.total_label = tk.Label(
            summary_inner,
            text="0",
            font=("Arial", 36, "bold"), # Total label font
            bg="#dbeafe",
            fg="#2563eb"
        )
        self.total_label.pack(anchor="center")
        
        self.file_info_label = tk.Label(
            summary_inner,
            text="",
            font=("Arial", 9), # File info font
            bg="#dbeafe",
            fg="#60a5fa"
        )
        self.file_info_label.pack(anchor="center", pady=(5, 0))  # Top padding only
        
        self.filter_info_label = tk.Label(
            summary_inner,
            text="",
            font=("Arial", 8), # Filter info font
            bg="#dbeafe",
            fg="#60a5fa"
        )
        self.filter_info_label.pack(anchor="center", pady=(2, 0))  # Minimal top padding
        
        # Toolbar Frame - contains action buttons
        toolbar_frame = tk.Frame(self.root, bg="#f0f0f0")
        toolbar_frame.pack(fill="x", padx=20, pady=(0, 10))
        
        # First row - Refresh button (full width)
        refresh_container = tk.Frame(toolbar_frame, bg="#f0f0f0")
        refresh_container.pack(fill="x", pady=(0, 5))
        
        self.refresh_btn = tk.Button(
            refresh_container,
            text="⭮ Refresh Data",
            command=self.refresh_data,
            font=("Arial", 11), # Refresh button font
            bg="#3b82f6",
            fg="white",
            relief="flat",
            cursor="hand2",
            pady=8
        )
        self.refresh_btn.pack(fill="x")
        
        # Second row - Cancelled/Broken and Other buttons (split width)
        button_row = tk.Frame(toolbar_frame, bg="#f0f0f0")
        button_row.pack(fill="x")
        
        self.cancelled_btn = tk.Button(
            button_row,
            text="🚫 Cancelled/Broken",
            command=self.show_cancelled_window,
            font=("Arial", 11), # Cancelled button font
            bg="#6b7280",
            fg="white",
            disabledforeground="#9ca3af",  # Light gray text when disabled
            relief="flat",
            cursor="hand2",
            pady=8,
            state="disabled"  # Initially disabled until data loads
        )
        self.cancelled_btn.pack(side="left", fill="x", expand=True, padx=(0, 2.5))
        
        self.other_btn = tk.Button(
            button_row,
            text="📋 Other",
            command=self.show_other_window,
            font=("Arial", 11),
            bg="#6b7280",
            fg="white",
            disabledforeground="#9ca3af",  # Light gray text when disabled
            relief="flat",
            cursor="hand2",
            pady=8,
            state="disabled"  # Initially disabled until data loads
        )
        self.other_btn.pack(side="left", fill="x", expand=True, padx=(2.5, 0))
        
        # Canvas with Scrollbar for Repository List - main scrollable content area
        canvas_frame = tk.Frame(self.root, bg="#f0f0f0")
        canvas_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        self.canvas = tk.Canvas(canvas_frame, bg="#f0f0f0", highlightthickness=0)  # No border highlight
        self.scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.canvas.yview)
        
        self.scrollable_frame = tk.Frame(self.canvas, bg="#f0f0f0")
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self._on_frame_configure()
        )
        
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(self.canvas_window, width=e.width))
        
        self.canvas.pack(side="left", fill="both", expand=True)
        
        # Bind mousewheel
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        
        # Empty state label
        self.empty_label = tk.Label(
            self.scrollable_frame,
            text="",
            font=("Arial", 14), # Empty label font
            bg="#f0f0f0",
            fg="#6b7280"
        )
        self.empty_label.pack(pady=50)  # Vertical centering
    
    def _on_frame_configure(self):
        """Update scroll region and show/hide scrollbar as needed.
        
        Dynamically shows scrollbar only when content exceeds visible height.
        """
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        
        # Show scrollbar only if content exceeds canvas height
        canvas_height = self.canvas.winfo_height()
        content_height = self.scrollable_frame.winfo_reqheight()
        
        if content_height > canvas_height:
            self.scrollbar.pack(side="right", fill="y")
        else:
            self.scrollbar.pack_forget()
    
    def _on_mousewheel(self, event):
        """Handle mousewheel scrolling for the main canvas."""
        self.canvas.yview_scroll(int(-1 * (event.delta / 100)), "units")  # Scroll speed
    
    def load_user_filter(self):
        """Load repository filter from config file.
        
        Returns:
            list: Repository names/patterns to filter, or empty list for all repos
        """
        if not FILTER_CONFIG.exists():
            return []
        
        try:
            with open(FILTER_CONFIG, 'r') as f:
                config = json.load(f)
            
            repos = config.get('repos', [])
            return repos if repos else []
        
        except Exception as e:
            print(f"Error loading filter config: {e}")
            return []
    
    def refresh_data(self):
        """Run code review updater then load data.
        
        Executes code_review_local.py in background thread,
        then loads the updated Excel data.
        """
        # Disable buttons during refresh
        self.refresh_btn.config(state="disabled", text="⭮ Updating...")
        self.cancelled_btn.config(state="disabled")
        self.other_btn.config(state="disabled")
        
        # Update UI to show progress
        self.empty_label.config(text="Running code review updater...\nThis may take a moment.")
        self.empty_label.pack(pady=50)
        
        def run_updater():
            """Run the updater script in a subprocess."""
            try:
                # Run the code_review_local.py script
                result = subprocess.run(
                    [sys.executable, str(UPDATER_SCRIPT)],
                    capture_output=True,
                    text=True,
                    cwd=str(SCRIPT_DIR)
                )
                
                # Schedule UI update on main thread
                self.root.after(0, lambda: self.on_updater_complete(result))
                
            except Exception as e:
                # Schedule error handling on main thread
                self.root.after(0, lambda: self.on_updater_error(str(e)))
        
        # Run updater in background thread
        thread = threading.Thread(target=run_updater, daemon=True)
        thread.start()
    
    def on_updater_complete(self, result):
        """Handle completion of updater script.
        
        Args:
            result: subprocess.CompletedProcess result
        """
        if result.returncode != 0:
            error_msg = result.stderr if result.stderr else "Unknown error occurred"
            messagebox.showerror(
                "Updater Failed",
                f"Code review updater failed:\n\n{error_msg[:500]}"
            )
        
        # Re-enable buttons and load data
        self.refresh_btn.config(state="normal", text="⭮ Refresh Data")
        self.load_data()
    
    def on_updater_error(self, error):
        """Handle error running updater script.
        
        Args:
            error: Error message string
        """
        messagebox.showerror(
            "Error",
            f"Failed to run code review updater:\n\n{error}"
        )
        self.refresh_btn.config(state="normal", text="⭮ Refresh Data")
        # Reload data to properly update button states
        self.load_data()
    
    def load_data(self):
        """Load code review data from the Excel file.
        
        Reads destination_current.xlsx and categorizes reviews as:
        - Outstanding: Report Uploaded is blank, Notes is empty
        - Cancelled/Broken: Report Uploaded is blank, Notes has content
        
        Applies user filters if configured.
        """
        if not DEST_FILE.exists():
            messagebox.showerror(
                "File Not Found",
                f"Could not find: {DEST_FILE}\n\n"
                "Please run first:\n"
                "python code_review_local.py"
            )
            self.refresh_btn.config(state="normal", text="⭮ Refresh Data")
            return
        
        try:
            wb = openpyxl.load_workbook(DEST_FILE, read_only=True, data_only=True)
            
            if 'Code' not in wb.sheetnames:
                messagebox.showerror("Error", "Code sheet not found in workbook")
                wb.close()
                return
            
            ws = wb['Code']
            rows = list(ws.iter_rows(values_only=True))
            
            self.repo_data = {}
            self.cancelled_data = {}
            self.other_data = {}
            self.total_outstanding = 0
            self.total_cancelled = 0
            self.total_other = 0
            
            # Skip header rows (first 3 rows typically contain headers)
            for row in rows[3:]:  # Start at row index 3
                if not row or all(cell is None for cell in row):
                    continue
                
                review_num = row[1] if len(row) > 1 else None  # Column B - Review Number
                date = row[2] if len(row) > 2 else None  # Column C - Date
                repo = row[3] if len(row) > 3 else None  # Column D - Repository
                description = row[4] if len(row) > 4 else None  # Column E - Description
                report_uploaded = row[5] if len(row) > 5 else None  # Column F - Report Uploaded status
                notes = row[8] if len(row) > 8 else None  # Column I - Notes
                
                # Check if report_uploaded is truly empty (None, empty string, whitespace, or 'None')
                is_uploaded = report_uploaded and str(report_uploaded).strip() and str(report_uploaded).strip().lower() != 'none'
                
                # Check if column contains 'no'
                if report_uploaded and 'no' in str(report_uploaded).strip().lower():
                    is_uploaded = False
                
                # Check if notes has content
                notes_str = str(notes).strip() if notes else ""
                has_notes = bool(notes_str)
                
                if repo and not is_uploaded:
                    if self.user_filter and not any(filter_term.lower() in repo.lower() for filter_term in self.user_filter):
                        continue
                    
                    review_info = {
                        'reviewNum': str(review_num) if review_num is not None else 'N/A',
                        'date': format_date_only(date),
                        'description': str(description) if description is not None else 'N/A',
                        'notes': notes_str if has_notes else None
                    }
                    
                    # Categorize based on notes content (match super dashboard logic)
                    if has_notes:
                        notes_lower = notes_str.lower()
                        # Cancelled/Broken: Notes contains "cancelled" or "pipeline"
                        if "cancelled" in notes_lower or "pipeline" in notes_lower:
                            if repo not in self.cancelled_data:
                                self.cancelled_data[repo] = []
                            self.cancelled_data[repo].append(review_info)
                            self.total_cancelled += 1
                        else:
                            # Other: Notes has content but not cancelled/pipeline
                            if repo not in self.other_data:
                                self.other_data[repo] = []
                            self.other_data[repo].append(review_info)
                            self.total_other += 1
                    else:
                        # Outstanding: No notes
                        if repo not in self.repo_data:
                            self.repo_data[repo] = []
                        self.repo_data[repo].append(review_info)
                        self.total_outstanding += 1
            
            wb.close()
            
            self.update_display()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data:\n{str(e)}")
    
    def update_display(self):
        """Update the UI display with current data.
        
        Updates:
        - Summary card with color-coded count
        - Repository cards for outstanding reviews
        - Cancelled/Broken button state and text
        """
        # Update summary with color-coding based on count thresholds
        count = self.total_outstanding
        
        # Color coding thresholds: red 10+, yellow 6-9, green 1-5, blue 0
        if count >= 10:  # Red threshold
            text_color = "#dc2626"  # Red
            bg_color = "#fef2f2"
            label_color = "#991b1b"
        elif count >= 6:  # Yellow threshold
            text_color = "#eab308"  # Yellow
            bg_color = "#fefce8"
            label_color = "#854d0e"
        elif count >= 1:  # Green threshold
            text_color = "#16a34a"  # Green
            bg_color = "#f0fdf4"
            label_color = "#166534"
        else:  # Blue threshold (0 items)
            text_color = "#2563eb"  # Blue
            bg_color = "#dbeafe"
            label_color = "#1e40af"
        
        # Update summary card colors
        self.summary_card.config(bg=bg_color)
        self.total_label.config(text=str(count), fg=text_color, bg=bg_color)
        self.file_info_label.config(
            text=f"Data Updated: {datetime.now().strftime('%m-%d %H:%M')}",
            bg=bg_color,
            fg=text_color
        )
        
        filter_text = f"Filter: {', '.join(self.user_filter)}" if self.user_filter else "Filter: All repos"
        self.filter_info_label.config(
            text=filter_text,
            bg=bg_color,
            fg=text_color
        )
        
        # Update summary label and inner frame
        for widget in self.summary_card.winfo_children():
            widget.config(bg=bg_color)
            for child in widget.winfo_children():
                if isinstance(child, tk.Label) and child not in [self.total_label, self.file_info_label, self.filter_info_label]:
                    child.config(bg=bg_color, fg=label_color)
        
        # Clear existing repo cards
        for widget in self.scrollable_frame.winfo_children():
            if widget != self.empty_label:
                widget.destroy()
        
        # Update cancelled button state and display count in label (BEFORE early return!)
        if self.total_cancelled > 0:
            self.cancelled_btn.config(state="normal")
            self.cancelled_btn.config(text=f"🚫 Cancelled/Broken ({self.total_cancelled})")
        else:
            self.cancelled_btn.config(state="disabled")
            self.cancelled_btn.config(text="🚫 Cancelled/Broken")
        
        # Update other button state and display count in label (BEFORE early return!)
        if self.total_other > 0:
            self.other_btn.config(state="normal")
            self.other_btn.config(text=f"📋 Other ({self.total_other})")
        else:
            self.other_btn.config(state="disabled")
            self.other_btn.config(text="📋 Other")
        
        if self.total_outstanding == 0:
            self.empty_label.config(
                text="✓ No Outstanding Reports!\nAll reports have been uploaded."
            )
            self.empty_label.pack(pady=50)  # Vertical centering
            return
        
        self.empty_label.pack_forget()
        
        # Sort repos by count (descending)
        sorted_repos = sorted(self.repo_data.items(), key=lambda x: len(x[1]), reverse=True)  # Most reviews first
        
        # Create repo cards
        for repo, reviews in sorted_repos:
            self.create_repo_card(repo, reviews)
    
    def _create_card(self, parent_frame, repo, reviews, card_config):
        """Create an expandable card for repository reviews (shared helper).
        
        Args:
            parent_frame: Parent frame to attach card to
            repo: Repository name
            reviews: List of review dictionaries
            card_config: Dictionary with styling configuration:
                - bg_color: Background color
                - fg_color: Foreground color for text
                - icon: Icon emoji
                - count_text_template: Template for count text (with {count} placeholder)
                - badge_color_fn: Function to determine badge colors, or None for grey
                - expanded_set: Set tracking expanded state
                - expansion_key: Key for expanded state (repo or cancelled_repo)
                - build_details_fn: Function to build details
                - details_bg: Background color for details frame
        """
        count = len(reviews)
        bg_color = card_config['bg_color']
        fg_color = card_config['fg_color']
        icon = card_config['icon']
        expanded_set = card_config['expanded_set']
        expansion_key = card_config['expansion_key']
        
        # Main card frame
        card_frame = tk.Frame(parent_frame, bg=bg_color, relief="solid", bd=1)  # 1px border
        card_frame.pack(fill="x", pady=5)  # Spacing between cards
        
        # Header (clickable)
        header_frame = tk.Frame(card_frame, bg=bg_color, cursor="hand2")
        header_frame.pack(fill="x", padx=15, pady=12)  # Header padding
        
        # Icon
        icon_label = tk.Label(
            header_frame,
            text=icon,
            font=("Arial", 20), # Icon font
            bg=bg_color
        )
        icon_label.pack(side="left", padx=(0, 10))  # Icon spacing
        
        # Repo name and count
        text_frame = tk.Frame(header_frame, bg=bg_color)
        text_frame.pack(side="left", fill="x", expand=True)
        
        repo_label = tk.Label(
            text_frame,
            text=repo,
            font=("Arial", 13, "bold"), # Repo label font
            bg=bg_color,
            fg=fg_color
        )
        repo_label.pack(anchor="w")
        
        count_text = card_config['count_text_template'].format(count=count, s='s' if count != 1 else '')
        count_label = tk.Label(
            text_frame,
            text=count_text,
            font=("Arial", 10), # Count label font
            bg=bg_color,
            fg="#9ca3af" if bg_color == "#f9fafb" else "#6b7280"
        )
        count_label.pack(anchor="w")
        
        # Badge with color-coding or grey
        if card_config['badge_color_fn']:
            badge_color, badge_bg = card_config['badge_color_fn'](count)
        else:
            badge_color = "#6b7280"
            badge_bg = "#e5e7eb"
        
        badge = tk.Label(
            header_frame,
            text=str(count),
            font=("Arial", 12, "bold"), # Badge font
            bg=badge_bg,
            fg=badge_color,
            padx=12,
            pady=4
        )
        badge.pack(side="right", padx=(10, 5))  # Badge spacing
        
        # Arrow
        arrow_label = tk.Label(
            header_frame,
            text="▼" if expansion_key in expanded_set else "▶",
            font=("Arial", 10), # Arrow font
            bg=bg_color,
            fg="#9ca3af"
        )
        arrow_label.pack(side="right")
        
        # Details frame (initially hidden, lazy loaded)
        details_frame = tk.Frame(card_frame, bg=card_config['details_bg'])
        
        # Bind click event to toggle expansion with lazy loading of details
        def toggle_details(event=None):
            """Toggle the expansion of this card."""
            if expansion_key in expanded_set:
                expanded_set.remove(expansion_key)
                details_frame.pack_forget()
                arrow_label.config(text="▶")
            else:
                expanded_set.add(expansion_key)
                details_frame.pack(fill="both", padx=15, pady=(0, 12))  # Details padding
                arrow_label.config(text="▼")
                # Lazy load details only when expanded
                card_config['build_details_fn'](details_frame, reviews)
        
        # If already expanded, show and populate details
        if expansion_key in expanded_set:
            details_frame.pack(fill="both", padx=15, pady=(0, 12))
            card_config['build_details_fn'](details_frame, reviews)
        
        # Bind click events to all header elements
        for widget in [header_frame, icon_label, repo_label, count_label, badge, arrow_label]:
            widget.bind("<Button-1>", toggle_details)
    
    def _get_badge_colors(self, count):
        """Get badge colors based on count (for outstanding reviews)."""
        if count >= 5:  # Red badge threshold
            return "#dc2626", "#fef2f2"
        elif count >= 3:  # Yellow badge threshold
            return "#eab308", "#fefce8"
        elif count >= 1:  # Green badge threshold
            return "#16a34a", "#f0fdf4"
        else:  # Blue badge threshold (0 items)
            return "#2563eb", "#eff6ff"
    
    def create_repo_card(self, repo, reviews):
        """Create an expandable card for a repository's outstanding reviews.
        
        Args:
            repo: Repository name
            reviews: List of review dictionaries with reviewNum, date, description
        """
        card_config = {
            'bg_color': 'white',
            'fg_color': '#1f2937',
            'icon': '📁',
            'count_text_template': '{count} outstanding report{s}',
            'badge_color_fn': self._get_badge_colors,
            'expanded_set': self.expanded_repos,
            'expansion_key': repo,
            'build_details_fn': self._build_details_for_repo,
            'details_bg': '#f9fafb'
        }
        self._create_card(self.scrollable_frame, repo, reviews, card_config)
    
    def _build_review_details(self, details_frame, reviews, bg_color, fg_color_header, fg_color_desc, show_notes=False):
        """Build the details view for reviews (shared helper).
        
        Args:
            details_frame: The frame to populate with review details
            reviews: List of review dictionaries to display
            bg_color: Background color for review frames
            fg_color_header: Foreground color for header text
            fg_color_desc: Foreground color for description text
        """
        # Clear any existing details
        for widget in details_frame.winfo_children():
            widget.destroy()
        
        # Add reviews to details
        for review in reviews:
            review_frame = tk.Frame(details_frame, bg=bg_color, relief="solid", bd=1)  # 1px border
            review_frame.pack(fill="x", pady=3)  # Spacing between reviews
            
            review_inner = tk.Frame(review_frame, bg=bg_color)
            review_inner.pack(fill="x", padx=10, pady=8)  # Review inner padding
            
            review_header = tk.Label(
                review_inner,
                text=f"Review #{review['reviewNum']} • {review['date']}",
                font=("Arial", 10, "bold"), # Review header font
                bg=bg_color,
                fg=fg_color_header
            )
            review_header.pack(anchor="w")
            
            desc_label = tk.Label(
                review_inner,
                text=review['description'],
                font=("Arial", 9), # Description font
                bg=bg_color,
                fg=fg_color_desc,
                wraplength=500,  # Max width before text wraps
                justify="left"
            )
            desc_label.pack(anchor="w", pady=(3, 0))  # Top padding only

            # Optionally show notes if requested and present
            if show_notes and review.get('notes'):
                notes_label = tk.Label(
                    review_inner,
                    text=f"Notes: {review['notes']}",
                    font=("Arial", 9, "italic"),  # Notes font
                    bg=bg_color,
                    fg=fg_color_desc,
                    wraplength=500,
                    justify="left"
                )
                notes_label.pack(anchor="w", pady=(3, 0))
    
    def _build_details_for_repo(self, details_frame, reviews):
        """Build the details view for a repo (lazy loaded).
        
        Args:
            details_frame: The frame to populate with review details
            reviews: List of review dictionaries to display
        """
        self._build_review_details(details_frame, reviews, "white", "#374151", "#6b7280")
    
    def show_cancelled_window(self):
        """Open a separate window to show cancelled/broken reviews.
        
        Creates a modal-style window with:
        - Header showing total cancelled count
        - Scrollable list of cancelled review cards
        - Close button
        
        These reviews have Report Uploaded blank AND Notes populated.
        """
        if self.total_cancelled == 0:
            return
        
        # Create new window
        cancelled_window = tk.Toplevel(self.root)
        cancelled_window.title("Cancelled/Broken Reviews")
        cancelled_window.geometry("400x600")  # Match main window size
        cancelled_window.configure(bg="#f0f0f0")
        
        # Header section
        header_frame = tk.Frame(cancelled_window, bg="#f9fafb", relief="solid", bd=1)
        header_frame.pack(fill="x", padx=20, pady=20)  # Header outer padding
        
        header_inner = tk.Frame(header_frame, bg="#f9fafb")
        header_inner.pack(fill="x", padx=20, pady=15)  # Header inner padding
        
        title_label = tk.Label(
            header_inner,
            text="🚫 Cancelled/Broken Reviews",
            font=("Arial", 16, "bold"), # Title font
            bg="#f9fafb",
            fg="#1f2937"
        )
        title_label.pack(anchor="center")
        
        count_label = tk.Label(
            header_inner,
            text=f"{self.total_cancelled} {'review' if self.total_cancelled == 1 else 'reviews'}",
            font=("Arial", 28, "bold"), # Count font
            bg="#f9fafb",
            fg="#1f2937"
        )
        count_label.pack(anchor="center", pady=(5, 0))  # Top padding only
        
        note_label = tk.Label(
            header_inner,
            text="Not counted in outstanding total",
            font=("Arial", 10), # Note font
            bg="#f9fafb",
            fg="#4b5563"
        )
        note_label.pack(anchor="center", pady=(5, 0))  # Top padding only
        
        # Scrollable content area
        canvas_frame = tk.Frame(cancelled_window, bg="#f0f0f0")
        canvas_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        canvas = tk.Canvas(canvas_frame, bg="#f0f0f0", highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        
        scrollable_frame = tk.Frame(canvas, bg="#f0f0f0")
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")  # Always show scrollbar in popup
        
        # Bind mousewheel
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 100)), "units")  # Scroll speed
        
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Sort cancelled repos by count
        sorted_cancelled = sorted(self.cancelled_data.items(), key=lambda x: len(x[1]), reverse=True)  # Most reviews first
        
        # Create cards for cancelled repos
        for repo, reviews in sorted_cancelled:
            self.create_cancelled_card_in_window(scrollable_frame, repo, reviews)
        
        # Close button at bottom
        close_btn = tk.Button(
            cancelled_window,
            text="Close",
            command=cancelled_window.destroy,
            font=("Arial", 11), # Close button font
            bg="#6b7280",
            fg="white",
            relief="flat",
            cursor="hand2",
            padx=30,
            pady=8
        )
        close_btn.pack(pady=(0, 20))  # Bottom padding
        
        # Cleanup mousewheel binding when window closes to avoid conflicts
        def on_close():
            canvas.unbind_all("<MouseWheel>")
            cancelled_window.destroy()
        
        cancelled_window.protocol("WM_DELETE_WINDOW", on_close)
    
    def create_cancelled_card_in_window(self, parent_frame, repo, reviews):
        """Create a card for cancelled/broken reviews in the popup window.
        
        Args:
            parent_frame: The parent frame to add the card to
            repo: Repository name
            reviews: List of cancelled review dictionaries
        """
        card_config = {
            'bg_color': '#f9fafb',
            'fg_color': '#1f2937',
            'icon': '🚫',
            'count_text_template': '{count} cancelled/broken review{s}',
            'badge_color_fn': None,  # Use grey badge
            'expanded_set': self.expanded_cancelled,
            'expansion_key': f"cancelled_{repo}",
            'build_details_fn': self._build_cancelled_details,
            'details_bg': '#f3f4f6'
        }
        self._create_card(parent_frame, repo, reviews, card_config)
    
    def _build_cancelled_details(self, details_frame, reviews):
        """Build the details view for cancelled reviews (lazy loaded).
        
        Args:
            details_frame: The frame to populate with review details
            reviews: List of cancelled review dictionaries to display
        """
        self._build_review_details(details_frame, reviews, "#f9fafb", "#1f2937", "#4b5563", show_notes=True)
    
    def show_other_window(self):
        """Open a separate window to show other reviews.
        
        Creates a modal-style window with:
        - Header showing total other count
        - Scrollable list of other review cards
        - Close button
        
        These reviews have Report Uploaded blank AND Notes with content (but not cancelled/broken).
        """
        if self.total_other == 0:
            return
        
        # Create new window
        other_window = tk.Toplevel(self.root)
        other_window.title("Other Reviews")
        other_window.geometry("400x600")  # Match main window size
        other_window.configure(bg="#f0f0f0")
        
        # Header section
        header_frame = tk.Frame(other_window, bg="#f9fafb", relief="solid", bd=1)
        header_frame.pack(fill="x", padx=20, pady=20)  # Header outer padding
        
        header_inner = tk.Frame(header_frame, bg="#f9fafb")
        header_inner.pack(fill="x", padx=20, pady=15)  # Header inner padding
        
        title_label = tk.Label(
            header_inner,
            text="📋 Other Reviews",
            font=("Arial", 16, "bold"),
            bg="#f9fafb",
            fg="#1f2937"
        )
        title_label.pack(anchor="center")
        
        count_label = tk.Label(
            header_inner,
            text=f"{self.total_other} {'review' if self.total_other == 1 else 'reviews'}",
            font=("Arial", 28, "bold"),
            bg="#f9fafb",
            fg="#1f2937"
        )   
        count_label.pack(anchor="center", pady=(5, 0))  # Top padding only
        
        note_label = tk.Label(
            header_inner,
            text="Not counted in outstanding total",
            font=("Arial", 10),
            bg="#f9fafb",
            fg="#4b5563"
        )
        note_label.pack(anchor="center", pady=(5, 0))  # Top padding only
        
        # Scrollable content area
        canvas_frame = tk.Frame(other_window, bg="#f0f0f0")
        canvas_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        canvas = tk.Canvas(canvas_frame, bg="#f0f0f0", highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        
        scrollable_frame = tk.Frame(canvas, bg="#f0f0f0")
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")  # Always show scrollbar in popup
        
        # Bind mousewheel
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 100)), "units")
        
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Sort other repos by count
        sorted_other = sorted(self.other_data.items(), key=lambda x: len(x[1]), reverse=True)
        
        # Create cards for other repos
        for repo, reviews in sorted_other:
            self.create_other_card_in_window(scrollable_frame, repo, reviews)
        
        # Close button at bottom
        close_btn = tk.Button(
            other_window,
            text="Close",
            command=other_window.destroy,
            font=("Arial", 11),
            bg="#6b7280",
            fg="white",
            relief="flat",
            cursor="hand2",
            padx=30,
            pady=8
        )
        close_btn.pack(pady=(0, 20))  # Bottom padding
        
        # Cleanup mousewheel binding when window closes to avoid conflicts
        def on_close():
            canvas.unbind_all("<MouseWheel>")
            other_window.destroy()
        
        other_window.protocol("WM_DELETE_WINDOW", on_close)
    
    def create_other_card_in_window(self, parent_frame, repo, reviews):
        """Create a card for other reviews in the popup window.
        
        Args:
            parent_frame: The parent frame to add the card to
            repo: Repository name
            reviews: List of other review dictionaries
        """
        card_config = {
            'bg_color': '#f9fafb',      # Same as cancelled popup cards
            'fg_color': '#1f2937',
            'icon': '📋',
            'count_text_template': '{count} other review{s}',
            'badge_color_fn': None,
            'expanded_set': self.expanded_other,
            'expansion_key': f"other_{repo}",
            'build_details_fn': self._build_other_details,
            'details_bg': '#f3f4f6'     # Same details background as cancelled
        }
        self._create_card(parent_frame, repo, reviews, card_config)
    
    def _build_other_details(self, details_frame, reviews):
        """Build the details view for other reviews (lazy loaded).
        
        Args:
            details_frame: The frame to populate with review details
            reviews: List of other review dictionaries to display
        """
        self._build_review_details(details_frame, reviews, "#f9fafb", "#1f2937", "#4b5563", show_notes=True)

def main():
    root = tk.Tk()
    app = DashboardApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
