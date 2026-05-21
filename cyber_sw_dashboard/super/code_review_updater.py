#!/usr/bin/env python3
"""
Code Review Updater
Downloads Excel from source SharePoint, processes data, uploads to destination SharePoint
"""

import logging
import sys
import tempfile
import os
import shutil
import time
import re
import webbrowser
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, date

# ============================================================================
# CONFIGURATION
# ============================================================================

# Source SharePoint configuration - where we download the original review status file
SOURCE_CONFIG = {
    "site_url": "https://space-us.external.lmco.com/sites/ngi",
    "file_path": "/fa/Software/Software/01-Common-Folder/NGI FSW Formal Review Status.xlsx",
    "output_path": str(Path(__file__).parent / "NGI FSW Formal Review Status.xlsx"),
    "session_file": str(Path(__file__).parent / "sp_session.json"),
}

# Destination SharePoint configuration - where we upload processed data
DEST_CONFIG = {
    "site_url": "https://lmco.sharepoint.us/sites/US-NGI-Cyber-Software",
    "file_path": "/Shared Documents/General/Software Support/Testing New Code Review Tracker.xlsx",
    "session_file": str(Path(__file__).parent / "sp_dest_session.json"),
}

# Target sheets in the destination Excel file for categorizing reviews
TARGET_SHEETS = ["Code", "Design", "Test", "Other"]

LOG_FILE = str(Path(__file__).parent / "code_review_updater.log")

# ============================================================================
# LOGGING SETUP
# ============================================================================

class ColoredFormatter(logging.Formatter):
    """Custom formatter that adds color to log level names for terminal output."""
    
    COLORS = {
        'INFO': '\033[92m',
        'ERROR': '\033[91m',
        'WARNING': '\033[93m',
        'RESET': '\033[0m'
    }
    
    def format(self, record):
        levelname = record.levelname
        if levelname in self.COLORS:
            record.levelname = f"{self.COLORS[levelname]}{levelname}{self.COLORS['RESET']}"
        return super().format(record)

def setup_logging():
    """Configure logging to output to both file and console."""
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    
    file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
    file_handler.setFormatter(logging.Formatter("%(levelname)-8s %(message)s"))
    logger.addHandler(file_handler)
    
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(ColoredFormatter("%(levelname)-8s %(message)s"))
    logger.addHandler(console_handler)

# ============================================================================
# SHAREPOINT AUTHENTICATION & DOWNLOAD
# ============================================================================

def is_login_page(url: str) -> bool:
    """Return True if the URL is a Microsoft login/auth page.
    
    Used to detect if authentication has expired and user needs to re-login.
    
    Args:
        url: The current page URL
        
    Returns:
        bool: True if URL is a Microsoft authentication page
    """
    auth_domains = [
        "login.microsoftonline.com",
        "login.live.com",
        "login.microsoft.com",
        "aadcdn.msftauth.net",
    ]
    return any(d in url.lower() for d in auth_domains)

def check_session_valid(playwright, session_path: Path, site_url: str) -> bool:
    """Check if saved session is still valid.
    
    Attempts to navigate to the site using saved session cookies.
    If redirected to login page, session has expired.
    
    Args:
        playwright: Playwright instance
        session_path: Path to saved session file
        site_url: SharePoint site URL to test
        
    Returns:
        bool: True if session is valid, False otherwise
    """
    if not session_path.exists():
        logging.info(f"No saved session found at: {session_path}")
        return False
    
    logging.info("Checking saved session...")
    browser = playwright.chromium.launch(headless=True)
    try:
        context = browser.new_context(storage_state=str(session_path))
        page = context.new_page()
        page.goto(site_url, wait_until="domcontentloaded", timeout=15000)
        time.sleep(0.5)
        valid = not is_login_page(page.url)
        if valid:
            logging.info("Session is valid.")
        else:
            logging.info("Session expired - re-authentication required.")
        return valid
    except Exception as e:
        error_type = type(e).__name__
        if "TimeoutError" in error_type:
            logging.warning(f"Session check timed out - network may be slow or unavailable")
        else:
            logging.warning(f"Session check failed ({error_type}): {e}")
        logging.warning(f"  Will attempt fresh authentication")
        return False
    finally:
        browser.close()

def authenticate(playwright, session_path: Path, site_url: str) -> bool:
    """Perform interactive login and save session.
    
    Opens a browser window for user to complete Microsoft authentication.
    Session is saved for future use to avoid repeated logins.
    
    Args:
        playwright: Playwright instance
        session_path: Path where session will be saved
        site_url: SharePoint site URL to authenticate to
        
    Returns:
        bool: True if authentication successful, False otherwise
    """
    logging.info("Opening browser for login…")
    print("\n" + "=" * 70)
    print(" ACTION REQUIRED: Browser opening for SharePoint login.")
    print(" 1. Log in and complete MFA")
    print(" 2. Navigate to your SharePoint site if not redirected")
    print(" 3. Press ENTER here after successful login")
    print("=" * 70 + "\n")

    browser = None
    try:
        browser = playwright.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()

        try:
            page.goto(site_url, timeout=15000)
        except Exception as e:
            logging.warning(f"Initial navigation had an issue: {e}")
            logging.info("Browser is open - please navigate to the SharePoint site manually")

        try:
            input("Press ENTER after successful login...")
            logging.info("User confirmed login.")
        except (KeyboardInterrupt, EOFError):
            logging.error("Login cancelled by user.")
            if browser:
                browser.close()
            return False

        try:
            # Verify we're not still on login page
            current_url = page.url
            if is_login_page(current_url):
                logging.error("Still on login page - authentication may not be complete")
                logging.error(f"  Current URL: {current_url}")
                if browser:
                    browser.close()
                return False
            
            context.storage_state(path=str(session_path))
            logging.info(f"Session saved to: {session_path}")
        except Exception as e:
            logging.error(f"Failed to save session: {e}")
            logging.error(f"  This may indicate the login was not successful")
            if browser:
                browser.close()
            return False

        browser.close()
        return True
    
    except Exception as e:
        logging.error(f"Authentication failed: {e}")
        if browser:
            browser.close()
        return False

def download_file_from_sharepoint(playwright, session_path: Path, site_url: str, 
                                  file_path: str, output_path: str) -> bool:
    """Download file from SharePoint using Playwright.
    
    Uses the ?download=1 parameter to trigger direct file download.
    
    Args:
        playwright: Playwright instance
        session_path: Path to saved session file
        site_url: SharePoint site URL
        file_path: Relative path to file on SharePoint
        output_path: Local path where file will be saved
        
    Returns:
        bool: True if download successful, False otherwise
    """
    download_url = f"{site_url.rstrip('/')}{file_path}?download=1"
    
    logging.info(f"Downloading: {download_url}")
    
    browser = playwright.chromium.launch(headless=True)
    page = None
    try:
        context = browser.new_context(storage_state=str(session_path), accept_downloads=True)
        page = context.new_page()
        
        download_info = {"download": None, "error": None}
        response_info = {"status": None, "url": None}
        
        def handle_download(download):
            try:
                logging.info(f"Download started: {download.suggested_filename}")
                download_info["download"] = download
            except Exception as e:
                download_info["error"] = str(e)
        
        def handle_response(response):
            response_info["status"] = response.status
            response_info["url"] = response.url
        
        page.on("download", handle_download)
        page.on("response", handle_response)
        
        navigation_error = None
        try:
            page.goto(download_url, wait_until="commit", timeout=30000)
        except Exception as e:
            navigation_error = e
            error_str = str(e)
            if "Download is starting" in error_str or "download" in error_str.lower():
                logging.info("Navigation interrupted by download (expected)")
            elif "TimeoutError" in type(e).__name__ or "timeout" in error_str.lower():
                logging.warning(f"Navigation timed out - will check if download started")
            else:
                logging.warning(f"Navigation error: {error_str}")
        
        page.wait_for_timeout(2000)
        
        if download_info["error"]:
            logging.error(f"Download handler error: {download_info['error']}")
            logging.error(f"  This indicates a problem capturing the download")
            return False
        
        if not download_info["download"]:
            logging.error("Download did not start")
            logging.error(f"  URL attempted: {download_url}")
            
            if navigation_error:
                logging.error(f"  Navigation error occurred: {str(navigation_error)}")
            
            if response_info["status"]:
                logging.error(f"  HTTP Status: {response_info['status']}")
                if response_info["status"] == 404:
                    logging.error(f"  File not found - check if the path is correct")
                elif response_info["status"] == 403:
                    logging.error(f"  Access denied - check permissions")
                elif response_info["status"] >= 500:
                    logging.error(f"  Server error - SharePoint may be experiencing issues")
            
            try:
                page_title = page.title()
                page_url = page.url
                logging.error(f"  Current page: {page_url}")
                logging.error(f"  Page title: {page_title}")
                
                if "sign" in page_title.lower() or "login" in page_title.lower():
                    logging.error(f"  Redirected to login page - session may have expired")
                elif "error" in page_title.lower():
                    logging.error(f"  Page shows an error - file may not exist or access denied")
            except:
                pass
            
            return False
        
        download = download_info["download"]
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            temp_path = tmp_file.name
        
        try:
            download.save_as(temp_path)
        except Exception as e:
            logging.error(f"Failed to save download: {str(e)}")
            logging.error(f"  This may indicate disk space issues or permission problems")
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return False
        
        if not os.path.exists(temp_path):
            logging.error(f"Download file was not created at: {temp_path}")
            return False
        
        temp_size = os.path.getsize(temp_path)
        if temp_size == 0:
            logging.error(f"Downloaded file is empty (0 bytes)")
            logging.error(f"  This usually means the download failed or returned an error page")
            os.remove(temp_path)
            return False
        elif temp_size < 1024:
            logging.warning(f"Downloaded file is very small ({temp_size} bytes)")
            logging.warning(f"  This may indicate an error page instead of the actual file")
        
        normalized_path = os.path.normpath(output_path)
        
        if os.path.exists(normalized_path):
            logging.info(f"Deleting existing file: {normalized_path}")
            os.remove(normalized_path)
        
        try:
            with open(temp_path, 'rb') as src:
                file_data = src.read()
                with open(normalized_path, 'wb') as dst:
                    dst.write(file_data)
        except Exception as e:
            logging.error(f"Failed to copy file to destination: {str(e)}")
            logging.error(f"  Source: {temp_path}")
            logging.error(f"  Destination: {normalized_path}")
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return False
        
        os.remove(temp_path)
        
        size_kb = os.path.getsize(normalized_path) / 1024
        logging.info(f"Saved to: {normalized_path} ({size_kb:.1f} KB)")
        
        if size_kb < 10:
            logging.warning(f"File size is unusually small for an Excel file")
            logging.warning(f"  Verify the file content is correct")
        
        return True
        
    except Exception as e:
        error_type = type(e).__name__
        logging.error(f"Unexpected error during download ({error_type}): {str(e)}")
        logging.error(f"  URL: {download_url}")
        
        import traceback
        logging.error(f"  Stack trace:")
        for line in traceback.format_exc().split('\n'):
            if line.strip():
                logging.error(f"    {line}")
        
        return False
    finally:
        if browser:
            browser.close()

def upload_file_to_sharepoint(playwright, session_path: Path, site_url: str, 
                               file_path: str, local_file: str) -> bool:
    """Upload file to SharePoint using REST API.
    
    Uses SharePoint REST API with form digest for authentication.
    Overwrites existing file at the destination path.
    
    Args:
        playwright: Playwright instance
        session_path: Path to saved session file
        site_url: SharePoint site URL
        file_path: Relative path where file will be uploaded on SharePoint
        local_file: Local path of file to upload
        
    Returns:
        bool: True if upload successful, False otherwise
    """
    logging.info("Uploading to SharePoint...")
    
    if not os.path.exists(local_file):
        logging.error(f"Local file not found: {local_file}")
        return False
    
    try:
        with open(local_file, 'rb') as f:
            file_content = f.read()
    except Exception as e:
        logging.error(f"Failed to read local file: {e}")
        logging.error(f"  File: {local_file}")
        return False
    
    file_size = len(file_content)
    logging.info(f"File size: {file_size / 1024:.2f} KB")
    
    if file_size == 0:
        logging.error(f"Cannot upload empty file")
        return False
    
    file_name = os.path.basename(file_path)
    folder_path = os.path.dirname(file_path)
    
    from urllib.parse import urlparse
    parsed_url = urlparse(site_url)
    site_relative_path = parsed_url.path.rstrip('/')
    
    full_folder_path = f"{site_relative_path}{folder_path}" if site_relative_path else folder_path
    
    upload_endpoint = f"{site_url}/_api/web/GetFolderByServerRelativeUrl('{full_folder_path}')/Files/add(url='{file_name}',overwrite=true)"
    
    logging.info(f"Upload endpoint: {upload_endpoint}")
    
    browser = None
    try:
        browser = playwright.chromium.launch(headless=True)
        context = browser.new_context(storage_state=str(session_path))
        page = context.new_page()
        
        digest_url = f"{site_url}/_api/contextinfo"
        
        try:
            response = page.request.post(
                digest_url,
                headers={
                    "Accept": "application/json;odata=verbose",
                    "Content-Type": "application/json;odata=verbose",
                }
            )
        except Exception as e:
            logging.error(f"Failed to request form digest: {e}")
            logging.error(f"  Digest URL: {digest_url}")
            logging.error(f"  This usually indicates session expiration or network issues")
            return False
        
        if response.status != 200:
            logging.error(f"Failed to get request digest")
            logging.error(f"  HTTP Status: {response.status}")
            logging.error(f"  URL: {digest_url}")
            
            if response.status == 401:
                logging.error(f"  Unauthorized - session may have expired")
            elif response.status == 403:
                logging.error(f"  Forbidden - insufficient permissions")
            elif response.status == 404:
                logging.error(f"  Not found - check site URL")
            
            try:
                error_text = response.text()
                if error_text:
                    logging.error(f"  Response: {error_text[:200]}")
            except:
                pass
            
            return False
        
        try:
            digest_data = response.json()
            form_digest = digest_data["d"]["GetContextWebInformation"]["FormDigestValue"]
            logging.info("Form digest obtained successfully")
        except Exception as e:
            logging.error(f"Failed to parse digest response: {e}")
            logging.error(f"  This indicates an unexpected response format")
            return False
        
        try:
            response = page.request.post(
                upload_endpoint,
                headers={
                    "Accept": "application/json;odata=verbose",
                    "X-RequestDigest": form_digest,
                    "Content-Type": "application/octet-stream",
                },
                data=file_content
            )
        except Exception as e:
            logging.error(f"Failed to upload file: {e}")
            logging.error(f"  Upload endpoint: {upload_endpoint}")
            logging.error(f"  This may indicate network timeout or connection issues")
            return False
        
        if response.status in (200, 201):
            logging.info("File uploaded successfully!")
            return True
        else:
            logging.error(f"Upload failed")
            logging.error(f"  HTTP Status: {response.status}")
            logging.error(f"  Upload endpoint: {upload_endpoint}")
            
            if response.status == 401:
                logging.error(f"  Unauthorized - form digest may have expired")
            elif response.status == 403:
                logging.error(f"  Forbidden - check write permissions to the folder")
            elif response.status == 404:
                logging.error(f"  Folder not found - check the folder path")
            elif response.status == 409:
                logging.error(f"  Conflict - file may be locked or checked out by another user")
            elif response.status >= 500:
                logging.error(f"  Server error - SharePoint may be experiencing issues")
            
            try:
                error_text = response.text()
                if error_text:
                    logging.error(f"  Response: {error_text[:500]}")
            except:
                pass
            
            return False
            
    except Exception as e:
        error_type = type(e).__name__
        logging.error(f"Unexpected upload error ({error_type}): {str(e)}")
        logging.error(f"  Target: {upload_endpoint}")
        
        import traceback
        logging.error(f"  Stack trace:")
        for line in traceback.format_exc().split('\n'):
            if line.strip():
                logging.error(f"    {line}")
        
        return False
    finally:
        if browser:
            browser.close()

# ============================================================================
# DATA PROCESSING & TRANSFORMATION
# ============================================================================

def map_column_d(col_c_value):
    """Map column C value to repository name.
    
    Extracts repository identifier from description field based on keywords.
    Supports multiple variants: MSV, KV, COMMON, NGICON, HSP, HS.
    
    Args:
        col_c_value: Value from column C (description field)
        
    Returns:
        str: Standardized repository name or 'ENTER REPO NAME' if no match
    """
    val = str(col_c_value or "").strip()
    val_upper = val.upper()
    
    if "MSV" in val_upper and "EM" in val_upper:
        return "MSV-EM"
    if "MSV" in val_upper and "FSW" in val_upper:
        return "MSV-FSW"
    if "MSV" in val_upper and "NGC" in val_upper:
        return "MSV-NGC"
    if "MSV" in val_upper and "PTS" in val_upper:
        return "MSV-PTS"
    
    if "KV" in val_upper and "EM" in val_upper:
        return "KV-EM"
    if "KV" in val_upper and "FSW" in val_upper:
        return "KV-FSW"
    if "KV" in val_upper and "NGC" in val_upper:
        return "KV-NGC"
    if "KV" in val_upper and "PTS" in val_upper:
        return "KV-PTS"
    
    if "COMMON" in val_upper and "ALGO" in val_upper:
        return "COMMON-ALGO"
    if "COMMON" in val_upper and "FSW" in val_upper:
        return "COMMON-FSW"
    
    if "NGICON" in val_upper:
        return "NGICON"
    if "HSP" in val_upper:
        return "HSP"
    if "HAWKEYE" in val_upper or "HS" in val_upper:
        return "HS"
    
    if val_upper == "FSW":
        return "COMMON-FSW"
    
    return "ENTER REPO NAME"

def extract_date_only(value):
    """Extract just the date from a value that may contain date and time.
    
    Handles multiple date formats: YYYY-MM-DD, M/D/YYYY, D-Mon-YYYY.
    Removes time portion if present.
    
    Args:
        value: Date value (datetime object, date object, or string)
        
    Returns:
        str: Date string without time, or None if no date found
    """
    if value is None:
        return None
    
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    
    value_str = str(value).strip()
    if not value_str:
        return None
    
    match = re.match(r'(\d{4}-\d{2}-\d{2})', value_str)
    if match:
        return match.group(1)
    
    match = re.match(r'(\d{1,2}/\d{1,2}/\d{4})', value_str)
    if match:
        return match.group(1)
    
    match = re.match(r'(\d{1,2}-[A-Za-z]{3}-\d{4})', value_str)
    if match:
        return match.group(1)
    
    return value_str

def classify(row):
    """Determine which target sheet this row belongs to.
    
    Classification rules:
    - 'test' in column C -> ignore (None)
    - 'code' in column D -> Code sheet
    - 'design' in column D -> Design sheet
    - 'test' in column D -> Test sheet
    - Otherwise -> Other sheet
    
    Args:
        row: Excel row with cell values
        
    Returns:
        str: Target sheet name ('Code', 'Design', 'Test', 'Other') or None to ignore
    """
    col_c = str(row[2].value or "").lower()
    col_d = str(row[3].value or "").lower()

    if "test" in col_c:
        return None
    if "code" in col_d:
        return "Code"
    if "design" in col_d:
        return "Design"
    if "test" in col_d:
        return "Test"
    return "Other"

def process_and_categorize(source_path: str, dest_path: str) -> None:
    """Process Excel data and categorize into different sheets.
    
    Reads source Excel file and:
    1. Classifies each row into Code/Design/Test/Other
    2. Maps repository names from descriptions
    3. Extracts date-only values
    4. Writes categorized data to destination file
    
    Args:
        source_path: Path to source Excel file
        dest_path: Path to output categorized Excel file
    """
    import openpyxl
    
    logging.info(f"Reading from: {source_path}")
    
    src_wb = openpyxl.load_workbook(source_path, data_only=True)
    src_ws = src_wb.active

    dst_wb = openpyxl.Workbook()
    dst_wb.active.title = "Other"
    
    for name in TARGET_SHEETS:
        if name not in dst_wb.sheetnames:
            dst_wb.create_sheet(name)

    counts = {sheet: 0 for sheet in TARGET_SHEETS}
    ignored = 0

    rows = list(src_ws.iter_rows())

    # Skip header row if first cell is not numeric (likely a header)
    start = 0
    if rows and not isinstance(rows[0][0].value, (int, float)):
        start = 1

    # Process each row and categorize
    for row in rows[start:]:
        # Skip empty rows
        if all(cell.value is None for cell in row):
            continue

        # Determine target sheet for this row
        target = classify(row)
        if target is None:
            ignored += 1
            continue

        col_a = row[0].value
        if col_a is not None:
            try:
                col_a = int(col_a)
            except (ValueError, TypeError):
                pass

        dst_ws = dst_wb[target]

        # Extract source data (columns A-I)
        src = [col_a] + [cell.value for cell in row[1:9]]
        
        # Build destination row with remapped columns
        new_row = [None] * 10
        new_row[1] = src[0]  # Review number: A → B
        new_row[3] = map_column_d(src[2])  # Repository: C → D (mapped)
        new_row[4] = src[8]  # Description: I → E
        new_row[2] = extract_date_only(src[6])  # Date: G → C (date only)
        
        dst_ws.append(new_row)
        counts[target] += 1

    dst_wb.save(dest_path)

    logging.info(f"Rows routed:")
    for sheet, n in counts.items():
        logging.info(f"  {sheet:8s}: {n}")
    logging.info(f"  ignored  : {ignored}")

# ============================================================================
# EXCEL TABLE OPERATIONS
# ============================================================================

def append_to_excel_tables(file_path: str, sheet_data: dict) -> None:
    """Append data to existing Excel tables in the file.
    
    For each sheet with data:
    1. Finds existing table and expands its range
    2. Appends new rows after existing data
    3. Preserves existing data and formatting
    
    Args:
        file_path: Path to Excel file with tables
        sheet_data: Dictionary mapping sheet names to lists of row data
    """
    import openpyxl
    from openpyxl.utils import range_boundaries, get_column_letter
    
    wb = openpyxl.load_workbook(file_path)
    
    for sheet_name, data in sheet_data.items():
        if not data:
            continue
            
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Find and expand existing table if present
            if ws.tables:
                table_name = list(ws.tables.keys())[0]
                table = ws.tables[table_name]
                table_ref = table.ref
                logging.info(f"Found table '{table_name}' with range {table_ref}")
                
                # Calculate new table range to include appended rows
                min_col, min_row, max_col, max_row = range_boundaries(table_ref)
                start_row = max_row + 1
                
                new_max_row = max_row + len(data)
                new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
                table.ref = new_ref
                logging.info(f"Expanded table to {new_ref}")
            else:
                last_row = ws.max_row
                start_row = last_row + 1 if last_row > 0 else 1
        else:
            ws = wb.create_sheet(sheet_name)
            start_row = 1
        
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        logging.info(f"Appended {len(data)} rows to '{sheet_name}' starting at row {start_row}")
    
    wb.save(file_path)

def open_sharepoint_urls():
    """Open SharePoint Excel files in browser after new data is added.
    
    Opens both the Code Review Tracker and Testing New Code Review Tracker
    in the default web browser.
    """
    urls = [
        "https://lmco.sharepoint.us/:x:/r/sites/US-NGI-Cyber-Software/_layouts/15/Doc.aspx?sourcedoc=%7B9c09dec9-da2f-44e0-9ef9-bf9f0179d308%7D&action=edit&activeCell=%27Code%27!F33&wdinitialsession=3c2a7e5c-9fbf-12e3-655e-6575b410cf22&wdrldsc=2&wdrldc=1&wdrldr=AccessTokenExpiredWarningAccessDenied%2CRefreshingEx",
        "https://lmco.sharepoint.us/:x:/r/sites/US-NGI-Cyber-Software/_layouts/15/Doc.aspx?sourcedoc=%7BFAD72989-75DE-4045-A95A-03E74CD9B673%7D&file=Testing%20New%20Code%20Review%20Tracker.xlsx&action=default&mobileredirect=true"
    ]
    
    logging.info("\nOpening SharePoint files in browser...")
    for url in urls:
        try:
            webbrowser.open(url)
            logging.info(f"✓ Opened URL in browser")
        except Exception as e:
            logging.warning(f"Failed to open URL: {e}")

# ============================================================================
# MAIN WORKFLOW
# ============================================================================

def main():
    """Main workflow function.
    
    Workflow steps:
    1. Download files from SharePoint (source and destination)
    2. Process and categorize source data
    3. Check for duplicates against destination
    4. Append only new rows to destination
    5. Upload updated destination file to SharePoint
    """
    setup_logging()
    logging.info("=" * 70)
    logging.info("Code Review Updater - Starting Workflow")
    logging.info("=" * 70)

    import argparse
    parser = argparse.ArgumentParser(description="Code Review Updater Workflow")
    parser.add_argument('--skip-download', action='store_true', 
                       help='Skip download step and use existing file')
    args = parser.parse_args()

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        logging.error("Playwright is not installed.")
        logging.error("Fix: pip install playwright openpyxl")
        logging.error("     playwright install chromium")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        logging.error("openpyxl is not installed.")
        logging.error("Fix: pip install openpyxl")
        sys.exit(1)

    source_file = SOURCE_CONFIG["output_path"]
    dest_file = Path(__file__).parent / "destination_current.xlsx"

    with sync_playwright() as playwright:
        # ==================================================================
        # STEP 1: Download files from SharePoint
        # Downloads both source (original data) and destination (current state)
        # Supports --skip-download flag to use existing source file
        # ==================================================================
        if args.skip_download:
            logging.info("\n" + "=" * 70)
            logging.info("STEP 1: SKIPPED - Using existing file")
            logging.info("=" * 70)
            
            if not Path(source_file).exists():
                logging.error(f"File not found: {source_file}")
                sys.exit(1)
            
            size_kb = Path(source_file).stat().st_size / 1024
            logging.info(f"Using existing file: {source_file} ({size_kb:.1f} KB)")
            
            # Download destination for comparison
            dest_session_path = Path(DEST_CONFIG["session_file"])
            session_ok = check_session_valid(playwright, dest_session_path, DEST_CONFIG["site_url"])
            if not session_ok:
                success = authenticate(playwright, dest_session_path, DEST_CONFIG["site_url"])
                if not success:
                    sys.exit(1)
            
            success = download_file_from_sharepoint(
                playwright, dest_session_path, 
                DEST_CONFIG["site_url"], DEST_CONFIG["file_path"], 
                str(dest_file)
            )
            if not success:
                sys.exit(1)
            
            # Save a copy for the dashboard to read
            logging.info(f"Saving copy for dashboard: {dest_file}")
        else:
            logging.info("\n" + "=" * 70)
            logging.info("STEP 1: Downloading files from SharePoint (PARALLEL)")
            logging.info("=" * 70)
            
            source_session_path = Path(SOURCE_CONFIG["session_file"])
            dest_session_path = Path(DEST_CONFIG["session_file"])
            
            # Validate sessions
            session_ok = check_session_valid(playwright, source_session_path, SOURCE_CONFIG["site_url"])
            if not session_ok:
                success = authenticate(playwright, source_session_path, SOURCE_CONFIG["site_url"])
                if not success:
                    sys.exit(1)
            
            session_ok = check_session_valid(playwright, dest_session_path, DEST_CONFIG["site_url"])
            if not session_ok:
                success = authenticate(playwright, dest_session_path, DEST_CONFIG["site_url"])
                if not success:
                    sys.exit(1)
            
            # Parallel downloads
            def download_source():
                with sync_playwright() as pw:
                    success = download_file_from_sharepoint(
                        pw, source_session_path,
                        SOURCE_CONFIG["site_url"], SOURCE_CONFIG["file_path"],
                        source_file
                    )
                    return ("source", success, source_file)
            
            def download_destination():
                with sync_playwright() as pw:
                    success = download_file_from_sharepoint(
                        pw, dest_session_path,
                        DEST_CONFIG["site_url"], DEST_CONFIG["file_path"],
                        str(dest_file)
                    )
                    return ("destination", success, str(dest_file))
            
            with ThreadPoolExecutor(max_workers=2) as executor:
                futures = {
                    executor.submit(download_source): "source",
                    executor.submit(download_destination): "destination"
                }
                
                for future in as_completed(futures):
                    file_type, success, filepath = future.result()
                    if success:
                        logging.info(f"✓ Downloaded {file_type} to: {filepath}")
                    else:
                        logging.error(f"✗ Failed to download {file_type}")
                        sys.exit(1)
            
            # Confirm destination file is ready for dashboard
            logging.info(f"\n✓ Dashboard file ready: {dest_file}")
        
        # ==================================================================
        # STEP 2: Process and categorize data
        # Reads source file, classifies reviews by type, maps repositories
        # ==================================================================
        logging.info("\n" + "=" * 70)
        logging.info("STEP 2: Processing and categorizing data")
        logging.info("=" * 70)
        
        processed_file = tempfile.mktemp(suffix='.xlsx')
        
        try:
            process_and_categorize(source_file, processed_file)
            logging.info(f"✓ Processed data saved to: {processed_file}")
        except Exception as e:
            logging.error(f"Failed to process data: {e}")
            sys.exit(1)
        
        # ==================================================================
        # STEP 3: Check for duplicates and upload
        # Compares processed data against destination to find new reviews
        # Only uploads rows with review numbers higher than current max
        # ==================================================================
        logging.info("\n" + "=" * 70)
        logging.info("STEP 3: Checking for duplicates and uploading")
        logging.info("=" * 70)
        
        # Read processed data from temporary file
        wb = openpyxl.load_workbook(processed_file)
        
        sheet_data = {}
        for sheet_name in TARGET_SHEETS:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = []
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        data.append(list(row))
                if data:
                    sheet_data[sheet_name] = data
        
        wb.close()
        
        # Check for duplicates by comparing review numbers
        dest_wb = openpyxl.load_workbook(dest_file, read_only=True, data_only=True)
        
        filtered_data = {}
        for sheet_name, source_data in sheet_data.items():
            if sheet_name in dest_wb.sheetnames:
                dest_ws = dest_wb[sheet_name]
                
                # Find the highest review number in destination sheet
                max_review_num = 0
                dest_rows = list(dest_ws.iter_rows(values_only=True))
                
                for row in dest_rows:
                    if row and len(row) > 1 and row[1] is not None:
                        try:
                            review_num = int(row[1])
                            if review_num > max_review_num:
                                max_review_num = review_num
                        except (ValueError, TypeError):
                            pass
                
                logging.info(f"  Sheet '{sheet_name}': Max Review # = {max_review_num}")
                
                # Filter source data to only include new review numbers
                new_rows = []
                skipped_count = 0
                for source_row in source_data:
                    if source_row and len(source_row) > 1 and source_row[1] is not None:
                        try:
                            source_review_num = int(source_row[1])
                            if source_review_num > max_review_num:
                                new_rows.append(source_row)
                            else:
                                skipped_count += 1
                        except (ValueError, TypeError):
                            skipped_count += 1
                    else:
                        skipped_count += 1
                
                filtered_data[sheet_name] = new_rows
                logging.info(f"  Sheet '{sheet_name}': {len(new_rows)} new rows to upload")
            else:
                filtered_data[sheet_name] = source_data
        
        dest_wb.close()
        
        # Check if there's any new data to upload
        total_rows = sum(len(data) for data in filtered_data.values())
        if total_rows == 0:
            logging.info("\n" + "=" * 70)
            logging.info("✓ NO NEW DATA - All rows already exist!")
            logging.info("=" * 70)
            os.remove(processed_file)
            return
        
        logging.info(f"\nUploading {total_rows} new rows...")
        
        # Append to destination file in a temporary copy
        temp_upload = tempfile.mktemp(suffix='.xlsx')
        shutil.copy(dest_file, temp_upload)
        
        append_to_excel_tables(temp_upload, filtered_data)
        
        # Upload modified file back to SharePoint
        dest_session_path = Path(DEST_CONFIG["session_file"])
        success = upload_file_to_sharepoint(
            playwright, dest_session_path,
            DEST_CONFIG["site_url"], DEST_CONFIG["file_path"],
            temp_upload
        )
        
        if success:
            logging.info("✓ Successfully uploaded all data to SharePoint!")
            
            # Update local destination_current.xlsx with the uploaded data
            shutil.copy(temp_upload, dest_file)
            logging.info(f"✓ Updated local file: {dest_file}")
            
            # Open SharePoint files in browser since new data was added
            open_sharepoint_urls()
        else:
            logging.error("✗ Upload failed.")
            os.remove(processed_file)
            os.remove(temp_upload)
            sys.exit(1)
        
        # Cleanup temporary files
        os.remove(processed_file)
        os.remove(temp_upload)

    logging.info("\n" + "=" * 70)
    logging.info("✓ WORKFLOW COMPLETED SUCCESSFULLY")
    logging.info("=" * 70)

if __name__ == "__main__":
    main()
